from annual_budget.api.phase_sheet import get_consolidated_report
from annual_budget.api.actuals import get_actuals_from_erp_prod, get_actuals_from_erp_month_wise
import frappe,io
from frappe import _
import frappe
import tempfile
import xlsxwriter
from frappe.utils import nowdate

# * ==============================================================   Budget Export Face Sheet =====================================================================================
@frappe.whitelist()
def export_phase_sheet_excel(
    financial_year=None,
    units=None,
    cost_center=None,
    location_code=None
):

    import io
    import frappe
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    data = get_consolidated_report(
        financial_year=financial_year,
        units=units,
        cost_center=cost_center,
        location_code=location_code
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Budget Summary"

    center = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    bold = Font(bold=True)
    white_bold = Font(bold=True, color="FFFFFF")

    header_fill = PatternFill("solid", fgColor="5D6D7E")
    head_fill = PatternFill("solid", fgColor="D6DBDF")
    subhead_fill = PatternFill("solid", fgColor="F2F3F4")
    subtotal_fill = PatternFill("solid", fgColor="EBF5FB")
    head_total_fill = PatternFill("solid", fgColor="D4E6F1")
    grand_fill = PatternFill("solid", fgColor="A9CCE3")

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)


    def style_row(row, fill=None, font=None, header=False):

        for col in range(2,22):

            cell = ws.cell(row=row,column=col)
            cell.border = border

            if header:
                cell.alignment = center
            else:
                if col == 2:
                    cell.alignment = center
                elif col in (3,4):
                    cell.alignment = left_align
                else:
                    cell.alignment = right_align

            if fill:
                cell.fill = fill

            if font:
                cell.font = font


    def format_numeric(row):
        for col in range(5,22):
            ws.cell(row=row,column=col).number_format = "#,##0.00"


    def build_formula(col, rows):
        return "=" + "+".join([f"{col}{r}" for r in rows]) if rows else 0


    def to_roman(num):

        val=[1000,900,500,400,100,90,50,40,10,9,5,4,1]
        syb=["M","CM","D","CD","C","XC","L","XL","X","IX","V","IV","I"]

        roman=""
        i=0

        while num>0:
            for _ in range(num//val[i]):
                roman+=syb[i]
                num-=val[i]
            i+=1

        return roman


    # TITLE
    ws.append(["","Azim Premji Foundation"])
    ws.merge_cells("B1:U1")
    ws["B1"].font=Font(size=14,bold=True)
    ws["B1"].alignment=left_align

    ws.append(["",f"Budget for the Financial Year {financial_year or ''}"])
    ws.merge_cells("B2:U2")
    ws["B2"].font=Font(size=12,bold=True)

    ws.append([])


    # HEADER
    ws.append([
        "", "Sl #","HEAD OF EXPENSE","TYPE OF EXPENSE",
        "QUARTER I","","",
        "QUARTER II","","",
        "QUARTER III","","",
        "QUARTER IV","","",
        "QTR-1","QTR-2","QTR-3","QTR-4",
        f"YEAR {financial_year}"
    ])

    r1=ws.max_row

    ws.merge_cells(start_row=r1,start_column=5,end_row=r1,end_column=7)
    ws.merge_cells(start_row=r1,start_column=8,end_row=r1,end_column=10)
    ws.merge_cells(start_row=r1,start_column=11,end_row=r1,end_column=13)
    ws.merge_cells(start_row=r1,start_column=14,end_row=r1,end_column=16)

    ws.append([
        "", "Sl #","HEAD OF EXPENSE","TYPE OF EXPENSE",
        "Apr","May","Jun",
        "Jul","Aug","Sep",
        "Oct","Nov","Dec",
        "Jan","Feb","Mar",
        "QTR-1","QTR-2","QTR-3","QTR-4",
        f"YEAR {financial_year}"
    ])

    r2=ws.max_row

    ws.merge_cells(start_row=r1,start_column=2,end_row=r2,end_column=2)
    ws.merge_cells(start_row=r1,start_column=3,end_row=r2,end_column=3)
    ws.merge_cells(start_row=r1,start_column=4,end_row=r2,end_column=4)

    for col in range(17,22):
        ws.merge_cells(start_row=r1,start_column=col,end_row=r2,end_column=col)

    style_row(r1,header_fill,white_bold,True)
    style_row(r2,header_fill,white_bold,True)

    ws.freeze_panes="E6"


    head_total_rows=[]
    head_counter=0


    for head in data:

        head_counter+=1
        alpha_index=chr(64+head_counter)
        head_name=(head.get("name") or "").strip().upper()


        # COVID SUPPORT
        if head_name=="COVID SUPPORT":

            ws.append([])

            item=head["items"][0]
            r=ws.max_row+1

            ws.append([
                "",alpha_index,
                head["name"],
                item["name"],
                *item["q1"],*item["q2"],*item["q3"],*item["q4"],
                f"=SUM(E{r}:G{r})",
                f"=SUM(H{r}:J{r})",
                f"=SUM(K{r}:M{r})",
                f"=SUM(N{r}:P{r})",
                f"=SUM(Q{r}:T{r})"
            ])

            style_row(ws.max_row)
            format_numeric(ws.max_row)
            ws.cell(row=ws.max_row,column=3).font=bold

            continue


        ws.append(["",alpha_index,head["name"]])
        r=ws.max_row

        ws.merge_cells(start_row=r,start_column=3,end_row=r,end_column=21)
        style_row(r,head_fill,bold)

        if head_name=="OPERATING EXPENSES":
            ws.append([])

        sub_total_rows=[]
        direct_item_rows=[]


        for item in head.get("items",[]):

            r=ws.max_row+1
            sub_val=item.get("sub_head_of_expense")
            head_display=sub_val.strip() if sub_val else ""

            ws.append([
                "", "", head_display,
                item["name"],
                *item["q1"],*item["q2"],*item["q3"],*item["q4"],
                f"=SUM(E{r}:G{r})",
                f"=SUM(H{r}:J{r})",
                f"=SUM(K{r}:M{r})",
                f"=SUM(N{r}:P{r})",
                f"=SUM(Q{r}:T{r})"
            ])

            style_row(ws.max_row)
            format_numeric(ws.max_row)
            direct_item_rows.append(ws.max_row)


        sub_counter=1

        for sub in head.get("sub_heads",[]):

            roman_index=to_roman(sub_counter)

            ws.append(["",roman_index,sub["name"]])
            r=ws.max_row

            ws.merge_cells(start_row=r,start_column=3,end_row=r,end_column=21)
            style_row(r,subhead_fill,bold)

            sub_item_rows=[]

            for item in sub.get("items",[]):

                r=ws.max_row+1
                item_sub=item.get("sub_head_of_expense")
                sub_name=sub.get("name")

                head_display=""

                if item_sub and str(item_sub).strip():

                    cleaned=str(item_sub).strip()

                    if cleaned.lower()!=str(sub_name).strip().lower():
                        head_display=cleaned

                ws.append([
                    "", "", head_display,
                    item["name"],
                    *item["q1"],*item["q2"],*item["q3"],*item["q4"],
                    f"=SUM(E{r}:G{r})",
                    f"=SUM(H{r}:J{r})",
                    f"=SUM(K{r}:M{r})",
                    f"=SUM(N{r}:P{r})",
                    f"=SUM(Q{r}:T{r})"
                ])

                style_row(ws.max_row)
                format_numeric(ws.max_row)
                sub_item_rows.append(ws.max_row)


            if sub_item_rows:

                ws.append([
                    "", "", "",
                    f"TOTAL - {sub['name']}",
                    *[build_formula(c,sub_item_rows) for c in list("EFGHIJKLMNOPQRSTU")]
                ])

                style_row(ws.max_row,subtotal_fill,bold)
                format_numeric(ws.max_row)
                sub_total_rows.append(ws.max_row)

            sub_counter+=1


        total_rows=sub_total_rows if sub_total_rows else direct_item_rows

        if total_rows:

            ws.append([
                "", "", "",
                f"TOTAL - {head['name']}",
                *[build_formula(c,total_rows) for c in list("EFGHIJKLMNOPQRSTU")]
            ])

            style_row(ws.max_row,head_total_fill,bold)
            format_numeric(ws.max_row)
            head_total_rows.append(ws.max_row)

            if head_name=="OPERATING EXPENSES":
                ws.append([])


    # CLEANUP EXTRA BLANK ROW
    for r in range(ws.max_row,1,-1):

        head_val=ws.cell(r,3).value
        prev_val=ws.cell(r-1,3).value

        if head_val=="COVID SUPPORT" and prev_val is None:
            ws.delete_rows(r-1)
            break


    # GRAND TOTAL
    if head_total_rows:

        ws.append([
            "", "", "",
            "GRAND TOTAL",
            *[build_formula(c,head_total_rows) for c in list("EFGHIJKLMNOPQRSTU")]
        ])

        style_row(ws.max_row,grand_fill,bold)
        format_numeric(ws.max_row)


    # COLUMN WIDTH
    for col in range(1,ws.max_column+1):

        letter=get_column_letter(col)
        max_len=0

        for row in range(1,ws.max_row+1):

            val=ws.cell(row=row,column=col).value

            if val:
                val=str(val)

                if val.startswith("="):
                    val="999,999,999.00"

                max_len=max(max_len,len(val))

        ws.column_dimensions[letter].width=min(max(max_len+3,10),50)


    stream=io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    frappe.response["filename"]=f"Budget Summary_{financial_year}.xlsx"
    frappe.response["filecontent"]=stream.getvalue()
    frappe.response["type"]="binary"

# * ==============================================================  Import Template Export  =====================================================================================
# @frappe.whitelist()
# def download_finance_budget_import_template(user):

#     if not user:
#         frappe.throw(_("User is required"))

#     # Get Finance User Access
#     doc_name = frappe.db.get_value(
#         "Finance user access",
#         {"user": user},
#         "name"
#     )

#     if not doc_name:
#         frappe.throw(_("No Finance User Access found"))

#     access_doc = frappe.get_doc("Finance user access", doc_name)

#     # Get Import Template ID from Finance User Access
#     if not access_doc.import_template_id:
#         frappe.throw(_("Import Template not linked in Finance User Access"))

#     import_template = frappe.get_doc(
#         "Import Templates",
#         access_doc.import_template_id
#     )

#     template_items = import_template.import_template_item_list

#     if not template_items:
#         frappe.throw(_("No Import Template Items found"))

#     financial_year = frappe.db.get_single_value(
#         "Master Settings",
#         "current_financial_year"
#     )

#     headers = [
#         "Entity / Unit",
#         "Entity / Unit Description",
#         "Cost Center",
#         "Cost Center(Original)",
#         "Cost Center Description",
#         "Location code",
#         "Location code(Original)",
#         "Function / Sub Unit / Division",
#         "State",
#         "Financial year",
#         "Uploaded By",
#         "Type of expense ID (Budget Amounts)",
#         "Head of expense (Budget Amounts)",
#         "Sub head of expense (Budget Amounts)",
#         "Type of expense (Budget Amounts)",
#         "April (Budget Amounts)",
#         "May (Budget Amounts)",
#         "June (Budget Amounts)",
#         "July (Budget Amounts)",
#         "August (Budget Amounts)",
#         "September (Budget Amounts)",
#         "October (Budget Amounts)",
#         "November (Budget Amounts)",
#         "December (Budget Amounts)",
#         "January (Budget Amounts)",
#         "February (Budget Amounts)",
#         "March (Budget Amounts)",
#         "Quarter 1 Total Amount (Budget Amounts)",
#         "Quarter 2 Total Amount (Budget Amounts)",
#         "Quarter 3 Total Amount (Budget Amounts)",
#         "Quarter 4 Total Amount (Budget Amounts)",
#         "Year Total Amount (Budget Amounts)"
#     ]

#     from openpyxl import Workbook
#     from openpyxl.styles import Font, Protection
#     from io import BytesIO
#     import datetime

#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Finance Budget Import"

#     ws.append(headers)

#     for cell in ws[1]:
#         cell.font = Font(bold=True)

#     ws.freeze_panes = "A2"

#     row_index = 2

#     for mapping in access_doc.mapping:

#         first_row = True

#         for item in template_items:

#             if first_row:
#                 parent_values = [
#                     mapping.unit,
#                     mapping.unit_description,
#                     mapping.cost_center,
#                     mapping.cost_center_erp,
#                     mapping.cost_center_description,
#                     mapping.location_code,
#                     mapping.location_code_erp,
#                     mapping.location_description,
#                     mapping.state,
#                     financial_year,
#                     user
#                 ]
#                 first_row = False
#             else:
#                 parent_values = [""] * 11

#             ws.append(parent_values + [
#                 item.type_of_expense_id,
#                 item.head_of_expense,
#                 item.sub_head_of_expense,
#                 item.type_of_expense,
#                 0.00, 0.00, 0.00,
#                 0.00, 0.00, 0.00,
#                 0.00, 0.00, 0.00,
#                 0.00, 0.00, 0.00,
#                 0.00, 0.00, 0.00, 0.00, 0.00
#             ])

#             # Quarter formulas
#             ws[f"AB{row_index}"] = f"=SUM(P{row_index}:R{row_index})"
#             ws[f"AC{row_index}"] = f"=SUM(S{row_index}:U{row_index})"
#             ws[f"AD{row_index}"] = f"=SUM(V{row_index}:X{row_index})"
#             ws[f"AE{row_index}"] = f"=SUM(Y{row_index}:AA{row_index})"

#             # Year total = Sum of quarters
#             ws[f"AF{row_index}"] = f"=SUM(AB{row_index}:AE{row_index})"

#             # Format numbers
#             for col in range(16, 33):
#                 ws.cell(row=row_index, column=col).number_format = '0.00'

#             row_index += 1

#     # Lock everything
#     for row in ws.iter_rows(min_row=1, max_row=row_index - 1):
#         for cell in row:
#             cell.protection = Protection(locked=True)

#     # Unlock only monthly columns (April–March)
#     for row in ws.iter_rows(min_row=2, max_row=row_index - 1, min_col=16, max_col=27):
#         for cell in row:
#             cell.protection = Protection(locked=False)

#     ws.protection.sheet = True
#     ws.protection.password = "[removed]"

#     # Auto column width
#     for column in ws.columns:
#         max_length = 0
#         column_letter = column[0].column_letter
#         for cell in column:
#             if cell.value:
#                 max_length = max(max_length, len(str(cell.value)))
#         ws.column_dimensions[column_letter].width = max_length + 2

#     output = BytesIO()
#     wb.save(output)
#     output.seek(0)

#     current_date = datetime.datetime.now().strftime("%Y-%m-%d")
#     frappe.response["filename"] = f"Budget_mis_Import_{current_date}.xlsx"
#     frappe.response["filecontent"] = output.getvalue()
#     frappe.response["type"] = "download"

# @frappe.whitelist()
# def download_finance_budget_import_template(user):

#     import frappe
#     from frappe import _
#     import datetime
#     from io import BytesIO
#     import xlsxwriter

#     if not user:
#         frappe.throw(_("User is required"))

#     # Get Finance User Access
#     doc_name = frappe.db.get_value(
#         "Finance user access",
#         {"user": user},
#         "name"
#     )

#     if not doc_name:
#         frappe.throw(_("No Finance User Access found"))

#     access_doc = frappe.get_doc("Finance user access", doc_name)

#     if not access_doc.import_template_id:
#         frappe.throw(_("Import Template not linked in Finance User Access"))

#     import_template = frappe.get_doc(
#         "Import Templates",
#         access_doc.import_template_id
#     )

#     template_items = import_template.import_template_item_list

#     if not template_items:
#         frappe.throw(_("No Import Template Items found"))

#     financial_year = frappe.db.get_single_value(
#         "Master Settings",
#         "current_financial_year"
#     )

#     headers = [
#         "Entity / Unit",
#         "Entity / Unit Description",
#         "Cost Center",
#         "Cost Center(Original)",
#         "Cost Center Description",
#         "Location code",
#         "Location code(Original)",
#         "Function / Sub Unit / Division",
#         "State",
#         "Financial year",
#         "Uploaded By",
#         "Type of expense ID (Budget Amounts)",
#         "Head of expense (Budget Amounts)",
#         "Sub head of expense (Budget Amounts)",
#         "Type of expense (Budget Amounts)",
#         "April (Budget Amounts)",
#         "May (Budget Amounts)",
#         "June (Budget Amounts)",
#         "July (Budget Amounts)",
#         "August (Budget Amounts)",
#         "September (Budget Amounts)",
#         "October (Budget Amounts)",
#         "November (Budget Amounts)",
#         "December (Budget Amounts)",
#         "January (Budget Amounts)",
#         "February (Budget Amounts)",
#         "March (Budget Amounts)",
#         "Quarter 1 Total Amount (Budget Amounts)",
#         "Quarter 2 Total Amount (Budget Amounts)",
#         "Quarter 3 Total Amount (Budget Amounts)",
#         "Quarter 4 Total Amount (Budget Amounts)",
#         "Year Total Amount (Budget Amounts)"
#     ]

#     output = BytesIO()

#     workbook = xlsxwriter.Workbook(output)
#     worksheet = workbook.add_worksheet("Finance Budget Import")

#     # Formats
#     header_format = workbook.add_format({
#         "bold": True,
#         "locked": True
#     })

#     locked_text = workbook.add_format({
#         "locked": True
#     })

#     locked_number = workbook.add_format({
#         "locked": True,
#         "num_format": "0.00"
#     })

#     unlocked_number = workbook.add_format({
#         "locked": False,
#         "num_format": "0.00"
#     })

#     # Write headers
#     for col, header in enumerate(headers):
#         worksheet.write(0, col, header, header_format)
#         worksheet.set_column(col, col, 22)

#     worksheet.freeze_panes(1, 0)

#     row_index = 1

#     for mapping in access_doc.mapping:

#         first_row = True

#         for item in template_items:

#             if first_row:
#                 parent_values = [
#                     mapping.unit,
#                     mapping.unit_description,
#                     mapping.cost_center,
#                     mapping.cost_center_erp,
#                     mapping.cost_center_description,
#                     mapping.location_code,
#                     mapping.location_code_erp,
#                     mapping.location_description,
#                     mapping.state,
#                     financial_year,
#                     user
#                 ]
#                 first_row = False
#             else:
#                 parent_values = [""] * 11

#             row_data = parent_values + [
#                 item.type_of_expense_id,
#                 item.head_of_expense,
#                 item.sub_head_of_expense,
#                 item.type_of_expense
#             ]

#             # Write locked columns
#             for col, value in enumerate(row_data):
#                 worksheet.write(row_index, col, value, locked_text)

#             # Write editable month columns (April–March)
#             for col in range(15, 27):
#                 worksheet.write(row_index, col, 0.00, unlocked_number)

#             # Quarter formulas
#             r = row_index + 1

#             worksheet.write_formula(row_index, 27, f"=SUM(P{r}:R{r})", locked_number)
#             worksheet.write_formula(row_index, 28, f"=SUM(S{r}:U{r})", locked_number)
#             worksheet.write_formula(row_index, 29, f"=SUM(V{r}:X{r})", locked_number)
#             worksheet.write_formula(row_index, 30, f"=SUM(Y{r}:AA{r})", locked_number)

#             # Year total
#             worksheet.write_formula(row_index, 31, f"=SUM(AB{r}:AE{r})", locked_number)

#             row_index += 1

#     # Enable sheet protection
#     worksheet.protect("[REDACTED-PASSWORD]")

#     workbook.close()

#     output.seek(0)

#     current_date = datetime.datetime.now().strftime("%Y-%m-%d")

#     frappe.response["filename"] = f"Budget_mis_Import_{current_date}.xlsx"
#     frappe.response["filecontent"] = output.getvalue()
#     frappe.response["type"] = "download"


# @frappe.whitelist()
# def start_budget_template_generation(user):

#     if not user:
#         frappe.throw("User required")

#     frappe.enqueue(
#         "annual_budget.api.export_reports.generate_budget_template",
#         queue="long",
#         timeout=1800,
#         user=user
#     )

#     return {"status": "started"}


# def generate_budget_template(user):

#     doc_name = frappe.db.get_value(
#         "Finance user access",
#         {"user": user},
#         "name"
#     )

#     if not doc_name:
#         return

#     access_doc = frappe.get_doc("Finance user access", doc_name)

#     import_template = frappe.get_doc(
#         "Import Templates",
#         access_doc.import_template_id
#     )

#     template_items = import_template.import_template_item_list

#     financial_year = frappe.db.get_single_value(
#         "Master Settings",
#         "current_financial_year"
#     )

#     tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")

#     workbook = xlsxwriter.Workbook(tmp.name)
#     worksheet = workbook.add_worksheet("Finance Budget Import")

#     header_format = workbook.add_format({"bold": True, "locked": True})
#     locked = workbook.add_format({"locked": True})
#     unlocked = workbook.add_format({"locked": False, "num_format": "0.00"})

#     headers = [
#         "Entity / Unit",
#         "Entity / Unit Description",
#         "Cost Center",
#         "Cost Center(Original)",
#         "Cost Center Description",
#         "Location code",
#         "Location code(Original)",
#         "Function / Sub Unit / Division",
#         "State",
#         "Financial year",
#         "Uploaded By",
#         "Type of expense ID (Budget Amounts)",
#         "Head of expense (Budget Amounts)",
#         "Sub head of expense (Budget Amounts)",
#         "Type of expense (Budget Amounts)",
#         "April (Budget Amounts)",
#         "May (Budget Amounts)",
#         "June (Budget Amounts)",
#         "July (Budget Amounts)",
#         "August (Budget Amounts)",
#         "September (Budget Amounts)",
#         "October (Budget Amounts)",
#         "November (Budget Amounts)",
#         "December (Budget Amounts)",
#         "January (Budget Amounts)",
#         "February (Budget Amounts)",
#         "March (Budget Amounts)",
#         "Quarter 1 Total Amount (Budget Amounts)",
#         "Quarter 2 Total Amount (Budget Amounts)",
#         "Quarter 3 Total Amount (Budget Amounts)",
#         "Quarter 4 Total Amount (Budget Amounts)",
#         "Year Total Amount (Budget Amounts)"
#     ]

#     # Track column width
#     col_widths = [len(h) for h in headers]

#     for col, header in enumerate(headers):
#         worksheet.write(0, col, header, header_format)

#     worksheet.freeze_panes(1, 0)

#     row_index = 1

#     for mapping in access_doc.mapping:

#         first_row = True

#         for item in template_items:

#             if first_row:
#                 parent_values = [
#                     mapping.unit,
#                     mapping.unit_description,
#                     mapping.cost_center,
#                     mapping.cost_center_erp,
#                     mapping.cost_center_description,
#                     mapping.location_code,
#                     mapping.location_code_erp,
#                     mapping.location_description,
#                     mapping.state,
#                     financial_year,
#                     user
#                 ]
#                 first_row = False
#             else:
#                 parent_values = [""] * 11

#             row = parent_values + [
#                 item.type_of_expense_id,
#                 item.head_of_expense,
#                 item.sub_head_of_expense,
#                 item.type_of_expense
#             ]

#             for col, val in enumerate(row):

#                 worksheet.write(row_index, col, val, locked)

#                 if val:
#                     col_widths[col] = max(col_widths[col], len(str(val)))

#             # Editable months
#             for col in range(15, 27):
#                 worksheet.write(row_index, col, 0, unlocked)

#             r = row_index + 1

#             worksheet.write_formula(row_index, 27, f"=SUM(P{r}:R{r})")
#             worksheet.write_formula(row_index, 28, f"=SUM(S{r}:U{r})")
#             worksheet.write_formula(row_index, 29, f"=SUM(V{r}:X{r})")
#             worksheet.write_formula(row_index, 30, f"=SUM(Y{r}:AA{r})")
#             worksheet.write_formula(row_index, 31, f"=SUM(AB{r}:AE{r})")

#             row_index += 1

#     worksheet.protect("[REDACTED-PASSWORD]")

#     # Apply calculated widths
#     for i, width in enumerate(col_widths):
#         worksheet.set_column(i, i, width + 3)

#     workbook.close()

#     frappe.cache().set_value(
#         f"budget_template_{user}",
#         tmp.name,
#         expires_in_sec=3600
#     )


# @frappe.whitelist()
# def download_generated_template(user):

#     path = frappe.cache().get_value(f"budget_template_{user}")

#     if not path:
#         return {"status": "processing"}

#     with open(path, "rb") as f:

#         frappe.response["filename"] = f"Budget_mis_Import_{nowdate()}.xlsx"
#         frappe.response["filecontent"] = f.read()
#         frappe.response["type"] = "download"



    # import frappe
    # import tempfile
    # import xlsxwriter
    # from frappe.utils import nowdate


    # @frappe.whitelist()
    # def start_budget_template_generation(user, entity_data=None):

    #     if not user:
    #         frappe.throw("User required")

    #     frappe.enqueue(
    #         "annual_budget.api.export_reports.generate_budget_template",
    #         queue="long",
    #         timeout=1800,
    #         user=user,
    #         entity_data=entity_data
    #     )

    #     return {"status": "started"}


    # def generate_budget_template(user, entity_data=None):

    #     # ── Parse entity_data ─────────────────────────────────
    #     # Can be:
    #     #   None          → use all mappings from access_doc
    #     #   JSON string   → parse into list or single dict
    #     #   list          → multiple selected rows
    #     #   dict          → single selected row (legacy)

    #     if entity_data and isinstance(entity_data, str):
    #         entity_data = frappe.parse_json(entity_data)

    #     # Normalise to a list (or None)
    #     if isinstance(entity_data, dict):
    #         entity_data = [entity_data]   # wrap single row in a list

    #     # entity_data is now either a list of dicts or None

    #     # ── Fetch Finance User Access ──────────────────────────
    #     doc_name = frappe.db.get_value(
    #         "Finance user access",
    #         {"user": user},
    #         "name"
    #     )

    #     access_doc = frappe.get_doc("Finance user access", doc_name) if doc_name else None

    #     if not access_doc:
    #         return

    #     # ── Import Template ────────────────────────────────────
    #     import_template = frappe.get_doc(
    #         "Import Templates",
    #         access_doc.import_template_id
    #     )

    #     template_items = import_template.import_template_item_list

    #     financial_year = frappe.db.get_single_value(
    #         "Master Settings",
    #         "current_financial_year"
    #     )

    #     # ── Decide mappings source ─────────────────────────────
    #     # If entity_data list provided → use those rows
    #     # Otherwise → use all rows from access_doc.mapping
    #     if entity_data:
    #         mappings       = entity_data          # list of plain dicts from frontend
    #         use_dict       = True                 # access via .get()
    #     else:
    #         mappings       = access_doc.mapping   # list of child doc objects
    #         use_dict       = False                # access via getattr()

    #     # ── Helper: read a field from either a dict or a child doc ──
    #     def get_field(field, mapping):
    #         if use_dict:
    #             return mapping.get(field) or ""
    #         return getattr(mapping, field, "") or ""

    #     # ── Create Excel ───────────────────────────────────────
    #     tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")

    #     workbook  = xlsxwriter.Workbook(tmp.name)
    #     worksheet = workbook.add_worksheet("Finance Budget Import")

    #     header_format = workbook.add_format({"bold": True, "locked": True})
    #     locked        = workbook.add_format({"locked": True})
    #     unlocked      = workbook.add_format({"locked": False, "num_format": "0.00"})

    #     headers = [
    #         "Entity / Unit",
    #         "Entity / Unit Description",
    #         "Cost Center",
    #         "Cost Center(Original)",
    #         "Cost Center Description",
    #         "Location code",
    #         "Location code(Original)",
    #         "Function / Sub Unit / Division",
    #         "State",
    #         "Financial year",
    #         "Uploaded By",
    #         "Type of expense ID (Budget Amounts)",
    #         "Head of expense (Budget Amounts)",
    #         "Sub head of expense (Budget Amounts)",
    #         "Type of expense (Budget Amounts)",
    #         "April (Budget Amounts)",
    #         "May (Budget Amounts)",
    #         "June (Budget Amounts)",
    #         "July (Budget Amounts)",
    #         "August (Budget Amounts)",
    #         "September (Budget Amounts)",
    #         "October (Budget Amounts)",
    #         "November (Budget Amounts)",
    #         "December (Budget Amounts)",
    #         "January (Budget Amounts)",
    #         "February (Budget Amounts)",
    #         "March (Budget Amounts)",
    #         "Quarter 1 Total Amount (Budget Amounts)",
    #         "Quarter 2 Total Amount (Budget Amounts)",
    #         "Quarter 3 Total Amount (Budget Amounts)",
    #         "Quarter 4 Total Amount (Budget Amounts)",
    #         "Year Total Amount (Budget Amounts)"
    #     ]

    #     col_widths = [len(h) for h in headers]

    #     for col, header in enumerate(headers):
    #         worksheet.write(0, col, header, header_format)

    #     worksheet.freeze_panes(1, 0)

    #     row_index = 1

    #     # ── Write one block of rows per mapping ───────────────
    #     for mapping in mappings:

    #         first_row = True

    #         for item in template_items:

    #             if first_row:
    #                 parent_values = [
    #                     get_field("unit",                   mapping),
    #                     get_field("unit_description",       mapping),
    #                     get_field("cost_center",            mapping),
    #                     get_field("cost_center_erp",        mapping),
    #                     get_field("cost_center_description",mapping),
    #                     get_field("location_code",          mapping),
    #                     get_field("location_code_erp",      mapping),
    #                     get_field("location_description",   mapping),
    #                     get_field("state",                  mapping),
    #                     financial_year,
    #                     user
    #                 ]
    #                 first_row = False
    #             else:
    #                 parent_values = [""] * 11

    #             row = parent_values + [
    #                 item.type_of_expense_id,
    #                 item.head_of_expense,
    #                 item.sub_head_of_expense,
    #                 item.type_of_expense
    #             ]

    #             for col, val in enumerate(row):
    #                 worksheet.write(row_index, col, val, locked)
    #                 if val:
    #                     col_widths[col] = max(col_widths[col], len(str(val)))

    #             for col in range(15, 27):
    #                 worksheet.write(row_index, col, 0, unlocked)

    #             r = row_index + 1
    #             worksheet.write_formula(row_index, 27, f"=SUM(P{r}:R{r})")
    #             worksheet.write_formula(row_index, 28, f"=SUM(S{r}:U{r})")
    #             worksheet.write_formula(row_index, 29, f"=SUM(V{r}:X{r})")
    #             worksheet.write_formula(row_index, 30, f"=SUM(Y{r}:AA{r})")
    #             worksheet.write_formula(row_index, 31, f"=SUM(AB{r}:AE{r})")

    #             row_index += 1

    #     worksheet.protect("[REDACTED-PASSWORD]")

    #     for i, width in enumerate(col_widths):
    #         worksheet.set_column(i, i, width + 3)

    #     workbook.close()

    #     frappe.cache().set_value(
    #         f"budget_template_{user}",
    #         tmp.name,
    #         expires_in_sec=3600
    #     )


    # @frappe.whitelist()
    # def download_generated_template(user):

    #     path = frappe.cache().get_value(f"budget_template_{user}")

    #     if not path:
    #         return {"status": "processing"}

    #     with open(path, "rb") as f:
    #         frappe.response["filename"]    = f"Budget_mis_Import_{nowdate()}.xlsx"
    #         frappe.response["filecontent"] = f.read()
    #         frappe.response["type"]        = "download"




import frappe
import tempfile
import xlsxwriter
from frappe.utils import nowdate


@frappe.whitelist()
def start_budget_template_generation(entity_data=None):

    user = frappe.session.user

    frappe.enqueue(
        "annual_budget.api.export_reports.generate_budget_template",
        queue="long",
        timeout=1800,
        user=user,
        entity_data=entity_data
    )

    return {"status": "started"}


def generate_budget_template(user=None, entity_data=None):

    user = user or frappe.session.user

    # ── Parse entity_data ─────────────────────────────────
    if entity_data and isinstance(entity_data, str):
        entity_data = frappe.parse_json(entity_data)

    if isinstance(entity_data, dict):
        entity_data = [entity_data]

    # ── Fetch Finance User Access ─────────────────────────
    doc_name = frappe.db.get_value(
        "Finance user access",
        {"user": user},
        "name"
    )

    access_doc = frappe.get_doc("Finance user access", doc_name) if doc_name else None

    if not access_doc:
        return

    allow_edit = access_doc.allow_edit_template or 0

    # ── Import Template ───────────────────────────────────
    import_template = frappe.get_doc(
        "Import Templates",
        access_doc.import_template_id
    )

    template_items = import_template.import_template_item_list

    financial_year = frappe.db.get_single_value(
        "Master Settings",
        "current_financial_year"
    )

    # ── Decide mappings source ─────────────────────────────
    if entity_data:
        mappings = entity_data
        use_dict = True
    else:
        mappings = access_doc.mapping
        use_dict = False

    # ── Helper ────────────────────────────────────────────
    def get_field(frontend_key, child_key, mapping):
        if use_dict:
            return mapping.get(frontend_key) or ""
        return getattr(mapping, child_key, "") or ""

    # ── Create Excel ──────────────────────────────────────
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")

    workbook  = xlsxwriter.Workbook(tmp.name)
    worksheet = workbook.add_worksheet("Finance Budget Import")

    header_format = workbook.add_format({"bold": True, "locked": True})

    if allow_edit:
        locked = workbook.add_format({"locked": False})
        unlocked = workbook.add_format({"locked": False, "num_format": "0.00"})
    else:
        locked = workbook.add_format({"locked": True})
        unlocked = workbook.add_format({"locked": False, "num_format": "0.00"})

    headers = [
        "Entity / Unit",
        "Entity / Unit Description",
        "Cost Center",
        "Cost Center(Original)",
        "Cost Center Description",
        "Location code",
        "Location code(Original)",
        "Location Description",
        "State",
        "Financial year",
        "Uploaded By",
        "Type of expense ID (Budget Amounts)",
        "Head of expense (Budget Amounts)",
        "Sub head of expense (Budget Amounts)",
        "Type of expense (Budget Amounts)",
        "April (Budget Amounts)",
        "May (Budget Amounts)",
        "June (Budget Amounts)",
        "July (Budget Amounts)",
        "August (Budget Amounts)",
        "September (Budget Amounts)",
        "October (Budget Amounts)",
        "November (Budget Amounts)",
        "December (Budget Amounts)",
        "January (Budget Amounts)",
        "February (Budget Amounts)",
        "March (Budget Amounts)",
        "Quarter 1 Total Amount (Budget Amounts)",
        "Quarter 2 Total Amount (Budget Amounts)",
        "Quarter 3 Total Amount (Budget Amounts)",
        "Quarter 4 Total Amount (Budget Amounts)",
        "Year Total Amount (Budget Amounts)"
    ]

    col_widths = [len(h) for h in headers]

    for col, header in enumerate(headers):
        worksheet.write(0, col, header, header_format)

    worksheet.freeze_panes(1, 0)

    row_index = 1

    for mapping in mappings:

        first_row = True

        for item in template_items:

            if first_row:
                parent_values = [
                    get_field("unit", "unit", mapping),
                    get_field("unit_description", "unit_description", mapping),

                    # ✅ FIXED COST CENTER
                    get_field("cost_center_id", "cost_center", mapping),
                    get_field("cost_center", "cost_center_erp", mapping),

                    get_field("cost_center_description", "cost_center_description", mapping),

                    # ✅ FIXED LOCATION
                    get_field("location_code_id", "location_code", mapping),
                    get_field("location_code", "location_code_erp", mapping),

                    get_field("location_description", "location_description", mapping),
                    get_field("state", "state", mapping),
                    financial_year,
                    user
                ]
                first_row = False
            else:
                parent_values = [""] * 11

            row = parent_values + [
                item.type_of_expense_id,
                item.head_of_expense,
                item.sub_head_of_expense,
                item.type_of_expense
            ]

            for col, val in enumerate(row):
                worksheet.write(row_index, col, val, locked)
                if val:
                    col_widths[col] = max(col_widths[col], len(str(val)))

            # Editable monthly fields
            for col in range(15, 27):
                worksheet.write(row_index, col, 0, unlocked)

            r = row_index + 1
            worksheet.write_formula(row_index, 27, f"=SUM(P{r}:R{r})")
            worksheet.write_formula(row_index, 28, f"=SUM(S{r}:U{r})")
            worksheet.write_formula(row_index, 29, f"=SUM(V{r}:X{r})")
            worksheet.write_formula(row_index, 30, f"=SUM(Y{r}:AA{r})")
            worksheet.write_formula(row_index, 31, f"=SUM(AB{r}:AE{r})")

            row_index += 1

    if not allow_edit:
        worksheet.protect("[REDACTED-PASSWORD]")

    for i, width in enumerate(col_widths):
        worksheet.set_column(i, i, width + 3)

    workbook.close()

    frappe.cache().set_value(
        f"budget_template_{user}",
        tmp.name,
        expires_in_sec=3600
    )


@frappe.whitelist()
def download_generated_template():

    user = frappe.session.user

    path = frappe.cache().get_value(f"budget_template_{user}")

    if not path:
        return {"status": "processing"}

    with open(path, "rb") as f:
        frappe.response["filename"]    = f"Budget_mis_Import_{nowdate()}.xlsx"
        frappe.response["filecontent"] = f.read()
        frappe.response["type"]        = "download"
        



# import tempfile
# import xlsxwriter
# from frappe.utils import nowdate
# import frappe


# @frappe.whitelist()
# def download_finance_budget_import_template(user):

#     if not user:
#         frappe.throw("User required")

#     doc_name = frappe.db.get_value(
#         "Finance user access",
#         {"user": user},
#         "name"
#     )

#     if not doc_name:
#         frappe.throw("No Finance User Access record found for this user.")

#     access_doc = frappe.get_doc("Finance user access", doc_name)

#     # ── allow_edit_template checkbox ──────────────────────
#     # Checked (1)   → skip sheet protection → entire sheet is editable
#     # Unchecked (0) → protect sheet         → only month columns are editable
#     is_editable = access_doc.allow_edit_template == 1

#     import_template = frappe.get_doc(
#         "Import Templates",
#         access_doc.import_template_id
#     )

#     template_items = import_template.import_template_item_list

#     financial_year = frappe.db.get_single_value(
#         "Master Settings",
#         "current_financial_year"
#     )

#     tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")

#     workbook  = xlsxwriter.Workbook(tmp.name)
#     worksheet = workbook.add_worksheet("Finance Budget Import")

#     header_format = workbook.add_format({"bold": True, "locked": True})
#     locked        = workbook.add_format({"locked": True})

#     # Month columns are always written with locked=False.
#     # When is_editable=False → sheet is protected → only these cells are editable.
#     # When is_editable=True  → sheet is not protected → all cells are editable anyway.
#     unlocked = workbook.add_format({"locked": False, "num_format": "0.00"})

#     headers = [
#         "Entity / Unit",
#         "Entity / Unit Description",
#         "Cost Center",
#         "Cost Center(Original)",
#         "Cost Center Description",
#         "Location code",
#         "Location code(Original)",
#         "Function / Sub Unit / Division",
#         "State",
#         "Financial year",
#         "Uploaded By",
#         "Type of expense ID (Budget Amounts)",
#         "Head of expense (Budget Amounts)",
#         "Sub head of expense (Budget Amounts)",
#         "Type of expense (Budget Amounts)",
#         "April (Budget Amounts)",
#         "May (Budget Amounts)",
#         "June (Budget Amounts)",
#         "July (Budget Amounts)",
#         "August (Budget Amounts)",
#         "September (Budget Amounts)",
#         "October (Budget Amounts)",
#         "November (Budget Amounts)",
#         "December (Budget Amounts)",
#         "January (Budget Amounts)",
#         "February (Budget Amounts)",
#         "March (Budget Amounts)",
#         "Quarter 1 Total Amount (Budget Amounts)",
#         "Quarter 2 Total Amount (Budget Amounts)",
#         "Quarter 3 Total Amount (Budget Amounts)",
#         "Quarter 4 Total Amount (Budget Amounts)",
#         "Year Total Amount (Budget Amounts)"
#     ]

#     col_widths = [len(h) for h in headers]

#     for col, header in enumerate(headers):
#         worksheet.write(0, col, header, header_format)

#     worksheet.freeze_panes(1, 0)

#     row_index = 1

#     for mapping in access_doc.mapping:

#         first_row = True

#         for item in template_items:

#             if first_row:
#                 parent_values = [
#                     mapping.unit,
#                     mapping.unit_description,
#                     mapping.cost_center,
#                     mapping.cost_center_erp,
#                     mapping.cost_center_description,
#                     mapping.location_code,
#                     mapping.location_code_erp,
#                     mapping.location_description,
#                     mapping.state,
#                     financial_year,
#                     user
#                 ]
#                 first_row = False
#             else:
#                 parent_values = [""] * 11

#             row = parent_values + [
#                 item.type_of_expense_id,
#                 item.head_of_expense,
#                 item.sub_head_of_expense,
#                 item.type_of_expense
#             ]

#             for col, val in enumerate(row):
#                 worksheet.write(row_index, col, val, locked)
#                 if val:
#                     col_widths[col] = max(col_widths[col], len(str(val)))

#             # Month columns (April–March)
#             for col in range(15, 27):
#                 worksheet.write(row_index, col, 0, unlocked)

#             r = row_index + 1

#             worksheet.write_formula(row_index, 27, f"=SUM(P{r}:R{r})")
#             worksheet.write_formula(row_index, 28, f"=SUM(S{r}:U{r})")
#             worksheet.write_formula(row_index, 29, f"=SUM(V{r}:X{r})")
#             worksheet.write_formula(row_index, 30, f"=SUM(Y{r}:AA{r})")
#             worksheet.write_formula(row_index, 31, f"=SUM(AB{r}:AE{r})")

#             row_index += 1

#     # Protect only when allow_edit_template is unchecked
#     if not is_editable:
#         worksheet.protect("[REDACTED-PASSWORD]")

#     for i, width in enumerate(col_widths):
#         worksheet.set_column(i, i, width + 3)

#     workbook.close()

#     with open(tmp.name, "rb") as f:
#         frappe.response["filename"]    = f"Budget_Import_{nowdate()}.xlsx"
#         frappe.response["filecontent"] = f.read()
#         frappe.response["type"]        = "download"






# import frappe
# import tempfile
# import xlsxwriter
# from frappe.utils import nowdate, cint


# # ✅ START JOB (with lock)
# @frappe.whitelist()
# def start_budget_template_generation(user):

#     if not user:
#         frappe.throw("User required")

#     job_key = f"budget_template_job_{user}"

#     # Prevent duplicate jobs
#     if frappe.cache().get_value(job_key):
#         return {"status": "already_running"}

#     frappe.cache().set_value(job_key, True, expires_in_sec=600)

#     frappe.enqueue(
#         "annual_budget.api.export_reports.generate_budget_template",
#         queue="long",
#         timeout=1800,
#         user=user
#     )

#     return {"status": "started"}


# # ✅ BACKGROUND JOB
# def generate_budget_template(user):

#     try:
#         frappe.cache().delete_value(f"budget_template_{user}")

#         doc_name = frappe.db.get_value(
#             "Finance user access",
#             {"user": user},
#             "name"
#         )

#         if not doc_name:
#             return

#         access_doc = frappe.get_doc("Finance user access", doc_name)

#         # ✅ FIX checkbox inconsistency
#         is_editable = cint(access_doc.allow_edit_template) == 1

#         import_template = frappe.get_doc(
#             "Import Templates",
#             access_doc.import_template_id
#         )

#         template_items = import_template.import_template_item_list

#         financial_year = frappe.db.get_single_value(
#             "Master Settings",
#             "current_financial_year"
#         )

#         tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")

#         # ✅ Faster workbook
#         workbook = xlsxwriter.Workbook(tmp.name, {'constant_memory': True})
#         worksheet = workbook.add_worksheet("Finance Budget Import")

#         # Formats
#         unlocked_format = workbook.add_format({
#             "locked": False,
#             "num_format": "0.00"
#         })

#         locked_format = workbook.add_format({"locked": True})

#         header_format = workbook.add_format({
#             "bold": True,
#             "locked": not is_editable
#         })

#         headers = [
#             "Entity / Unit", "Entity / Unit Description", "Cost Center",
#             "Cost Center(Original)", "Cost Center Description",
#             "Location code", "Location code(Original)",
#             "Function / Sub Unit / Division", "State",
#             "Financial year", "Uploaded By",
#             "Type of expense ID (Budget Amounts)",
#             "Head of expense (Budget Amounts)",
#             "Sub head of expense (Budget Amounts)",
#             "Type of expense (Budget Amounts)",
#             "April (Budget Amounts)", "May (Budget Amounts)",
#             "June (Budget Amounts)", "July (Budget Amounts)",
#             "August (Budget Amounts)", "September (Budget Amounts)",
#             "October (Budget Amounts)", "November (Budget Amounts)",
#             "December (Budget Amounts)", "January (Budget Amounts)",
#             "February (Budget Amounts)", "March (Budget Amounts)",
#             "Quarter 1 Total Amount (Budget Amounts)",
#             "Quarter 2 Total Amount (Budget Amounts)",
#             "Quarter 3 Total Amount (Budget Amounts)",
#             "Quarter 4 Total Amount (Budget Amounts)",
#             "Year Total Amount (Budget Amounts)"
#         ]

#         # Write headers
#         for col, header in enumerate(headers):
#             worksheet.write(0, col, header, header_format)

#         worksheet.freeze_panes(1, 0)

#         row_index = 1

#         base_format = unlocked_format if is_editable else locked_format

#         for mapping in access_doc.mapping:

#             first_row = True

#             for item in template_items:

#                 if first_row:
#                     parent_values = [
#                         mapping.unit,
#                         mapping.unit_description,
#                         mapping.cost_center,
#                         mapping.cost_center_erp,
#                         mapping.cost_center_description,
#                         mapping.location_code,
#                         mapping.location_code_erp,
#                         mapping.location_description,
#                         mapping.state,
#                         financial_year,
#                         user
#                     ]
#                     first_row = False
#                 else:
#                     parent_values = [""] * 11

#                 row = parent_values + [
#                     item.type_of_expense_id,
#                     item.head_of_expense,
#                     item.sub_head_of_expense,
#                     item.type_of_expense
#                 ]

#                 # 🚀 Fast row write
#                 worksheet.write_row(row_index, 0, row, base_format)

#                 # Month columns always editable
#                 for col in range(15, 27):
#                     worksheet.write(row_index, col, 0, unlocked_format)

#                 r = row_index + 1

#                 worksheet.write_formula(row_index, 27, f"=SUM(P{r}:R{r})")
#                 worksheet.write_formula(row_index, 28, f"=SUM(S{r}:U{r})")
#                 worksheet.write_formula(row_index, 29, f"=SUM(V{r}:X{r})")
#                 worksheet.write_formula(row_index, 30, f"=SUM(Y{r}:AA{r})")
#                 worksheet.write_formula(row_index, 31, f"=SUM(AB{r}:AE{r})")

#                 row_index += 1

#         # ✅ Apply filter ONLY when editable
#         if is_editable and row_index > 1:
#             worksheet.autofilter(0, 0, row_index - 1, len(headers) - 1)

#         # Protect if needed
#         if not is_editable:
#             worksheet.protect("[REDACTED-PASSWORD]")

#         worksheet.set_column(0, len(headers) - 1, 22)

#         workbook.close()

#         frappe.cache().set_value(
#             f"budget_template_{user}",
#             tmp.name,
#             expires_in_sec=3600
#         )

#     finally:
#         # Always release lock
#         frappe.cache().delete_value(f"budget_template_job_{user}")


# # ✅ STATUS API
# @frappe.whitelist()
# def check_template_status(user):

#     path = frappe.cache().get_value(f"budget_template_{user}")

#     if path:
#         return {"status": "ready"}

#     return {"status": "processing"}


# # ✅ DOWNLOAD API
# @frappe.whitelist()
# def download_generated_template(user):

#     path = frappe.cache().get_value(f"budget_template_{user}")

#     if not path:
#         return {"status": "processing"}

#     with open(path, "rb") as f:
#         frappe.response["filename"] = f"Budget_mis_Import_{nowdate()}.xlsx"
#         frappe.response["filecontent"] = f.read()
#         frappe.response["type"] = "download"





















# annual_budget/api/export_reports.py
#
# Server-side Excel export — design matches the screenshot:
#   Row 1       : Org name  (no fill, bold 14pt)
#   Row 2       : Subtitle  (no fill, bold 12pt)
#   Row 3       : Blank spacer
#   Row 4-5     : Quarter / month headers  (grey #5D6D7E, white bold)
#   Section rows: A / B … label  (grey #D6DBDF, bold, merged)
#   Sub-head    : I / II … label  (light grey #F2F3F4, bold, merged)
#   Item rows   : white, normal
#   Sub-total   : #EBF5FB, bold
#   Head total  : #D4E6F1, bold
#   Grand total : #A9CCE3, bold
#
# Numbers use Excel SUM() formulas exactly like the reference code.
# Requires:  bench pip install openpyxl
# Place at:  annual_budget/api/export_reports.py




# import io
# import base64
# import json

# import frappe
# from openpyxl import Workbook
# from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
# from openpyxl.utils import get_column_letter

# # ═══════════════════════════════════════════════════════════════════════════════
# # SHARED STYLE CONSTANTS  (matching reference code exactly)
# # ═══════════════════════════════════════════════════════════════════════════════

# _CENTER  = Alignment(horizontal="center", vertical="center")
# _LEFT    = Alignment(horizontal="left",   vertical="center")
# _RIGHT   = Alignment(horizontal="right",  vertical="center")
# _WRAP_L  = Alignment(horizontal="left",   vertical="center", wrap_text=True)

# _BOLD       = Font(bold=True, name="Calibri", size=9)
# _WHITE_BOLD = Font(bold=True, color="FFFFFF", name="Calibri", size=9)
# _NORMAL     = Font(name="Calibri", size=9)

# _FILL_HDR      = PatternFill("solid", fgColor="5D6D7E")   # dark blue-grey  — column headers
# _FILL_HEAD     = PatternFill("solid", fgColor="D6DBDF")   # medium grey     — section rows (A/B/…)
# _FILL_SUBHEAD  = PatternFill("solid", fgColor="F2F3F4")   # light grey      — sub-head rows (I/II/…)
# _FILL_SUBTOTAL = PatternFill("solid", fgColor="EBF5FB")   # pale blue       — sub-total rows
# _FILL_HTOTAL   = PatternFill("solid", fgColor="D4E6F1")   # mid blue        — head total rows
# _FILL_GRAND    = PatternFill("solid", fgColor="A9CCE3")   # stronger blue   — grand total
# _FILL_WHITE    = PatternFill("solid", fgColor="FFFFFF")

# _THIN   = Side(style="thin")
# _BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# NUM_FMT = "#,##0.00"

# # Column indices (1-based)
# COL_SI    = 2   # Sl #
# COL_HEAD  = 3   # HEAD OF EXPENSE
# COL_TYPE  = 4   # TYPE OF EXPENSE
# COL_START = 5   # Apr  (first data column)
# COL_END   = 21  # YEAR total  (last data column)
# # Cols 5-7  = Apr May Jun
# # Cols 8-10 = Jul Aug Sep
# # Cols 11-13= Oct Nov Dec
# # Cols 14-16= Jan Feb Mar
# # Cols 17-20= QTR-1 QTR-2 QTR-3 QTR-4
# # Col  21   = YEAR total


# # ═══════════════════════════════════════════════════════════════════════════════
# # LOW-LEVEL HELPERS
# # ═══════════════════════════════════════════════════════════════════════════════

# def _merge(ws, r1, c1, r2, c2):
#     ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)


# def _to_roman(num):
#     vals = [1000,900,500,400,100,90,50,40,10,9,5,4,1]
#     syms = ["M","CM","D","CD","C","XC","L","XL","X","IX","V","IV","I"]
#     out = ""
#     i = 0
#     while num > 0:
#         for _ in range(num // vals[i]):
#             out += syms[i]
#             num -= vals[i]
#         i += 1
#     return out


# def _sum_formula(col_letter, rows):
#     """Build =R1+R2+… formula from a list of row numbers."""
#     if not rows:
#         return 0
#     return "=" + "+".join(f"{col_letter}{r}" for r in rows)


# def _qtr_formulas(r):
#     """
#     Return the 5 formula strings for a data row r:
#       QTR-1 = SUM(E:G), QTR-2 = SUM(H:J), QTR-3 = SUM(K:M),
#       QTR-4 = SUM(N:P), YEAR  = SUM(Q:T)
#     """
#     return [
#         f"=SUM(E{r}:G{r})",
#         f"=SUM(H{r}:J{r})",
#         f"=SUM(K{r}:M{r})",
#         f"=SUM(N{r}:P{r})",
#         f"=SUM(Q{r}:T{r})",
#     ]


# def _build_formula(col_letter, rows):
#     """Same as reference code's build_formula."""
#     if not rows:
#         return 0
#     return "=" + "+".join(f"{col_letter}{r}" for r in rows)


# def _totals_from_rows(rows):
#     """Return 17 formula cells (E..U) that sum the given row list."""
#     return [_build_formula(c, rows)
#             for c in list("EFGHIJKLMNOPQRSTU")]


# def _style_row(ws, row, fill=None, font=None, is_header=False):
#     """Apply border, alignment, fill, font to cols 2-21 (matching reference)."""
#     for col in range(COL_SI, COL_END + 1):
#         cell = ws.cell(row=row, column=col)
#         cell.border = _BORDER
#         if is_header:
#             cell.alignment = _CENTER
#         else:
#             if col == COL_SI:
#                 cell.alignment = _CENTER
#             elif col in (COL_HEAD, COL_TYPE):
#                 cell.alignment = _WRAP_L
#             else:
#                 cell.alignment = _RIGHT
#         if fill:
#             cell.fill = fill
#         if font:
#             cell.font = font


# def _fmt_numeric(ws, row):
#     for col in range(COL_START, COL_END + 1):
#         ws.cell(row=row, column=col).number_format = NUM_FMT


# def _auto_col_widths(ws):
#     for col in range(1, ws.max_column + 1):
#         letter = get_column_letter(col)
#         max_len = 0
#         for row in range(1, ws.max_row + 1):
#             val = ws.cell(row=row, column=col).value
#             if val:
#                 s = str(val)
#                 if s.startswith("="):
#                     s = "999,999,999.00"
#                 max_len = max(max_len, len(s))
#         ws.column_dimensions[letter].width = min(max(max_len + 3, 10), 50)


# def _wb_to_b64(wb):
#     buf = io.BytesIO()
#     wb.save(buf)
#     buf.seek(0)
#     return base64.b64encode(buf.read()).decode("utf-8")


# # ═══════════════════════════════════════════════════════════════════════════════
# # SHEET BUILDER — Annual Budget  (directly mirrors reference code logic)
# # ═══════════════════════════════════════════════════════════════════════════════

# def _sheet_annual(wb, sheet_name, data, fy,
#                   org_name="Azim Premji Foundation"):
#     ws = wb.create_sheet(title=sheet_name[:31])

#     # ── Row 1: Org name ───────────────────────────────────────────────────────
#     ws.append(["", org_name])
#     ws.merge_cells("B1:U1")
#     ws["B1"].font      = Font(size=14, bold=True, name="Calibri")
#     ws["B1"].alignment = _LEFT

#     # ── Row 2: Subtitle ───────────────────────────────────────────────────────
#     ws.append(["", f"Budget for the Financial Year {fy or ''}"])
#     ws.merge_cells("B2:U2")
#     ws["B2"].font      = Font(size=12, bold=True, name="Calibri")
#     ws["B2"].alignment = _LEFT

#     # ── Row 3: Blank ──────────────────────────────────────────────────────────
#     ws.append([])

#     # ── Rows 4-5: Headers — style BEFORE merging ─────────────────────────────
#     ws.append([
#         "", "Sl #", "HEAD OF EXPENSE", "TYPE OF EXPENSE",
#         "QUARTER I",  "", "",
#         "QUARTER II", "", "",
#         "QUARTER III","", "",
#         "QUARTER IV", "", "",
#         "QTR-1","QTR-2","QTR-3","QTR-4",
#         f"YEAR {fy}",
#     ])
#     r1 = ws.max_row

#     ws.append([
#         "", "Sl #", "HEAD OF EXPENSE", "TYPE OF EXPENSE",
#         "Apr","May","Jun",
#         "Jul","Aug","Sep",
#         "Oct","Nov","Dec",
#         "Jan","Feb","Mar",
#         "QTR-1","QTR-2","QTR-3","QTR-4",
#         f"YEAR {fy}",
#     ])
#     r2 = ws.max_row

#     # Style both header rows FIRST
#     _style_row(ws, r1, _FILL_HDR, _WHITE_BOLD, is_header=True)
#     _style_row(ws, r2, _FILL_HDR, _WHITE_BOLD, is_header=True)

#     # Now merge (covered cells already styled)
#     ws.merge_cells(start_row=r1, start_column=5,  end_row=r1, end_column=7)
#     ws.merge_cells(start_row=r1, start_column=8,  end_row=r1, end_column=10)
#     ws.merge_cells(start_row=r1, start_column=11, end_row=r1, end_column=13)
#     ws.merge_cells(start_row=r1, start_column=14, end_row=r1, end_column=16)
#     ws.merge_cells(start_row=r1, start_column=2,  end_row=r2, end_column=2)
#     ws.merge_cells(start_row=r1, start_column=3,  end_row=r2, end_column=3)
#     ws.merge_cells(start_row=r1, start_column=4,  end_row=r2, end_column=4)
#     for col in range(17, 22):
#         ws.merge_cells(start_row=r1, start_column=col, end_row=r2, end_column=col)

#     ws.freeze_panes = "E6"

#     # ── Data rows ─────────────────────────────────────────────────────────────
#     head_total_rows = []
#     head_counter    = 0

#     for head in data:
#         head_counter += 1
#         alpha_index = chr(64 + head_counter)
#         head_name   = (head.get("name") or "").strip().upper()

#         # ── COVID SUPPORT (special single-row treatment) ──────────────────────
#         if head_name == "COVID SUPPORT":
#             ws.append([])
#             item = head["items"][0] if head.get("items") else {}
#             r = ws.max_row + 1
#             ws.append([
#                 "", alpha_index,
#                 head["name"],
#                 item.get("name", ""),
#                 *item.get("q1", [0,0,0]),
#                 *item.get("q2", [0,0,0]),
#                 *item.get("q3", [0,0,0]),
#                 *item.get("q4", [0,0,0]),
#                 *_qtr_formulas(r),
#             ])
#             _style_row(ws, ws.max_row, font=_NORMAL)
#             _fmt_numeric(ws, ws.max_row)
#             ws.cell(row=ws.max_row, column=COL_HEAD).font = _BOLD
#             continue

#         # ── Section label row (A / B …) ───────────────────────────────────────
#         ws.append(["", alpha_index, head["name"]])
#         r_sec = ws.max_row
#         # Style all cells in merge span BEFORE merging
#         for col in range(COL_HEAD, COL_END + 1):
#             c = ws.cell(row=r_sec, column=col)
#             c.fill   = _FILL_HEAD
#             c.font   = _BOLD
#             c.border = _BORDER
#             c.alignment = _LEFT
#         ws.cell(row=r_sec, column=COL_SI).fill   = _FILL_HEAD
#         ws.cell(row=r_sec, column=COL_SI).border = _BORDER
#         ws.merge_cells(start_row=r_sec, start_column=COL_HEAD,
#                        end_row=r_sec,   end_column=COL_END)

#         if head_name == "OPERATING EXPENSES":
#             ws.append([])

#         sub_total_rows   = []
#         direct_item_rows = []

#         # ── Direct items (no sub-head) ────────────────────────────────────────
#         for item in head.get("items", []):
#             r = ws.max_row + 1
#             sub_val      = item.get("sub_head_of_expense") or ""
#             head_display = sub_val.strip()
#             ws.append([
#                 "", "",
#                 head_display,
#                 item["name"],
#                 *item.get("q1", [0,0,0]),
#                 *item.get("q2", [0,0,0]),
#                 *item.get("q3", [0,0,0]),
#                 *item.get("q4", [0,0,0]),
#                 *_qtr_formulas(r),
#             ])
#             _style_row(ws, ws.max_row, font=_NORMAL)
#             _fmt_numeric(ws, ws.max_row)
#             direct_item_rows.append(ws.max_row)

#         # ── Sub-heads (I / II …) ──────────────────────────────────────────────
#         sub_counter = 1
#         for sub in head.get("sub_heads", []):
#             roman_index = _to_roman(sub_counter)

#             ws.append(["", roman_index, sub["name"]])
#             r_sub = ws.max_row
#             # Style BEFORE merge
#             for col in range(COL_HEAD, COL_END + 1):
#                 c = ws.cell(row=r_sub, column=col)
#                 c.fill      = _FILL_SUBHEAD
#                 c.font      = _BOLD
#                 c.border    = _BORDER
#                 c.alignment = _LEFT
#             ws.cell(row=r_sub, column=COL_SI).fill   = _FILL_SUBHEAD
#             ws.cell(row=r_sub, column=COL_SI).border = _BORDER
#             ws.merge_cells(start_row=r_sub, start_column=COL_HEAD,
#                            end_row=r_sub,   end_column=COL_END)

#             sub_item_rows = []
#             for item in sub.get("items", []):
#                 r = ws.max_row + 1
#                 item_sub  = item.get("sub_head_of_expense") or ""
#                 sub_name  = sub.get("name") or ""
#                 head_display = ""
#                 cleaned = item_sub.strip()
#                 if cleaned and cleaned.lower() != sub_name.strip().lower():
#                     head_display = cleaned
#                 ws.append([
#                     "", "",
#                     head_display,
#                     item["name"],
#                     *item.get("q1", [0,0,0]),
#                     *item.get("q2", [0,0,0]),
#                     *item.get("q3", [0,0,0]),
#                     *item.get("q4", [0,0,0]),
#                     *_qtr_formulas(r),
#                 ])
#                 _style_row(ws, ws.max_row, font=_NORMAL)
#                 _fmt_numeric(ws, ws.max_row)
#                 sub_item_rows.append(ws.max_row)

#             if sub_item_rows:
#                 ws.append([
#                     "", "", "",
#                     f"TOTAL - {sub['name']}",
#                     *_totals_from_rows(sub_item_rows),
#                 ])
#                 _style_row(ws, ws.max_row, _FILL_SUBTOTAL, _BOLD)
#                 _fmt_numeric(ws, ws.max_row)
#                 sub_total_rows.append(ws.max_row)

#             sub_counter += 1

#         # ── Head total row ────────────────────────────────────────────────────
#         total_rows = sub_total_rows if sub_total_rows else direct_item_rows
#         if total_rows:
#             ws.append([
#                 "", "", "",
#                 f"TOTAL - {head['name']}",
#                 *_totals_from_rows(total_rows),
#             ])
#             _style_row(ws, ws.max_row, _FILL_HTOTAL, _BOLD)
#             _fmt_numeric(ws, ws.max_row)
#             head_total_rows.append(ws.max_row)

#             if head_name == "OPERATING EXPENSES":
#                 ws.append([])

#     # ── Clean up stray blank row before COVID SUPPORT ─────────────────────────
#     for r in range(ws.max_row, 1, -1):
#         if (ws.cell(r, COL_HEAD).value == "COVID SUPPORT" and
#                 ws.cell(r - 1, COL_HEAD).value is None):
#             ws.delete_rows(r - 1)
#             break

#     # ── Grand total ───────────────────────────────────────────────────────────
#     if head_total_rows:
#         ws.append([
#             "", "", "",
#             "GRAND TOTAL",
#             *_totals_from_rows(head_total_rows),
#         ])
#         _style_row(ws, ws.max_row, _FILL_GRAND, _BOLD)
#         _fmt_numeric(ws, ws.max_row)

#     _auto_col_widths(ws)
#     return ws


# # ═══════════════════════════════════════════════════════════════════════════════
# # SHEET BUILDER — Estimate Consolidated  (same structure as Annual)
# # ═══════════════════════════════════════════════════════════════════════════════

# def _sheet_estimate(wb, sheet_name, data, fy,
#                     org_name="Azim Premji Foundation"):
#     """
#     estimate data items carry Q1..Q4 as floats and a months dict.
#     We map months → q1/q2/q3/q4 lists so the same row-writing
#     logic as Annual works unchanged.
#     """
#     MONTH_MAP = {
#         "q1": ["4",  "5",  "6"],
#         "q2": ["7",  "8",  "9"],
#         "q3": ["10", "11", "12"],
#         "q4": ["1",  "2",  "3"],
#     }

#     def _to_qlist(obj):
#         """Convert months dict → {q1:[v,v,v], q2:…, q3:…, q4:…}."""
#         m = obj.get("months") or {}
#         out = {}
#         for qk, keys in MONTH_MAP.items():
#             out[qk] = [float(m.get(k, 0) or 0) for k in keys]
#         return out

#     def _normalise(obj):
#         """Add q1..q4 lists to an item/sub/head dict in-place."""
#         ql = _to_qlist(obj)
#         obj.update(ql)
#         for item in obj.get("items", []):
#             ql2 = _to_qlist(item)
#             item.update(ql2)
#         for sub in obj.get("sub_heads", []):
#             ql3 = _to_qlist(sub)
#             sub.update(ql3)
#             for item in sub.get("items", []):
#                 ql4 = _to_qlist(item)
#                 item.update(ql4)
#         return obj

#     normalised = [_normalise(dict(h)) for h in (data or [])]

#     ws = wb.create_sheet(title=sheet_name[:31])

#     # Row 1-2: title
#     ws.append(["", org_name])
#     ws.merge_cells("B1:U1")
#     ws["B1"].font      = Font(size=14, bold=True, name="Calibri")
#     ws["B1"].alignment = _LEFT

#     ws.append(["", f"Estimate for the Financial Year {fy or ''}"])
#     ws.merge_cells("B2:U2")
#     ws["B2"].font      = Font(size=12, bold=True, name="Calibri")
#     ws["B2"].alignment = _LEFT

#     ws.append([])

#     # Rows 4-5: headers — style then merge
#     ws.append([
#         "", "Sl #", "HEAD OF EXPENSE", "TYPE OF EXPENSE",
#         "QUARTER I","","","QUARTER II","","",
#         "QUARTER III","","","QUARTER IV","","",
#         "QTR-1","QTR-2","QTR-3","QTR-4", f"YEAR {fy}",
#     ])
#     r1 = ws.max_row
#     ws.append([
#         "", "Sl #", "HEAD OF EXPENSE", "TYPE OF EXPENSE",
#         "Apr","May","Jun","Jul","Aug","Sep",
#         "Oct","Nov","Dec","Jan","Feb","Mar",
#         "QTR-1","QTR-2","QTR-3","QTR-4", f"YEAR {fy}",
#     ])
#     r2 = ws.max_row

#     _style_row(ws, r1, _FILL_HDR, _WHITE_BOLD, is_header=True)
#     _style_row(ws, r2, _FILL_HDR, _WHITE_BOLD, is_header=True)

#     ws.merge_cells(start_row=r1, start_column=5,  end_row=r1, end_column=7)
#     ws.merge_cells(start_row=r1, start_column=8,  end_row=r1, end_column=10)
#     ws.merge_cells(start_row=r1, start_column=11, end_row=r1, end_column=13)
#     ws.merge_cells(start_row=r1, start_column=14, end_row=r1, end_column=16)
#     ws.merge_cells(start_row=r1, start_column=2,  end_row=r2, end_column=2)
#     ws.merge_cells(start_row=r1, start_column=3,  end_row=r2, end_column=3)
#     ws.merge_cells(start_row=r1, start_column=4,  end_row=r2, end_column=4)
#     for col in range(17, 22):
#         ws.merge_cells(start_row=r1, start_column=col, end_row=r2, end_column=col)

#     ws.freeze_panes = "E6"

#     # Reuse the same data-writing loop as Annual (data is now normalised)
#     head_total_rows = []
#     head_counter    = 0

#     for head in normalised:
#         head_counter += 1
#         alpha_index = chr(64 + head_counter)
#         head_name   = (head.get("name") or "").strip().upper()

#         if head_name == "COVID SUPPORT":
#             ws.append([])
#             item = head["items"][0] if head.get("items") else {}
#             r = ws.max_row + 1
#             ws.append([
#                 "", alpha_index,
#                 head["name"], item.get("name", ""),
#                 *item.get("q1",[0,0,0]), *item.get("q2",[0,0,0]),
#                 *item.get("q3",[0,0,0]), *item.get("q4",[0,0,0]),
#                 *_qtr_formulas(r),
#             ])
#             _style_row(ws, ws.max_row, font=_NORMAL)
#             _fmt_numeric(ws, ws.max_row)
#             ws.cell(row=ws.max_row, column=COL_HEAD).font = _BOLD
#             continue

#         ws.append(["", alpha_index, head["name"]])
#         r_sec = ws.max_row
#         for col in range(COL_HEAD, COL_END + 1):
#             c = ws.cell(row=r_sec, column=col)
#             c.fill = _FILL_HEAD; c.font = _BOLD
#             c.border = _BORDER;  c.alignment = _LEFT
#         ws.cell(row=r_sec, column=COL_SI).fill   = _FILL_HEAD
#         ws.cell(row=r_sec, column=COL_SI).border = _BORDER
#         ws.merge_cells(start_row=r_sec, start_column=COL_HEAD,
#                        end_row=r_sec,   end_column=COL_END)

#         if head_name == "OPERATING EXPENSES":
#             ws.append([])

#         sub_total_rows = []
#         direct_item_rows = []

#         for item in head.get("items", []):
#             r = ws.max_row + 1
#             sub_val = item.get("sub_head_of_expense") or ""
#             ws.append([
#                 "", "", sub_val.strip(), item["name"],
#                 *item.get("q1",[0,0,0]), *item.get("q2",[0,0,0]),
#                 *item.get("q3",[0,0,0]), *item.get("q4",[0,0,0]),
#                 *_qtr_formulas(r),
#             ])
#             _style_row(ws, ws.max_row, font=_NORMAL)
#             _fmt_numeric(ws, ws.max_row)
#             direct_item_rows.append(ws.max_row)

#         sub_counter = 1
#         for sub in head.get("sub_heads", []):
#             roman_index = _to_roman(sub_counter)
#             ws.append(["", roman_index, sub["name"]])
#             r_sub = ws.max_row
#             for col in range(COL_HEAD, COL_END + 1):
#                 c = ws.cell(row=r_sub, column=col)
#                 c.fill = _FILL_SUBHEAD; c.font = _BOLD
#                 c.border = _BORDER;     c.alignment = _LEFT
#             ws.cell(row=r_sub, column=COL_SI).fill   = _FILL_SUBHEAD
#             ws.cell(row=r_sub, column=COL_SI).border = _BORDER
#             ws.merge_cells(start_row=r_sub, start_column=COL_HEAD,
#                            end_row=r_sub,   end_column=COL_END)

#             sub_item_rows = []
#             for item in sub.get("items", []):
#                 r = ws.max_row + 1
#                 item_sub = (item.get("sub_head_of_expense") or "").strip()
#                 sub_name = (sub.get("name") or "").strip()
#                 head_display = item_sub if item_sub.lower() != sub_name.lower() else ""
#                 ws.append([
#                     "", "", head_display, item["name"],
#                     *item.get("q1",[0,0,0]), *item.get("q2",[0,0,0]),
#                     *item.get("q3",[0,0,0]), *item.get("q4",[0,0,0]),
#                     *_qtr_formulas(r),
#                 ])
#                 _style_row(ws, ws.max_row, font=_NORMAL)
#                 _fmt_numeric(ws, ws.max_row)
#                 sub_item_rows.append(ws.max_row)

#             if sub_item_rows:
#                 ws.append([
#                     "", "", "", f"TOTAL - {sub['name']}",
#                     *_totals_from_rows(sub_item_rows),
#                 ])
#                 _style_row(ws, ws.max_row, _FILL_SUBTOTAL, _BOLD)
#                 _fmt_numeric(ws, ws.max_row)
#                 sub_total_rows.append(ws.max_row)

#             sub_counter += 1

#         total_rows = sub_total_rows if sub_total_rows else direct_item_rows
#         if total_rows:
#             ws.append([
#                 "", "", "", f"TOTAL - {head['name']}",
#                 *_totals_from_rows(total_rows),
#             ])
#             _style_row(ws, ws.max_row, _FILL_HTOTAL, _BOLD)
#             _fmt_numeric(ws, ws.max_row)
#             head_total_rows.append(ws.max_row)
#             if head_name == "OPERATING EXPENSES":
#                 ws.append([])

#     for r in range(ws.max_row, 1, -1):
#         if (ws.cell(r, COL_HEAD).value == "COVID SUPPORT" and
#                 ws.cell(r-1, COL_HEAD).value is None):
#             ws.delete_rows(r - 1)
#             break

#     if head_total_rows:
#         ws.append([
#             "", "", "", "GRAND TOTAL",
#             *_totals_from_rows(head_total_rows),
#         ])
#         _style_row(ws, ws.max_row, _FILL_GRAND, _BOLD)
#         _fmt_numeric(ws, ws.max_row)

#     _auto_col_widths(ws)
#     return ws


# # ═══════════════════════════════════════════════════════════════════════════════
# # SHEET BUILDER — PPT / Foundation level
# # ═══════════════════════════════════════════════════════════════════════════════
# #
# # Columns: Unit | Budget Opex | Budget Capex | Budget Total | Est Opex | Est Capex | Est Total
# #

# def _sheet_ppt(wb, sheet_name, rows, budget_label, est_label,
#                fy="2025-26", org_name="Azim Premji Foundation"):
#     ws = wb.create_sheet(title=sheet_name[:31])

#     # Row 1-2
#     ws.append(["", org_name])
#     ws.merge_cells("B1:H1")
#     ws["B1"].font      = Font(size=14, bold=True, name="Calibri")
#     ws["B1"].alignment = _LEFT
#     ws.append(["", f"Foundation Level Metrics – {fy}"])
#     ws.merge_cells("B2:H2")
#     ws["B2"].font      = Font(size=12, bold=True, name="Calibri")
#     ws["B2"].alignment = _LEFT
#     ws.append([])

#     # Rows 4-5: headers — style BEFORE merge
#     for col in range(2, 9):
#         c4 = ws.cell(4, col)
#         c4.fill = _FILL_HDR; c4.font = _WHITE_BOLD
#         c4.border = _BORDER; c4.alignment = _CENTER
#         c5 = ws.cell(5, col)
#         c5.fill = _FILL_HDR; c5.font = _WHITE_BOLD
#         c5.border = _BORDER; c5.alignment = _CENTER

#     ws.cell(4, 2, "Unit")
#     ws.cell(4, 3, budget_label)
#     ws.cell(4, 6, est_label)
#     for col, lbl in zip(range(3, 9), ["Opex","Capex","Total","Opex","Capex","Total"]):
#         ws.cell(5, col, lbl)

#     ws.merge_cells(start_row=4, start_column=2, end_row=5, end_column=2)  # Unit
#     ws.merge_cells(start_row=4, start_column=3, end_row=4, end_column=5)  # Budget
#     ws.merge_cells(start_row=4, start_column=6, end_row=4, end_column=8)  # Estimate

#     thin = Side(style="thin")
#     border = Border(left=thin, right=thin, top=thin, bottom=thin)

#     for row in rows:
#         is_total = row.get("is_total", False)
#         label    = row.get("label", "")
#         bO = float(row.get("bOpex",  0) or 0)
#         bC = float(row.get("bCapex", 0) or 0)
#         eO = float(row.get("eOpex",  0) or 0)
#         eC = float(row.get("eCapex", 0) or 0)

#         ws.append(["", label, bO, bC, bO + bC, eO, eC, eO + eC])
#         r = ws.max_row
#         fill = _FILL_HTOTAL if is_total else _FILL_WHITE
#         font = _BOLD if is_total else _NORMAL
#         for col in range(2, 9):
#             cell = ws.cell(r, col)
#             cell.fill   = fill
#             cell.font   = font
#             cell.border = border
#             cell.alignment = _RIGHT if col > 2 else _LEFT
#             if col > 2:
#                 cell.number_format = NUM_FMT

#     _auto_col_widths(ws)
#     return ws


# # ═══════════════════════════════════════════════════════════════════════════════
# # SHEET BUILDER — Budget & Estimate
# # ═══════════════════════════════════════════════════════════════════════════════

# def _sheet_be(wb, sheet_name, be_data, fy, plan_label, est_label,
#               org_name="Azim Premji Foundation"):
#     ws = wb.create_sheet(title=sheet_name[:31])

#     entities   = be_data or []
#     n          = len(entities)
#     # Col layout: B=HEAD, C=TYPE, then 2 cols per entity, then 2 grand-total cols
#     # All 1-based; col B = 2
#     TOTAL_COLS = 1 + 2 + n * 2 + 2   # col A unused + HEAD + TYPE + entities + grand
#     grand_p_col = 2 + 2 + n * 2 + 1  # first grand total col (1-based)
#     grand_e_col = grand_p_col + 1

#     # Row 1-2
#     ws.append(["", org_name])
#     for col in range(2, grand_e_col + 1):
#         ws.cell(1, col)  # ensure cells exist
#     ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=grand_e_col)
#     ws["B1"].font      = Font(size=14, bold=True, name="Calibri")
#     ws["B1"].alignment = _LEFT

#     ws.append(["", f"Budget & Estimate – {fy}"])
#     ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=grand_e_col)
#     ws["B2"].font      = Font(size=12, bold=True, name="Calibri")
#     ws["B2"].alignment = _LEFT
#     ws.append([])

#     # Rows 4-5: headers — style ALL cells BEFORE any merges
#     for col in range(2, grand_e_col + 1):
#         for row in [4, 5]:
#             c = ws.cell(row, col)
#             c.fill = _FILL_HDR; c.font = _WHITE_BOLD
#             c.border = _BORDER; c.alignment = _CENTER

#     ws.cell(4, 2, "HEAD OF EXPENSE"); ws.cell(4, 3, "TYPE OF EXPENSE")
#     for ei, entity in enumerate(entities):
#         cs = 4 + ei * 2
#         ws.cell(4, cs, entity.get("label", ""))
#     ws.cell(4, grand_p_col, "GRAND TOTAL")

#     for ei in range(n):
#         cs = 4 + ei * 2
#         ws.cell(5, cs,     plan_label)
#         ws.cell(5, cs + 1, est_label)
#     ws.cell(5, grand_p_col,     plan_label)
#     ws.cell(5, grand_e_col,     est_label)

#     # Merges AFTER styling
#     ws.merge_cells(start_row=4, start_column=2, end_row=5, end_column=2)
#     ws.merge_cells(start_row=4, start_column=3, end_row=5, end_column=3)
#     for ei in range(n):
#         cs = 4 + ei * 2
#         ws.merge_cells(start_row=4, start_column=cs, end_row=4, end_column=cs + 1)
#     ws.merge_cells(start_row=4, start_column=grand_p_col, end_row=4, end_column=grand_e_col)

#     ws.freeze_panes = ws.cell(6, 4).coordinate

#     # ── Value helpers ─────────────────────────────────────────────────────────
#     PF = "ytd"; EF = "total_posted_amt_ytd"; IF_E = "total_posted_amt"

#     def _sec_v(entity, sname, f):
#         for s in entity.get("actuals", []):
#             if s.get("name") == sname:
#                 return float(s.get(f, 0) or 0)
#         return 0.0

#     def _sub_v(entity, sname, subname, f):
#         for s in entity.get("actuals", []):
#             if s.get("name") == sname:
#                 for sub in s.get("sub_heads", []):
#                     if sub.get("name") == subname:
#                         return float(sub.get(f, 0) or 0)
#         return 0.0

#     def _item_v(entity, iname, use_est=False):
#         f = IF_E if use_est else PF
#         for s in entity.get("actuals", []):
#             for item in s.get("items", []):
#                 if item.get("name") == iname:
#                     return float(item.get(f, 0) or 0)
#             for sub in s.get("sub_heads", []):
#                 for item in sub.get("items", []):
#                     if item.get("name") == iname:
#                         return float(item.get(f, 0) or 0)
#         return 0.0

#     def _grand_v(entity, f):
#         return sum(float(s.get(f, 0) or 0) for s in entity.get("actuals", []))

#     thin2  = Side(style="thin")
#     brd2   = Border(left=thin2, right=thin2, top=thin2, bottom=thin2)

#     def _write_be_row(head_val, type_val, vp_list, ve_list, fill, font):
#         ws.append(["", head_val, type_val] + [""] * (grand_e_col - 3))
#         r = ws.max_row
#         gp = ge = 0.0
#         for ei, (p, e) in enumerate(zip(vp_list, ve_list)):
#             cs = 4 + ei * 2
#             cp = ws.cell(r, cs);     cp.value = p
#             ce = ws.cell(r, cs + 1); ce.value = e
#             gp += p; ge += e
#         ws.cell(r, grand_p_col).value = gp
#         ws.cell(r, grand_e_col).value = ge
#         for col in range(2, grand_e_col + 1):
#             c = ws.cell(r, col)
#             c.fill = fill; c.font = font; c.border = brd2
#             c.alignment = _RIGHT if col > 3 else _LEFT
#             if col > 3:
#                 c.number_format = NUM_FMT

#     if not entities:
#         return ws

#     for sec in entities[0].get("actuals", []):
#         sname = sec.get("name", "")

#         # Section row
#         ws.append(["", sname])
#         r_sec = ws.max_row
#         for col in range(2, grand_e_col + 1):
#             c = ws.cell(r_sec, col)
#             c.fill = _FILL_HEAD; c.font = _BOLD
#             c.border = brd2;     c.alignment = _LEFT
#         ws.merge_cells(start_row=r_sec, start_column=2,
#                        end_row=r_sec,   end_column=grand_e_col)

#         for sub in sec.get("sub_heads", []):
#             subname = sub.get("name", "")
#             vp = [_sub_v(e, sname, subname, PF) for e in entities]
#             ve = [_sub_v(e, sname, subname, EF) for e in entities]
#             _write_be_row(sname, subname, vp, ve, _FILL_SUBHEAD, _BOLD)
#             for item in sub.get("items", []):
#                 iname = item.get("name", "")
#                 vp = [_item_v(e, iname, use_est=False) for e in entities]
#                 ve = [_item_v(e, iname, use_est=True)  for e in entities]
#                 _write_be_row("", iname, vp, ve, _FILL_WHITE, _NORMAL)

#         for item in sec.get("items", []):
#             iname = item.get("name", "")
#             vp = [_item_v(e, iname, use_est=False) for e in entities]
#             ve = [_item_v(e, iname, use_est=True)  for e in entities]
#             _write_be_row(sname, iname, vp, ve, _FILL_WHITE, _NORMAL)

#         vp = [_sec_v(e, sname, PF) for e in entities]
#         ve = [_sec_v(e, sname, EF) for e in entities]
#         _write_be_row(f"TOTAL - {sname}", "", vp, ve, _FILL_HTOTAL, _BOLD)

#     vp = [_grand_v(e, PF) for e in entities]
#     ve = [_grand_v(e, EF) for e in entities]
#     _write_be_row("GRAND TOTAL", "", vp, ve, _FILL_GRAND, _BOLD)

#     _auto_col_widths(ws)
#     return ws


# # ═══════════════════════════════════════════════════════════════════════════════
# # FY HELPERS
# # ═══════════════════════════════════════════════════════════════════════════════

# def _fy_labels(fy):
#     parts    = (fy or "2025-26").split("-")
#     start_yy = (parts[0] or "2025")[-2:]
#     end_yy   = (parts[1] if len(parts) > 1 else "26")[-2:]
#     prev_s   = str(int(start_yy) - 1).zfill(2)
#     prev_e   = str(int(end_yy)   - 1).zfill(2)
#     return {
#         "plan": f"FY{start_yy}-{end_yy} Plan",
#         "est":  f"FY{prev_s}-{prev_e} Estimate",
#     }

# def _prev_fy(fy):
#     parts = (fy or "2025-26").split("-")
#     s = int(parts[0] or 2025) - 1
#     e = int(parts[1] or 26)   - 1
#     return f"{s}-{str(e).zfill(2)}"


# # ═══════════════════════════════════════════════════════════════════════════════
# # WHITELISTED API ENDPOINTS
# # ═══════════════════════════════════════════════════════════════════════════════

# @frappe.whitelist()
# def export_ppt(financial_year, ppt_rows, prev_ppt_rows,
#                budget_label, est_label, prev_budget_label, prev_est_label):
#     fy      = financial_year or "2025-26"
#     prev_fy = _prev_fy(fy)
#     wb = Workbook(); wb.remove(wb.active)
#     _sheet_ppt(wb, f"Foundation Metrics ({fy})",
#                json.loads(ppt_rows), budget_label, est_label, fy=fy)
#     _sheet_ppt(wb, f"Foundation Metrics ({prev_fy})",
#                json.loads(prev_ppt_rows), prev_budget_label, prev_est_label, fy=prev_fy)
#     return {"filename": f"CB_Foundation_Metrics_{fy}.xlsx", "data": _wb_to_b64(wb)}


# @frappe.whitelist()
# def export_annual(financial_year, annual_data):
#     fy = financial_year or "2025-26"
#     wb = Workbook(); wb.remove(wb.active)
#     _sheet_annual(wb, "Annual Budget", json.loads(annual_data), fy)
#     return {"filename": f"CB_Annual_Budget_{fy}.xlsx", "data": _wb_to_b64(wb)}


# @frappe.whitelist()
# def export_estimate(financial_year, estimate_data):
#     fy = financial_year or "2025-26"
#     wb = Workbook(); wb.remove(wb.active)
#     _sheet_estimate(wb, "Estimate", json.loads(estimate_data), fy)
#     return {"filename": f"CB_Estimate_{fy}.xlsx", "data": _wb_to_b64(wb)}


# @frappe.whitelist()
# def export_budget_estimate(financial_year, be_data):
#     fy     = financial_year or "2025-26"
#     labels = _fy_labels(fy)
#     wb = Workbook(); wb.remove(wb.active)
#     _sheet_be(wb, "Budget & Estimate", json.loads(be_data),
#               fy, labels["plan"], labels["est"])
#     return {"filename": f"CB_BudgetEstimate_{fy}.xlsx", "data": _wb_to_b64(wb)}


# @frappe.whitelist()
# def export_all(financial_year, ppt_rows, prev_ppt_rows,
#                budget_label, est_label, prev_budget_label, prev_est_label,
#                annual_data, estimate_data, be_data):
#     fy      = financial_year or "2025-26"
#     prev_fy = _prev_fy(fy)
#     labels  = _fy_labels(fy)
#     wb = Workbook(); wb.remove(wb.active)
#     _sheet_ppt(wb, f"Foundation Metrics ({fy})",
#                json.loads(ppt_rows), budget_label, est_label, fy=fy)
#     _sheet_ppt(wb, f"Foundation Metrics ({prev_fy})",
#                json.loads(prev_ppt_rows), prev_budget_label, prev_est_label, fy=prev_fy)
#     _sheet_annual(wb,   "Annual Budget",     json.loads(annual_data),   fy)
#     _sheet_estimate(wb, "Estimate",          json.loads(estimate_data), fy)
#     _sheet_be(wb,       "Budget & Estimate", json.loads(be_data),
#               fy, labels["plan"], labels["est"])
#     return {"filename": f"CB_Consolidated_All_{fy}.xlsx", "data": _wb_to_b64(wb)}




# annual_budget/api/export_reports.py
#
# Server-side Excel export — design matches the screenshot:
#   Row 1       : Org name  (no fill, bold 14pt)
#   Row 2       : Subtitle  (no fill, bold 12pt)
#   Row 3       : Blank spacer
#   Row 4-5     : Quarter / month headers  (grey #5D6D7E, white bold)
#   Section rows: A / B … label  (grey #D6DBDF, bold, merged)
#   Sub-head    : I / II … label  (light grey #F2F3F4, bold, merged)
#   Item rows   : white, normal
#   Sub-total   : #EBF5FB, bold
#   Head total  : #D4E6F1, bold
#   Grand total : #A9CCE3, bold
#
# Numbers use Excel SUM() formulas exactly like the reference code.
# Requires:  bench pip install openpyxl
# Place at:  annual_budget/api/export_reports.py





































































# import io
# import base64
# import json

# import frappe
# from openpyxl import Workbook
# from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
# from openpyxl.utils import get_column_letter

# # ═══════════════════════════════════════════════════════════════════════════════
# # SHARED STYLE CONSTANTS  (matching reference code exactly)
# # ═══════════════════════════════════════════════════════════════════════════════

# _CENTER  = Alignment(horizontal="center", vertical="center")
# _LEFT    = Alignment(horizontal="left",   vertical="center")
# _RIGHT   = Alignment(horizontal="right",  vertical="center")
# _WRAP_L  = Alignment(horizontal="left",   vertical="center", wrap_text=True)

# _BOLD       = Font(bold=True, name="Calibri", size=9)
# _WHITE_BOLD = Font(bold=True, color="FFFFFF", name="Calibri", size=9)
# _NORMAL     = Font(name="Calibri", size=9)

# _FILL_HDR      = PatternFill("solid", fgColor="5D6D7E")   # dark blue-grey  — column headers
# _FILL_HEAD     = PatternFill("solid", fgColor="D6DBDF")   # medium grey     — section rows (A/B/…)
# _FILL_SUBHEAD  = PatternFill("solid", fgColor="F2F3F4")   # light grey      — sub-head rows (I/II/…)
# _FILL_SUBTOTAL = PatternFill("solid", fgColor="EBF5FB")   # pale blue       — sub-total rows
# _FILL_HTOTAL   = PatternFill("solid", fgColor="D4E6F1")   # mid blue        — head total rows
# _FILL_GRAND    = PatternFill("solid", fgColor="A9CCE3")   # stronger blue   — grand total
# _FILL_WHITE    = PatternFill("solid", fgColor="FFFFFF")

# _THIN   = Side(style="thin")
# _BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# NUM_FMT = "#,##0.00"

# # Column indices (1-based)
# COL_SI    = 2   # Sl #
# COL_HEAD  = 3   # HEAD OF EXPENSE
# COL_TYPE  = 4   # TYPE OF EXPENSE
# COL_START = 5   # Apr  (first data column)
# COL_END   = 21  # YEAR total  (last data column)
# # Cols 5-7  = Apr May Jun
# # Cols 8-10 = Jul Aug Sep
# # Cols 11-13= Oct Nov Dec
# # Cols 14-16= Jan Feb Mar
# # Cols 17-20= QTR-1 QTR-2 QTR-3 QTR-4
# # Col  21   = YEAR total


# # ═══════════════════════════════════════════════════════════════════════════════
# # LOW-LEVEL HELPERS
# # ═══════════════════════════════════════════════════════════════════════════════

# def _merge(ws, r1, c1, r2, c2):
#     ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)


# def _to_roman(num):
#     vals = [1000,900,500,400,100,90,50,40,10,9,5,4,1]
#     syms = ["M","CM","D","CD","C","XC","L","XL","X","IX","V","IV","I"]
#     out = ""
#     i = 0
#     while num > 0:
#         for _ in range(num // vals[i]):
#             out += syms[i]
#             num -= vals[i]
#         i += 1
#     return out


# def _sum_formula(col_letter, rows):
#     """Build =R1+R2+… formula from a list of row numbers."""
#     if not rows:
#         return 0
#     return "=" + "+".join(f"{col_letter}{r}" for r in rows)


# def _qtr_formulas(r):
#     """
#     Return the 5 formula strings for a data row r:
#       QTR-1 = SUM(E:G), QTR-2 = SUM(H:J), QTR-3 = SUM(K:M),
#       QTR-4 = SUM(N:P), YEAR  = SUM(Q:T)
#     """
#     return [
#         f"=SUM(E{r}:G{r})",
#         f"=SUM(H{r}:J{r})",
#         f"=SUM(K{r}:M{r})",
#         f"=SUM(N{r}:P{r})",
#         f"=SUM(Q{r}:T{r})",
#     ]


# def _build_formula(col_letter, rows):
#     """Same as reference code's build_formula."""
#     if not rows:
#         return 0
#     return "=" + "+".join(f"{col_letter}{r}" for r in rows)


# def _totals_from_rows(rows):
#     """Return 17 formula cells (E..U) that sum the given row list."""
#     return [_build_formula(c, rows)
#             for c in list("EFGHIJKLMNOPQRSTU")]


# def _style_row(ws, row, fill=None, font=None, is_header=False):
#     """Apply border, alignment, fill, font to cols 2-21 (matching reference)."""
#     for col in range(COL_SI, COL_END + 1):
#         cell = ws.cell(row=row, column=col)
#         cell.border = _BORDER
#         if is_header:
#             cell.alignment = _CENTER
#         else:
#             if col == COL_SI:
#                 cell.alignment = _CENTER
#             elif col in (COL_HEAD, COL_TYPE):
#                 cell.alignment = _WRAP_L
#             else:
#                 cell.alignment = _RIGHT
#         if fill:
#             cell.fill = fill
#         if font:
#             cell.font = font


# def _fmt_numeric(ws, row):
#     for col in range(COL_START, COL_END + 1):
#         ws.cell(row=row, column=col).number_format = NUM_FMT


# def _auto_col_widths(ws):
#     for col in range(1, ws.max_column + 1):
#         letter = get_column_letter(col)
#         max_len = 0
#         for row in range(1, ws.max_row + 1):
#             val = ws.cell(row=row, column=col).value
#             if val:
#                 s = str(val)
#                 if s.startswith("="):
#                     s = "999,999,999.00"
#                 max_len = max(max_len, len(s))
#         ws.column_dimensions[letter].width = min(max(max_len + 3, 10), 50)


# def _wb_to_b64(wb):
#     buf = io.BytesIO()
#     wb.save(buf)
#     buf.seek(0)
#     return base64.b64encode(buf.read()).decode("utf-8")


# # ═══════════════════════════════════════════════════════════════════════════════
# # SHEET BUILDER — Annual Budget  (directly mirrors reference code logic)
# # ═══════════════════════════════════════════════════════════════════════════════

# def _sheet_annual(wb, sheet_name, data, fy,
#                   org_name="Azim Premji Foundation"):
#     ws = wb.create_sheet(title=sheet_name[:31])

#     # ── Row 1: Org name ───────────────────────────────────────────────────────
#     ws.append(["", org_name])
#     ws.merge_cells("B1:U1")
#     ws["B1"].font      = Font(size=14, bold=True, name="Calibri")
#     ws["B1"].alignment = _LEFT

#     # ── Row 2: Subtitle ───────────────────────────────────────────────────────
#     ws.append(["", f"Budget for the Financial Year {fy or ''}"])
#     ws.merge_cells("B2:U2")
#     ws["B2"].font      = Font(size=12, bold=True, name="Calibri")
#     ws["B2"].alignment = _LEFT

#     # ── Row 3: Blank ──────────────────────────────────────────────────────────
#     ws.append([])

#     # ── Rows 4-5: Headers — style BEFORE merging ─────────────────────────────
#     ws.append([
#         "", "Sl #", "HEAD OF EXPENSE", "TYPE OF EXPENSE",
#         "QUARTER I",  "", "",
#         "QUARTER II", "", "",
#         "QUARTER III","", "",
#         "QUARTER IV", "", "",
#         "QTR-1","QTR-2","QTR-3","QTR-4",
#         f"YEAR {fy}",
#     ])
#     r1 = ws.max_row

#     ws.append([
#         "", "Sl #", "HEAD OF EXPENSE", "TYPE OF EXPENSE",
#         "Apr","May","Jun",
#         "Jul","Aug","Sep",
#         "Oct","Nov","Dec",
#         "Jan","Feb","Mar",
#         "QTR-1","QTR-2","QTR-3","QTR-4",
#         f"YEAR {fy}",
#     ])
#     r2 = ws.max_row

#     # Style both header rows FIRST
#     _style_row(ws, r1, _FILL_HDR, _WHITE_BOLD, is_header=True)
#     _style_row(ws, r2, _FILL_HDR, _WHITE_BOLD, is_header=True)

#     # Now merge (covered cells already styled)
#     ws.merge_cells(start_row=r1, start_column=5,  end_row=r1, end_column=7)
#     ws.merge_cells(start_row=r1, start_column=8,  end_row=r1, end_column=10)
#     ws.merge_cells(start_row=r1, start_column=11, end_row=r1, end_column=13)
#     ws.merge_cells(start_row=r1, start_column=14, end_row=r1, end_column=16)
#     ws.merge_cells(start_row=r1, start_column=2,  end_row=r2, end_column=2)
#     ws.merge_cells(start_row=r1, start_column=3,  end_row=r2, end_column=3)
#     ws.merge_cells(start_row=r1, start_column=4,  end_row=r2, end_column=4)
#     for col in range(17, 22):
#         ws.merge_cells(start_row=r1, start_column=col, end_row=r2, end_column=col)

#     ws.freeze_panes = "E6"

#     # ── Data rows ─────────────────────────────────────────────────────────────
#     head_total_rows = []
#     head_counter    = 0

#     for head in data:
#         head_counter += 1
#         alpha_index = chr(64 + head_counter)
#         head_name   = (head.get("name") or "").strip().upper()

#         # ── COVID SUPPORT (special single-row treatment) ──────────────────────
#         if head_name == "COVID SUPPORT":
#             ws.append([])
#             item = head["items"][0] if head.get("items") else {}
#             r = ws.max_row + 1
#             ws.append([
#                 "", alpha_index,
#                 head["name"],
#                 item.get("name", ""),
#                 *item.get("q1", [0,0,0]),
#                 *item.get("q2", [0,0,0]),
#                 *item.get("q3", [0,0,0]),
#                 *item.get("q4", [0,0,0]),
#                 *_qtr_formulas(r),
#             ])
#             _style_row(ws, ws.max_row, font=_NORMAL)
#             _fmt_numeric(ws, ws.max_row)
#             ws.cell(row=ws.max_row, column=COL_HEAD).font = _BOLD
#             continue

#         # ── Section label row (A / B …) ───────────────────────────────────────
#         ws.append(["", alpha_index, head["name"]])
#         r_sec = ws.max_row
#         # Style all cells in merge span BEFORE merging
#         for col in range(COL_HEAD, COL_END + 1):
#             c = ws.cell(row=r_sec, column=col)
#             c.fill   = _FILL_HEAD
#             c.font   = _BOLD
#             c.border = _BORDER
#             c.alignment = _LEFT
#         ws.cell(row=r_sec, column=COL_SI).fill   = _FILL_HEAD
#         ws.cell(row=r_sec, column=COL_SI).border = _BORDER
#         ws.merge_cells(start_row=r_sec, start_column=COL_HEAD,
#                        end_row=r_sec,   end_column=COL_END)

#         if head_name == "OPERATING EXPENSES":
#             ws.append([])

#         sub_total_rows   = []
#         direct_item_rows = []

#         # ── Direct items (no sub-head) ────────────────────────────────────────
#         for item in head.get("items", []):
#             r = ws.max_row + 1
#             sub_val      = item.get("sub_head_of_expense") or ""
#             head_display = sub_val.strip()
#             ws.append([
#                 "", "",
#                 head_display,
#                 item["name"],
#                 *item.get("q1", [0,0,0]),
#                 *item.get("q2", [0,0,0]),
#                 *item.get("q3", [0,0,0]),
#                 *item.get("q4", [0,0,0]),
#                 *_qtr_formulas(r),
#             ])
#             _style_row(ws, ws.max_row, font=_NORMAL)
#             _fmt_numeric(ws, ws.max_row)
#             direct_item_rows.append(ws.max_row)

#         # ── Sub-heads (I / II …) ──────────────────────────────────────────────
#         sub_counter = 1
#         for sub in head.get("sub_heads", []):
#             roman_index = _to_roman(sub_counter)

#             ws.append(["", roman_index, sub["name"]])
#             r_sub = ws.max_row
#             # Style BEFORE merge
#             for col in range(COL_HEAD, COL_END + 1):
#                 c = ws.cell(row=r_sub, column=col)
#                 c.fill      = _FILL_SUBHEAD
#                 c.font      = _BOLD
#                 c.border    = _BORDER
#                 c.alignment = _LEFT
#             ws.cell(row=r_sub, column=COL_SI).fill   = _FILL_SUBHEAD
#             ws.cell(row=r_sub, column=COL_SI).border = _BORDER
#             ws.merge_cells(start_row=r_sub, start_column=COL_HEAD,
#                            end_row=r_sub,   end_column=COL_END)

#             sub_item_rows = []
#             for item in sub.get("items", []):
#                 r = ws.max_row + 1
#                 item_sub  = item.get("sub_head_of_expense") or ""
#                 sub_name  = sub.get("name") or ""
#                 head_display = ""
#                 cleaned = item_sub.strip()
#                 if cleaned and cleaned.lower() != sub_name.strip().lower():
#                     head_display = cleaned
#                 ws.append([
#                     "", "",
#                     head_display,
#                     item["name"],
#                     *item.get("q1", [0,0,0]),
#                     *item.get("q2", [0,0,0]),
#                     *item.get("q3", [0,0,0]),
#                     *item.get("q4", [0,0,0]),
#                     *_qtr_formulas(r),
#                 ])
#                 _style_row(ws, ws.max_row, font=_NORMAL)
#                 _fmt_numeric(ws, ws.max_row)
#                 sub_item_rows.append(ws.max_row)

#             if sub_item_rows:
#                 ws.append([
#                     "", "", "",
#                     f"TOTAL - {sub['name']}",
#                     *_totals_from_rows(sub_item_rows),
#                 ])
#                 _style_row(ws, ws.max_row, _FILL_SUBTOTAL, _BOLD)
#                 _fmt_numeric(ws, ws.max_row)
#                 sub_total_rows.append(ws.max_row)

#             sub_counter += 1

#         # ── Head total row ────────────────────────────────────────────────────
#         total_rows = sub_total_rows if sub_total_rows else direct_item_rows
#         if total_rows:
#             ws.append([
#                 "", "", "",
#                 f"TOTAL - {head['name']}",
#                 *_totals_from_rows(total_rows),
#             ])
#             _style_row(ws, ws.max_row, _FILL_HTOTAL, _BOLD)
#             _fmt_numeric(ws, ws.max_row)
#             head_total_rows.append(ws.max_row)

#             if head_name == "OPERATING EXPENSES":
#                 ws.append([])

#     # ── Clean up stray blank row before COVID SUPPORT ─────────────────────────
#     for r in range(ws.max_row, 1, -1):
#         if (ws.cell(r, COL_HEAD).value == "COVID SUPPORT" and
#                 ws.cell(r - 1, COL_HEAD).value is None):
#             ws.delete_rows(r - 1)
#             break

#     # ── Grand total ───────────────────────────────────────────────────────────
#     if head_total_rows:
#         ws.append([
#             "", "", "",
#             "GRAND TOTAL",
#             *_totals_from_rows(head_total_rows),
#         ])
#         _style_row(ws, ws.max_row, _FILL_GRAND, _BOLD)
#         _fmt_numeric(ws, ws.max_row)

#     _auto_col_widths(ws)
#     return ws


# # ═══════════════════════════════════════════════════════════════════════════════
# # SHEET BUILDER — Estimate Consolidated  (same structure as Annual)
# # ═══════════════════════════════════════════════════════════════════════════════

# def _sheet_estimate(wb, sheet_name, data, fy,
#                     org_name="Azim Premji Foundation"):
#     """
#     estimate data items carry Q1..Q4 as floats and a months dict.
#     We map months → q1/q2/q3/q4 lists so the same row-writing
#     logic as Annual works unchanged.
#     """
#     MONTH_MAP = {
#         "q1": ["4",  "5",  "6"],
#         "q2": ["7",  "8",  "9"],
#         "q3": ["10", "11", "12"],
#         "q4": ["1",  "2",  "3"],
#     }

#     def _to_qlist(obj):
#         """Convert months dict → {q1:[v,v,v], q2:…, q3:…, q4:…}."""
#         m = obj.get("months") or {}
#         out = {}
#         for qk, keys in MONTH_MAP.items():
#             out[qk] = [float(m.get(k, 0) or 0) for k in keys]
#         return out

#     def _normalise(obj):
#         """Add q1..q4 lists to an item/sub/head dict in-place."""
#         ql = _to_qlist(obj)
#         obj.update(ql)
#         for item in obj.get("items", []):
#             ql2 = _to_qlist(item)
#             item.update(ql2)
#         for sub in obj.get("sub_heads", []):
#             ql3 = _to_qlist(sub)
#             sub.update(ql3)
#             for item in sub.get("items", []):
#                 ql4 = _to_qlist(item)
#                 item.update(ql4)
#         return obj

#     normalised = [_normalise(dict(h)) for h in (data or [])]

#     ws = wb.create_sheet(title=sheet_name[:31])

#     # Row 1-2: title
#     ws.append(["", org_name])
#     ws.merge_cells("B1:U1")
#     ws["B1"].font      = Font(size=14, bold=True, name="Calibri")
#     ws["B1"].alignment = _LEFT

#     ws.append(["", f"Estimate for the Financial Year {fy or ''}"])
#     ws.merge_cells("B2:U2")
#     ws["B2"].font      = Font(size=12, bold=True, name="Calibri")
#     ws["B2"].alignment = _LEFT

#     ws.append([])

#     # Rows 4-5: headers — style then merge
#     ws.append([
#         "", "Sl #", "HEAD OF EXPENSE", "TYPE OF EXPENSE",
#         "QUARTER I","","","QUARTER II","","",
#         "QUARTER III","","","QUARTER IV","","",
#         "QTR-1","QTR-2","QTR-3","QTR-4", f"YEAR {fy}",
#     ])
#     r1 = ws.max_row
#     ws.append([
#         "", "Sl #", "HEAD OF EXPENSE", "TYPE OF EXPENSE",
#         "Apr","May","Jun","Jul","Aug","Sep",
#         "Oct","Nov","Dec","Jan","Feb","Mar",
#         "QTR-1","QTR-2","QTR-3","QTR-4", f"YEAR {fy}",
#     ])
#     r2 = ws.max_row

#     _style_row(ws, r1, _FILL_HDR, _WHITE_BOLD, is_header=True)
#     _style_row(ws, r2, _FILL_HDR, _WHITE_BOLD, is_header=True)

#     ws.merge_cells(start_row=r1, start_column=5,  end_row=r1, end_column=7)
#     ws.merge_cells(start_row=r1, start_column=8,  end_row=r1, end_column=10)
#     ws.merge_cells(start_row=r1, start_column=11, end_row=r1, end_column=13)
#     ws.merge_cells(start_row=r1, start_column=14, end_row=r1, end_column=16)
#     ws.merge_cells(start_row=r1, start_column=2,  end_row=r2, end_column=2)
#     ws.merge_cells(start_row=r1, start_column=3,  end_row=r2, end_column=3)
#     ws.merge_cells(start_row=r1, start_column=4,  end_row=r2, end_column=4)
#     for col in range(17, 22):
#         ws.merge_cells(start_row=r1, start_column=col, end_row=r2, end_column=col)

#     ws.freeze_panes = "E6"

#     # Reuse the same data-writing loop as Annual (data is now normalised)
#     head_total_rows = []
#     head_counter    = 0

#     for head in normalised:
#         head_counter += 1
#         alpha_index = chr(64 + head_counter)
#         head_name   = (head.get("name") or "").strip().upper()

#         if head_name == "COVID SUPPORT":
#             ws.append([])
#             item = head["items"][0] if head.get("items") else {}
#             r = ws.max_row + 1
#             ws.append([
#                 "", alpha_index,
#                 head["name"], item.get("name", ""),
#                 *item.get("q1",[0,0,0]), *item.get("q2",[0,0,0]),
#                 *item.get("q3",[0,0,0]), *item.get("q4",[0,0,0]),
#                 *_qtr_formulas(r),
#             ])
#             _style_row(ws, ws.max_row, font=_NORMAL)
#             _fmt_numeric(ws, ws.max_row)
#             ws.cell(row=ws.max_row, column=COL_HEAD).font = _BOLD
#             continue

#         ws.append(["", alpha_index, head["name"]])
#         r_sec = ws.max_row
#         for col in range(COL_HEAD, COL_END + 1):
#             c = ws.cell(row=r_sec, column=col)
#             c.fill = _FILL_HEAD; c.font = _BOLD
#             c.border = _BORDER;  c.alignment = _LEFT
#         ws.cell(row=r_sec, column=COL_SI).fill   = _FILL_HEAD
#         ws.cell(row=r_sec, column=COL_SI).border = _BORDER
#         ws.merge_cells(start_row=r_sec, start_column=COL_HEAD,
#                        end_row=r_sec,   end_column=COL_END)

#         if head_name == "OPERATING EXPENSES":
#             ws.append([])

#         sub_total_rows = []
#         direct_item_rows = []

#         for item in head.get("items", []):
#             r = ws.max_row + 1
#             sub_val = item.get("sub_head_of_expense") or ""
#             ws.append([
#                 "", "", sub_val.strip(), item["name"],
#                 *item.get("q1",[0,0,0]), *item.get("q2",[0,0,0]),
#                 *item.get("q3",[0,0,0]), *item.get("q4",[0,0,0]),
#                 *_qtr_formulas(r),
#             ])
#             _style_row(ws, ws.max_row, font=_NORMAL)
#             _fmt_numeric(ws, ws.max_row)
#             direct_item_rows.append(ws.max_row)

#         sub_counter = 1
#         for sub in head.get("sub_heads", []):
#             roman_index = _to_roman(sub_counter)
#             ws.append(["", roman_index, sub["name"]])
#             r_sub = ws.max_row
#             for col in range(COL_HEAD, COL_END + 1):
#                 c = ws.cell(row=r_sub, column=col)
#                 c.fill = _FILL_SUBHEAD; c.font = _BOLD
#                 c.border = _BORDER;     c.alignment = _LEFT
#             ws.cell(row=r_sub, column=COL_SI).fill   = _FILL_SUBHEAD
#             ws.cell(row=r_sub, column=COL_SI).border = _BORDER
#             ws.merge_cells(start_row=r_sub, start_column=COL_HEAD,
#                            end_row=r_sub,   end_column=COL_END)

#             sub_item_rows = []
#             for item in sub.get("items", []):
#                 r = ws.max_row + 1
#                 item_sub = (item.get("sub_head_of_expense") or "").strip()
#                 sub_name = (sub.get("name") or "").strip()
#                 head_display = item_sub if item_sub.lower() != sub_name.lower() else ""
#                 ws.append([
#                     "", "", head_display, item["name"],
#                     *item.get("q1",[0,0,0]), *item.get("q2",[0,0,0]),
#                     *item.get("q3",[0,0,0]), *item.get("q4",[0,0,0]),
#                     *_qtr_formulas(r),
#                 ])
#                 _style_row(ws, ws.max_row, font=_NORMAL)
#                 _fmt_numeric(ws, ws.max_row)
#                 sub_item_rows.append(ws.max_row)

#             if sub_item_rows:
#                 ws.append([
#                     "", "", "", f"TOTAL - {sub['name']}",
#                     *_totals_from_rows(sub_item_rows),
#                 ])
#                 _style_row(ws, ws.max_row, _FILL_SUBTOTAL, _BOLD)
#                 _fmt_numeric(ws, ws.max_row)
#                 sub_total_rows.append(ws.max_row)

#             sub_counter += 1

#         total_rows = sub_total_rows if sub_total_rows else direct_item_rows
#         if total_rows:
#             ws.append([
#                 "", "", "", f"TOTAL - {head['name']}",
#                 *_totals_from_rows(total_rows),
#             ])
#             _style_row(ws, ws.max_row, _FILL_HTOTAL, _BOLD)
#             _fmt_numeric(ws, ws.max_row)
#             head_total_rows.append(ws.max_row)
#             if head_name == "OPERATING EXPENSES":
#                 ws.append([])

#     for r in range(ws.max_row, 1, -1):
#         if (ws.cell(r, COL_HEAD).value == "COVID SUPPORT" and
#                 ws.cell(r-1, COL_HEAD).value is None):
#             ws.delete_rows(r - 1)
#             break

#     if head_total_rows:
#         ws.append([
#             "", "", "", "GRAND TOTAL",
#             *_totals_from_rows(head_total_rows),
#         ])
#         _style_row(ws, ws.max_row, _FILL_GRAND, _BOLD)
#         _fmt_numeric(ws, ws.max_row)

#     _auto_col_widths(ws)
#     return ws


# # ═══════════════════════════════════════════════════════════════════════════════
# # SHEET BUILDER — PPT / Foundation level
# # ═══════════════════════════════════════════════════════════════════════════════
# #
# # Columns: Unit | Budget Opex | Budget Capex | Budget Total | Est Opex | Est Capex | Est Total
# #

# def _sheet_ppt_combined(wb, sheet_name,
#                         rows,      budget_label,      est_label,
#                         prev_rows, prev_budget_label, prev_est_label,
#                         fy="2025-26", prev_fy="2024-25",
#                         org_name="Azim Premji Foundation"):
#     """Both PPT tables (current FY + previous FY) in a single sheet."""
#     ws = wb.create_sheet(title=sheet_name[:31])

#     def _write_block(ws, start_row, block_rows, b_label, e_label, title):
#         # Title row
#         ws.cell(start_row, 2, title)
#         for col in range(2, 9):
#             c = ws.cell(start_row, col)
#             c.fill = _FILL_HEAD; c.font = _BOLD
#             c.border = _BORDER;  c.alignment = _LEFT
#         ws.merge_cells(start_row=start_row, start_column=2,
#                        end_row=start_row,   end_column=8)

#         h1 = start_row + 1
#         h2 = start_row + 2

#         # Style all header cells BEFORE merging
#         for col in range(2, 9):
#             for hr in [h1, h2]:
#                 c = ws.cell(hr, col)
#                 c.fill = _FILL_HDR; c.font = _WHITE_BOLD
#                 c.border = _BORDER; c.alignment = _CENTER

#         ws.cell(h1, 2, "Unit")
#         ws.cell(h1, 3, b_label)
#         ws.cell(h1, 6, e_label)
#         for col, lbl in zip(range(3, 9), ["Opex","Capex","Total","Opex","Capex","Total"]):
#             ws.cell(h2, col, lbl)

#         ws.merge_cells(start_row=h1, start_column=2, end_row=h2, end_column=2)
#         ws.merge_cells(start_row=h1, start_column=3, end_row=h1, end_column=5)
#         ws.merge_cells(start_row=h1, start_column=6, end_row=h1, end_column=8)

#         # Data rows
#         next_row = h2 + 1
#         thin = Side(style="thin")
#         brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
#         for row in block_rows:
#             is_total = row.get("is_total", False)
#             label    = row.get("label", "")
#             bO = float(row.get("bOpex",  0) or 0)
#             bC = float(row.get("bCapex", 0) or 0)
#             eO = float(row.get("eOpex",  0) or 0)
#             eC = float(row.get("eCapex", 0) or 0)
#             fill = _FILL_HTOTAL if is_total else _FILL_WHITE
#             font = _BOLD        if is_total else _NORMAL
#             for col, val in enumerate([label, bO, bC, bO+bC, eO, eC, eO+eC], start=2):
#                 c = ws.cell(next_row, col, val)
#                 c.fill = fill; c.font = font; c.border = brd
#                 c.alignment = _RIGHT if col > 2 else _LEFT
#                 if col > 2:
#                     c.number_format = NUM_FMT
#             next_row += 1

#         return next_row  # first row after this block

#     # Org name + subtitle
#     ws.append(["", org_name])
#     ws.merge_cells("B1:H1")
#     ws["B1"].font = Font(size=14, bold=True, name="Calibri")
#     ws["B1"].alignment = _LEFT
#     ws.append(["", f"Foundation Level Metrics – {fy}"])
#     ws.merge_cells("B2:H2")
#     ws["B2"].font = Font(size=12, bold=True, name="Calibri")
#     ws["B2"].alignment = _LEFT
#     ws.append([])  # row 3 blank

#     # Ensure enough rows exist before writing by pre-filling
#     # (openpyxl creates rows on ws.cell() access, so just call _write_block)
#     next_r = _write_block(ws, 4, rows, budget_label, est_label,
#                           f"Overall Foundation – Budget vs. Estimate ({fy})")

#     # Blank separator row between the two blocks
#     next_r += 1

#     _write_block(ws, next_r, prev_rows, prev_budget_label, prev_est_label,
#                  f"Overall Foundation – Previous Year Budget vs. Estimate ({prev_fy})")

#     _auto_col_widths(ws)
#     return ws


# def _sheet_ppt(wb, sheet_name, rows, budget_label, est_label,
#                fy="2025-26", org_name="Azim Premji Foundation"):
#     ws = wb.create_sheet(title=sheet_name[:31])

#     # Row 1-2
#     ws.append(["", org_name])
#     ws.merge_cells("B1:H1")
#     ws["B1"].font      = Font(size=14, bold=True, name="Calibri")
#     ws["B1"].alignment = _LEFT
#     ws.append(["", f"Foundation Level Metrics – {fy}"])
#     ws.merge_cells("B2:H2")
#     ws["B2"].font      = Font(size=12, bold=True, name="Calibri")
#     ws["B2"].alignment = _LEFT
#     ws.append([])

#     # Rows 4-5: headers — style BEFORE merge
#     for col in range(2, 9):
#         c4 = ws.cell(4, col)
#         c4.fill = _FILL_HDR; c4.font = _WHITE_BOLD
#         c4.border = _BORDER; c4.alignment = _CENTER
#         c5 = ws.cell(5, col)
#         c5.fill = _FILL_HDR; c5.font = _WHITE_BOLD
#         c5.border = _BORDER; c5.alignment = _CENTER

#     ws.cell(4, 2, "Unit")
#     ws.cell(4, 3, budget_label)
#     ws.cell(4, 6, est_label)
#     for col, lbl in zip(range(3, 9), ["Opex","Capex","Total","Opex","Capex","Total"]):
#         ws.cell(5, col, lbl)

#     ws.merge_cells(start_row=4, start_column=2, end_row=5, end_column=2)  # Unit
#     ws.merge_cells(start_row=4, start_column=3, end_row=4, end_column=5)  # Budget
#     ws.merge_cells(start_row=4, start_column=6, end_row=4, end_column=8)  # Estimate

#     thin = Side(style="thin")
#     border = Border(left=thin, right=thin, top=thin, bottom=thin)

#     for row in rows:
#         is_total = row.get("is_total", False)
#         label    = row.get("label", "")
#         bO = float(row.get("bOpex",  0) or 0)
#         bC = float(row.get("bCapex", 0) or 0)
#         eO = float(row.get("eOpex",  0) or 0)
#         eC = float(row.get("eCapex", 0) or 0)

#         ws.append(["", label, bO, bC, bO + bC, eO, eC, eO + eC])
#         r = ws.max_row
#         fill = _FILL_HTOTAL if is_total else _FILL_WHITE
#         font = _BOLD if is_total else _NORMAL
#         for col in range(2, 9):
#             cell = ws.cell(r, col)
#             cell.fill   = fill
#             cell.font   = font
#             cell.border = border
#             cell.alignment = _RIGHT if col > 2 else _LEFT
#             if col > 2:
#                 cell.number_format = NUM_FMT

#     _auto_col_widths(ws)
#     return ws


# # ═══════════════════════════════════════════════════════════════════════════════
# # SHEET BUILDER — Budget & Estimate
# # ═══════════════════════════════════════════════════════════════════════════════

# def _sheet_be(wb, sheet_name, be_data, fy, plan_label, est_label,
#               org_name="Azim Premji Foundation"):
#     ws = wb.create_sheet(title=sheet_name[:31])

#     entities   = be_data or []
#     n          = len(entities)
#     # Col layout: B=HEAD, C=TYPE, then 2 cols per entity, then 2 grand-total cols
#     # All 1-based; col B = 2
#     TOTAL_COLS = 1 + 2 + n * 2 + 2   # col A unused + HEAD + TYPE + entities + grand
#     grand_p_col = 2 + 2 + n * 2 + 1  # first grand total col (1-based)
#     grand_e_col = grand_p_col + 1

#     # Row 1-2
#     ws.append(["", org_name])
#     for col in range(2, grand_e_col + 1):
#         ws.cell(1, col)  # ensure cells exist
#     ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=grand_e_col)
#     ws["B1"].font      = Font(size=14, bold=True, name="Calibri")
#     ws["B1"].alignment = _LEFT

#     ws.append(["", f"Budget & Estimate – {fy}"])
#     ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=grand_e_col)
#     ws["B2"].font      = Font(size=12, bold=True, name="Calibri")
#     ws["B2"].alignment = _LEFT
#     ws.append([])

#     # Rows 4-5: headers — style ALL cells BEFORE any merges
#     for col in range(2, grand_e_col + 1):
#         for row in [4, 5]:
#             c = ws.cell(row, col)
#             c.fill = _FILL_HDR; c.font = _WHITE_BOLD
#             c.border = _BORDER; c.alignment = _CENTER

#     ws.cell(4, 2, "HEAD OF EXPENSE"); ws.cell(4, 3, "TYPE OF EXPENSE")
#     for ei, entity in enumerate(entities):
#         cs = 4 + ei * 2
#         ws.cell(4, cs, entity.get("label", ""))
#     ws.cell(4, grand_p_col, "GRAND TOTAL")

#     for ei in range(n):
#         cs = 4 + ei * 2
#         ws.cell(5, cs,     plan_label)
#         ws.cell(5, cs + 1, est_label)
#     ws.cell(5, grand_p_col,     plan_label)
#     ws.cell(5, grand_e_col,     est_label)

#     # Merges AFTER styling
#     ws.merge_cells(start_row=4, start_column=2, end_row=5, end_column=2)
#     ws.merge_cells(start_row=4, start_column=3, end_row=5, end_column=3)
#     for ei in range(n):
#         cs = 4 + ei * 2
#         ws.merge_cells(start_row=4, start_column=cs, end_row=4, end_column=cs + 1)
#     ws.merge_cells(start_row=4, start_column=grand_p_col, end_row=4, end_column=grand_e_col)

#     ws.freeze_panes = ws.cell(6, 4).coordinate

#     # ── Value helpers ─────────────────────────────────────────────────────────
#     PF = "ytd"; EF = "total_posted_amt_ytd"; IF_E = "total_posted_amt"

#     def _sec_v(entity, sname, f):
#         for s in entity.get("actuals", []):
#             if s.get("name") == sname:
#                 return float(s.get(f, 0) or 0)
#         return 0.0

#     def _sub_v(entity, sname, subname, f):
#         for s in entity.get("actuals", []):
#             if s.get("name") == sname:
#                 for sub in s.get("sub_heads", []):
#                     if sub.get("name") == subname:
#                         return float(sub.get(f, 0) or 0)
#         return 0.0

#     def _item_v(entity, iname, use_est=False):
#         f = IF_E if use_est else PF
#         for s in entity.get("actuals", []):
#             for item in s.get("items", []):
#                 if item.get("name") == iname:
#                     return float(item.get(f, 0) or 0)
#             for sub in s.get("sub_heads", []):
#                 for item in sub.get("items", []):
#                     if item.get("name") == iname:
#                         return float(item.get(f, 0) or 0)
#         return 0.0

#     def _grand_v(entity, f):
#         return sum(float(s.get(f, 0) or 0) for s in entity.get("actuals", []))

#     thin2  = Side(style="thin")
#     brd2   = Border(left=thin2, right=thin2, top=thin2, bottom=thin2)

#     def _write_be_row(head_val, type_val, vp_list, ve_list, fill, font):
#         ws.append(["", head_val, type_val] + [""] * (grand_e_col - 3))
#         r = ws.max_row
#         gp = ge = 0.0
#         for ei, (p, e) in enumerate(zip(vp_list, ve_list)):
#             cs = 4 + ei * 2
#             cp = ws.cell(r, cs);     cp.value = p
#             ce = ws.cell(r, cs + 1); ce.value = e
#             gp += p; ge += e
#         ws.cell(r, grand_p_col).value = gp
#         ws.cell(r, grand_e_col).value = ge
#         for col in range(2, grand_e_col + 1):
#             c = ws.cell(r, col)
#             c.fill = fill; c.font = font; c.border = brd2
#             c.alignment = _RIGHT if col > 3 else _LEFT
#             if col > 3:
#                 c.number_format = NUM_FMT

#     if not entities:
#         return ws

#     for sec in entities[0].get("actuals", []):
#         sname = sec.get("name", "")

#         # Section row
#         ws.append(["", sname])
#         r_sec = ws.max_row
#         for col in range(2, grand_e_col + 1):
#             c = ws.cell(r_sec, col)
#             c.fill = _FILL_HEAD; c.font = _BOLD
#             c.border = brd2;     c.alignment = _LEFT
#         ws.merge_cells(start_row=r_sec, start_column=2,
#                        end_row=r_sec,   end_column=grand_e_col)

#         for sub in sec.get("sub_heads", []):
#             subname = sub.get("name", "")
#             vp = [_sub_v(e, sname, subname, PF) for e in entities]
#             ve = [_sub_v(e, sname, subname, EF) for e in entities]
#             _write_be_row(sname, subname, vp, ve, _FILL_SUBHEAD, _BOLD)
#             for item in sub.get("items", []):
#                 iname = item.get("name", "")
#                 vp = [_item_v(e, iname, use_est=False) for e in entities]
#                 ve = [_item_v(e, iname, use_est=True)  for e in entities]
#                 _write_be_row("", iname, vp, ve, _FILL_WHITE, _NORMAL)

#         for item in sec.get("items", []):
#             iname = item.get("name", "")
#             vp = [_item_v(e, iname, use_est=False) for e in entities]
#             ve = [_item_v(e, iname, use_est=True)  for e in entities]
#             _write_be_row(sname, iname, vp, ve, _FILL_WHITE, _NORMAL)

#         vp = [_sec_v(e, sname, PF) for e in entities]
#         ve = [_sec_v(e, sname, EF) for e in entities]
#         _write_be_row(f"TOTAL - {sname}", "", vp, ve, _FILL_HTOTAL, _BOLD)

#     vp = [_grand_v(e, PF) for e in entities]
#     ve = [_grand_v(e, EF) for e in entities]
#     _write_be_row("GRAND TOTAL", "", vp, ve, _FILL_GRAND, _BOLD)

#     _auto_col_widths(ws)
#     return ws


# # ═══════════════════════════════════════════════════════════════════════════════
# # FY HELPERS
# # ═══════════════════════════════════════════════════════════════════════════════

# def _fy_labels(fy):
#     parts    = (fy or "2025-26").split("-")
#     start_yy = (parts[0] or "2025")[-2:]
#     end_yy   = (parts[1] if len(parts) > 1 else "26")[-2:]
#     prev_s   = str(int(start_yy) - 1).zfill(2)
#     prev_e   = str(int(end_yy)   - 1).zfill(2)
#     return {
#         "plan": f"FY{start_yy}-{end_yy} Plan",
#         "est":  f"FY{prev_s}-{prev_e} Estimate",
#     }

# def _prev_fy(fy):
#     parts = (fy or "2025-26").split("-")
#     s = int(parts[0] or 2025) - 1
#     e = int(parts[1] or 26)   - 1
#     return f"{s}-{str(e).zfill(2)}"


# # ═══════════════════════════════════════════════════════════════════════════════
# # WHITELISTED API ENDPOINTS
# # ═══════════════════════════════════════════════════════════════════════════════

# @frappe.whitelist()
# def export_ppt(financial_year, ppt_rows, prev_ppt_rows,
#                budget_label, est_label, prev_budget_label, prev_est_label):
#     fy      = financial_year or "2025-26"
#     prev_fy = _prev_fy(fy)
#     wb = Workbook(); wb.remove(wb.active)
#     _sheet_ppt_combined(
#         wb, "Foundation Metrics",
#         json.loads(ppt_rows),      budget_label,      est_label,
#         json.loads(prev_ppt_rows), prev_budget_label, prev_est_label,
#         fy=fy, prev_fy=prev_fy,
#     )
#     return {"filename": f"Foundation_Metrics_{fy}.xlsx", "data": _wb_to_b64(wb)}


# @frappe.whitelist()
# def export_annual(financial_year, annual_data):
#     fy = financial_year or "2025-26"
#     wb = Workbook(); wb.remove(wb.active)
#     _sheet_annual(wb, "Annual Budget Consolidated", json.loads(annual_data), fy)
#     return {"filename": f"Annual_Budget_Consolidated_{fy}.xlsx", "data": _wb_to_b64(wb)}


# @frappe.whitelist()
# def export_estimate(financial_year, estimate_data):
#     fy = financial_year or "2025-26"
#     wb = Workbook(); wb.remove(wb.active)
#     _sheet_estimate(wb, "EstimEstimate Consolidated", json.loads(estimate_data), fy)
#     return {"filename": f"Estimate_Consolidated{fy}.xlsx", "data": _wb_to_b64(wb)}


# @frappe.whitelist()
# def export_budget_estimate(financial_year, be_data):
#     fy     = financial_year or "2025-26"
#     labels = _fy_labels(fy)
#     wb = Workbook(); wb.remove(wb.active)
#     _sheet_be(wb, "Budget & Estimate", json.loads(be_data),
#               fy, labels["plan"], labels["est"])
#     return {"filename": f"Budget_and_Estimate_{fy}.xlsx", "data": _wb_to_b64(wb)}


# @frappe.whitelist()
# def export_all(financial_year, ppt_rows, prev_ppt_rows,
#                budget_label, est_label, prev_budget_label, prev_est_label,
#                annual_data, estimate_data, be_data):
#     fy      = financial_year or "2025-26"
#     prev_fy = _prev_fy(fy)
#     labels  = _fy_labels(fy)
#     wb = Workbook(); wb.remove(wb.active)
#     _sheet_ppt_combined(
#         wb, "Foundation Metrics",
#         json.loads(ppt_rows),      budget_label,      est_label,
#         json.loads(prev_ppt_rows), prev_budget_label, prev_est_label,
#         fy=fy, prev_fy=prev_fy,
#     )
#     _sheet_annual(wb,   "Annual Budget Consolidated",     json.loads(annual_data),   fy)
#     _sheet_estimate(wb, "Estimate Consolidated",          json.loads(estimate_data), fy)
#     _sheet_be(wb,       "Budget & Estimate", json.loads(be_data),
#               fy, labels["plan"], labels["est"])
#     return {"filename": f"Foundation - Consolidated Budget_{fy}.xlsx", "data": _wb_to_b64(wb)}









# import io
# import base64
# import json

# import frappe
# from openpyxl import Workbook
# from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
# from openpyxl.utils import get_column_letter

# # ═══════════════════════════════════════════════════════════════════════════════
# # SHARED STYLE CONSTANTS
# # ═══════════════════════════════════════════════════════════════════════════════

# _CENTER  = Alignment(horizontal="center", vertical="center", wrap_text=False)
# _LEFT    = Alignment(horizontal="left",   vertical="center")
# _RIGHT   = Alignment(horizontal="right",  vertical="center")
# _WRAP_L  = Alignment(horizontal="left",   vertical="center", wrap_text=True)

# _BOLD       = Font(bold=True,  name="Calibri", size=10)
# _WHITE_BOLD = Font(bold=True,  color="FFFFFF", name="Calibri", size=10)
# _NORMAL     = Font(bold=False, name="Calibri", size=10)
# _ITALIC     = Font(bold=False, italic=True, name="Calibri", size=9, color="444444")

# _FILL_HDR      = PatternFill("solid", fgColor="0076B6")   # Blue header
# _FILL_HDR2     = PatternFill("solid", fgColor="F26B21")   # Orange sub-header
# _FILL_HEAD     = PatternFill("solid", fgColor="E9F4FB")   # Light blue  — section rows
# _FILL_SUBHEAD  = PatternFill("solid", fgColor="FFF3E6")   # Light orange — sub-head rows
# _FILL_SUBTOTAL = PatternFill("solid", fgColor="EBF5FB")   # Pale blue   — sub-total rows
# _FILL_HTOTAL   = PatternFill("solid", fgColor="D4E6F1")   # Mid blue    — head total rows
# _FILL_GRAND    = PatternFill("solid", fgColor="0076B6")   # Blue        — grand total (white text)
# _FILL_WHITE    = PatternFill("solid", fgColor="FFFFFF")
# _FILL_GT_PLAN  = PatternFill("solid", fgColor="DDEAF7")   # Summary INR gt-plan row
# _FILL_GT_ACT   = PatternFill("solid", fgColor="DDEAF7")   # Summary INR gt-act row

# _THIN   = Side(style="thin")
# _BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
# _THICK_TOP = Border(left=_THIN, right=_THIN, top=Side(style="medium"), bottom=_THIN)

# NUM_FMT    = "#,##0.00"
# NUM_FMT_CR = "#,##0.00"   # All values are already in crores when we reach export

# # ─── Annual / Estimate sheet column layout (1-based) ─────────────────────────
# COL_SI    = 2   # Sl #
# COL_HEAD  = 3   # HEAD OF EXPENSE
# COL_TYPE  = 4   # TYPE OF EXPENSE
# COL_START = 5   # Apr  (first data column)
# COL_END   = 21  # YEAR total  (last data column)
# # Cols 5-7  = Apr May Jun   (Q1)
# # Cols 8-10 = Jul Aug Sep   (Q2)
# # Cols 11-13= Oct Nov Dec   (Q3)
# # Cols 14-16= Jan Feb Mar   (Q4)
# # Cols 17-20= QTR-1 QTR-2 QTR-3 QTR-4
# # Col  21   = YEAR total


# # ═══════════════════════════════════════════════════════════════════════════════
# # LOW-LEVEL HELPERS
# # ═══════════════════════════════════════════════════════════════════════════════

# def _merge(ws, r1, c1, r2, c2):
#     ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)


# def _to_roman(num):
#     vals = [1000,900,500,400,100,90,50,40,10,9,5,4,1]
#     syms = ["M","CM","D","CD","C","XC","L","XL","X","IX","V","IV","I"]
#     out, i = "", 0
#     while num > 0:
#         for _ in range(num // vals[i]):
#             out += syms[i]; num -= vals[i]
#         i += 1
#     return out


# def _sum_formula(col_letter, rows):
#     if not rows: return 0
#     return "=" + "+".join(f"{col_letter}{r}" for r in rows)


# def _qtr_formulas(r):
#     return [
#         f"=SUM(E{r}:G{r})",
#         f"=SUM(H{r}:J{r})",
#         f"=SUM(K{r}:M{r})",
#         f"=SUM(N{r}:P{r})",
#         f"=SUM(Q{r}:T{r})",
#     ]


# def _build_formula(col_letter, rows):
#     if not rows: return 0
#     return "=" + "+".join(f"{col_letter}{r}" for r in rows)


# def _totals_from_rows(rows):
#     return [_build_formula(c, rows) for c in list("EFGHIJKLMNOPQRSTU")]


# def _style_row(ws, row, fill=None, font=None, is_header=False, white_text=False):
#     for col in range(COL_SI, COL_END + 1):
#         cell = ws.cell(row=row, column=col)
#         cell.border = _BORDER
#         if is_header:
#             cell.alignment = _CENTER
#         else:
#             if col == COL_SI:
#                 cell.alignment = _CENTER
#             elif col in (COL_HEAD, COL_TYPE):
#                 cell.alignment = _WRAP_L
#             else:
#                 cell.alignment = _RIGHT
#         if fill:
#             cell.fill = fill
#         if font:
#             cell.font = font
#         elif white_text:
#             cell.font = _WHITE_BOLD


# def _fmt_numeric(ws, row, start_col=COL_START, end_col=COL_END):
#     for col in range(start_col, end_col + 1):
#         ws.cell(row=row, column=col).number_format = NUM_FMT


# def _auto_col_widths(ws):
#     for col in range(1, ws.max_column + 1):
#         letter = get_column_letter(col)
#         max_len = 0
#         for row in range(1, ws.max_row + 1):
#             val = ws.cell(row=row, column=col).value
#             if val:
#                 s = str(val)
#                 if s.startswith("="): s = "999,999,999.00"
#                 max_len = max(max_len, len(s))
#         ws.column_dimensions[letter].width = min(max(max_len + 3, 10), 60)


# def _wb_to_b64(wb):
#     buf = io.BytesIO()
#     wb.save(buf)
#     buf.seek(0)
#     return base64.b64encode(buf.read()).decode("utf-8")


# def _is_grand_total_section(sec):
#     name = (sec.get("name") or "").upper().replace("  ", " ").strip()
#     return sec.get("sequence_id") == 9999 or name == "GRAND TOTAL"


# # ═══════════════════════════════════════════════════════════════════════════════
# # SHEET: PPT — Foundation Level / Overall Metrics (two tables in one sheet)
# # ═══════════════════════════════════════════════════════════════════════════════

# def _sheet_ppt_combined(wb, sheet_name,
#                         rows, budget_label, est_label,
#                         prev_rows, prev_budget_label, prev_est_label,
#                         fy="2025-26", prev_fy="2024-25",
#                         org_name="Azim Premji Foundation"):
#     ws = wb.create_sheet(title=sheet_name[:31])

#     thin = Side(style="thin")
#     brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

#     def _write_block(start_row, block_rows, b_label, e_label, title):
#         # ── Title bar ──
#         for col in range(2, 9):
#             c = ws.cell(start_row, col)
#             c.fill = _FILL_HEAD; c.font = Font(bold=True, name="Calibri", size=11, color="003B63")
#             c.border = brd; c.alignment = _LEFT
#         ws.cell(start_row, 2, title)
#         _merge(ws, start_row, 2, start_row, 8)

#         h1 = start_row + 1
#         h2 = start_row + 2

#         # ── Header rows — style ALL cells BEFORE merging ──
#         for col in range(2, 9):
#             for hr in (h1, h2):
#                 c = ws.cell(hr, col)
#                 c.fill = _FILL_HDR; c.font = _WHITE_BOLD
#                 c.border = brd; c.alignment = _CENTER

#         ws.cell(h1, 2, "Unit")
#         ws.cell(h1, 3, b_label)
#         ws.cell(h1, 6, e_label)
#         for col, lbl in zip(range(3, 9), ["Opex","Capex","Total","Opex","Capex","Total"]):
#             ws.cell(h2, col, lbl)

#         _merge(ws, h1, 2, h2, 2)
#         _merge(ws, h1, 3, h1, 5)
#         _merge(ws, h1, 6, h1, 8)

#         # ── Note: amounts in Cr. ──
#         note_row = h2 + 1
#         ws.cell(note_row, 2, "₹ Cr.")
#         ws.cell(note_row, 2).font = _ITALIC
#         ws.cell(note_row, 2).alignment = _RIGHT
#         _merge(ws, note_row, 2, note_row, 8)

#         # ── Data rows ──
#         next_row = note_row + 1
#         for row in block_rows:
#             is_total = row.get("is_total", False)
#             label    = row.get("label", "")
#             bO = float(row.get("bOpex",  0) or 0)
#             bC = float(row.get("bCapex", 0) or 0)
#             eO = float(row.get("eOpex",  0) or 0)
#             eC = float(row.get("eCapex", 0) or 0)
#             fill = _FILL_HTOTAL if is_total else _FILL_WHITE
#             font = _BOLD        if is_total else _NORMAL
#             wt   = Font(bold=True, color="FFFFFF", name="Calibri", size=10) if is_total else _NORMAL

#             for col, val in enumerate([label, bO, bC, bO+bC, eO, eC, eO+eC], start=2):
#                 c = ws.cell(next_row, col, val if val != 0 or col == 2 else None)
#                 c.fill = fill; c.font = font; c.border = brd
#                 c.alignment = _RIGHT if col > 2 else _LEFT
#                 if col > 2: c.number_format = NUM_FMT_CR
#             next_row += 1

#         return next_row

#     # ── Sheet header ──
#     ws.append(["", org_name])
#     _merge(ws, 1, 2, 1, 8)
#     ws["B1"].font = Font(size=14, bold=True, name="Calibri", color="003B63")
#     ws["B1"].alignment = _LEFT

#     ws.append(["", f"Foundation Level Metrics – {fy}"])
#     _merge(ws, 2, 2, 2, 8)
#     ws["B2"].font = Font(size=12, bold=True, name="Calibri")
#     ws["B2"].alignment = _LEFT

#     ws.append([])  # row 3 blank

#     next_r = _write_block(4, rows, budget_label, est_label,
#                           f"Overall Foundation – Budget vs. Actual ({fy})")

#     next_r += 1  # blank separator

#     _write_block(next_r, prev_rows, prev_budget_label, prev_est_label,
#                  f"Overall Foundation – Previous Year Budget vs. Actual ({prev_fy})")

#     _auto_col_widths(ws)
#     return ws


# # ═══════════════════════════════════════════════════════════════════════════════
# # SHEET: Summary in INR — Table A (Unit Wise Plan) + Table B (Breakdown)
# # ═══════════════════════════════════════════════════════════════════════════════

# def _sheet_summary_inr(wb, sheet_name, api_data, fy,
#                        org_name="Azim Premji Foundation"):
#     ws = wb.create_sheet(title=sheet_name[:31])
#     thin = Side(style="thin")
#     brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

#     fy_parts   = (fy or "2025-26").split("-")
#     fy_start   = int(fy_parts[0] or "2025")
#     fy_end_yy  = int(fy_parts[1] or "26")
#     plan_label = f"{fy} Budget"
#     prev_start = fy_start - 1
#     prev_end   = str(fy_end_yy - 1).zfill(2)
#     act_label  = f"{prev_start}-{prev_end} Est"

#     def _norm(s): return (s or "").replace("  ", " ").strip().upper()
#     def _fv(v): return float(v or 0)

#     def _extract_a(actuals):
#         op, cp = 0.0, 0.0
#         oa, ca = 0.0, 0.0
#         for sec in (actuals or []):
#             nm = _norm(sec.get("name",""))
#             if "OPERATING" in nm:
#                 op += _fv(sec.get("ytd"))
#                 oa += _fv(sec.get("total_posted_amt_ytd"))
#             elif "CAPITAL" in nm:
#                 cp += _fv(sec.get("ytd"))
#                 ca += _fv(sec.get("total_posted_amt_ytd"))
#         return op, cp, oa, ca

#     # ── Sheet title ──
#     ws.append(["", org_name])
#     _merge(ws, 1, 2, 1, 8)
#     ws["B1"].font = Font(size=14, bold=True, name="Calibri", color="003B63")
#     ws["B1"].alignment = _LEFT
#     ws.append(["", f"Summary in INR – {fy}"])
#     _merge(ws, 2, 2, 2, 8)
#     ws["B2"].font = Font(size=12, bold=True, name="Calibri"); ws["B2"].alignment = _LEFT
#     ws.append([])

#     # ════════════════════════════════
#     # TABLE A — Unit Wise Plan
#     # ════════════════════════════════
#     cur_row = 4
#     ws.cell(cur_row, 2, "A. Unit Wise Plan")
#     ws.cell(cur_row, 2).font = Font(bold=True, name="Calibri", size=11, underline="single")
#     cur_row += 1

#     # Note
#     ws.cell(cur_row, 8, "₹ Cr.")
#     ws.cell(cur_row, 8).font = _ITALIC; ws.cell(cur_row, 8).alignment = _RIGHT
#     cur_row += 1

#     # Header row 1
#     h1 = cur_row
#     for col in range(2, 9):
#         c = ws.cell(h1, col); c.fill = _FILL_HDR; c.font = _WHITE_BOLD
#         c.border = brd; c.alignment = _CENTER
#     ws.cell(h1, 2, "Unit / Function")
#     ws.cell(h1, 3, plan_label)
#     ws.cell(h1, 6, act_label)
#     _merge(ws, h1, 2, h1+1, 2)
#     _merge(ws, h1, 3, h1,   5)
#     _merge(ws, h1, 6, h1,   8)
#     cur_row += 1

#     # Header row 2
#     h2 = cur_row
#     for col in range(2, 9):
#         c = ws.cell(h2, col); c.fill = _FILL_HDR2; c.font = _WHITE_BOLD
#         c.border = brd; c.alignment = _CENTER
#     for col, lbl in zip(range(3, 9), ["Opex","Capex","Total","Opex","Capex","Total"]):
#         ws.cell(h2, col, lbl)
#     cur_row += 1

#     sorted_data = sorted((api_data or []), key=lambda e: e.get("sequence_id", 0))
#     normal_rows = [e for e in sorted_data if "covid" not in (e.get("label","")).lower()]
#     covid_rows  = [e for e in sorted_data if "covid"     in (e.get("label","")).lower()]

#     tot_op = tot_cp = tot_oa = tot_ca = 0.0
#     total_data_rows = []

#     for entry in normal_rows:
#         op, cp, oa, ca = _extract_a(entry.get("actuals",[]))
#         is_sub = entry.get("is_this_sub_item") == 1
#         fill   = _FILL_WHITE
#         font   = _NORMAL
#         indent = "    " if is_sub else ""
#         for col, val in enumerate(
#             [indent + (entry.get("label") or "").strip(),
#              op/1e7, cp/1e7, (op+cp)/1e7, oa/1e7, ca/1e7, (oa+ca)/1e7], start=2):
#             c = ws.cell(cur_row, col, val)
#             c.fill = fill; c.font = font; c.border = brd
#             c.alignment = _LEFT if col == 2 else _RIGHT
#             if col > 2: c.number_format = NUM_FMT_CR
#         if not is_sub:
#             tot_op += op; tot_cp += cp; tot_oa += oa; tot_ca += ca
#             total_data_rows.append(cur_row)
#         cur_row += 1

#     # Normal total row
#     for col, val in enumerate(
#         ["Total", tot_op/1e7, tot_cp/1e7, (tot_op+tot_cp)/1e7,
#          tot_oa/1e7, tot_ca/1e7, (tot_oa+tot_ca)/1e7], start=2):
#         c = ws.cell(cur_row, col, val)
#         c.fill = _FILL_HTOTAL; c.font = _BOLD; c.border = _THICK_TOP
#         c.alignment = _LEFT if col == 2 else _RIGHT
#         if col > 2: c.number_format = NUM_FMT_CR
#     cur_row += 1

#     # Covid rows
#     cov_op = cov_cp = cov_oa = cov_ca = 0.0
#     for entry in covid_rows:
#         op, cp, oa, ca = _extract_a(entry.get("actuals",[]))
#         for col, val in enumerate(
#             [(entry.get("label") or "").strip(),
#              op/1e7, cp/1e7, (op+cp)/1e7, oa/1e7, ca/1e7, (oa+ca)/1e7], start=2):
#             c = ws.cell(cur_row, col, val)
#             c.fill = _FILL_WHITE; c.font = _NORMAL; c.border = brd
#             c.alignment = _LEFT if col == 2 else _RIGHT
#             if col > 2: c.number_format = NUM_FMT_CR
#         cov_op += op; cov_cp += cp; cov_oa += oa; cov_ca += ca
#         cur_row += 1

#     if covid_rows:
#         grand_op = tot_op + cov_op; grand_cp = tot_cp + cov_cp
#         grand_oa = tot_oa + cov_oa; grand_ca = tot_ca + cov_ca
#         for col, val in enumerate(
#             ["Grand Total", grand_op/1e7, grand_cp/1e7, (grand_op+grand_cp)/1e7,
#              grand_oa/1e7, grand_ca/1e7, (grand_oa+grand_ca)/1e7], start=2):
#             c = ws.cell(cur_row, col, val)
#             c.fill = _FILL_GRAND
#             c.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
#             c.border = brd
#             c.alignment = _LEFT if col == 2 else _RIGHT
#             if col > 2: c.number_format = NUM_FMT_CR
#         cur_row += 1

#     cur_row += 1  # spacer

#     # ════════════════════════════════
#     # TABLE B — Breakdown of Unit Wise Plan (sub-heads)
#     # ════════════════════════════════
#     ws.cell(cur_row, 2, "B. Breakdown of Unit Wise Plan")
#     ws.cell(cur_row, 2).font = Font(bold=True, name="Calibri", size=11, underline="single")
#     cur_row += 1

#     ws.cell(cur_row, 8, "₹ Cr.")
#     ws.cell(cur_row, 8).font = _ITALIC; ws.cell(cur_row, 8).alignment = _RIGHT
#     cur_row += 1

#     # Collect sub-head names from OPERATING EXPENSES
#     seen_sh = {}; sub_head_names = []
#     main_entries = [e for e in sorted_data if e.get("is_this_sub_item") != 1]
#     for entry in main_entries:
#         for sec in (entry.get("actuals") or []):
#             nm = _norm(sec.get("name",""))
#             if "OPERATING" not in nm: continue
#             for sh in (sec.get("sub_heads") or []):
#                 n = (sh.get("name") or "").strip()
#                 if n and n not in seen_sh:
#                     seen_sh[n] = True; sub_head_names.append(n)

#     n_sh   = len(sub_head_names)
#     # Col layout: B=Unit, C..C+n_sh-1 = sub-heads, C+n_sh = Opex Total, C+n_sh+1 = Capex, C+n_sh+2 = Total
#     # B=2, sub-head cols = 3..2+n_sh, opex_tot=3+n_sh, capex=4+n_sh, total=5+n_sh
#     c_sh_start = 3
#     c_opex_tot = c_sh_start + n_sh
#     c_capex    = c_opex_tot + 1
#     c_total    = c_capex + 1
#     last_col   = c_total

#     def _sh_val(actuals, shname, field):
#         v = 0.0
#         for sec in (actuals or []):
#             if "OPERATING" not in _norm(sec.get("name","")): continue
#             for sh in (sec.get("sub_heads") or []):
#                 if sh.get("name","").strip() == shname:
#                     v += _fv(sh.get(field))
#         return v

#     def _opex_total(actuals, field):
#         v = 0.0
#         for sec in (actuals or []):
#             if "OPERATING" not in _norm(sec.get("name","")): continue
#             v += _fv(sec.get(field))
#         return v

#     def _capex_total(actuals, field):
#         v = 0.0
#         for sec in (actuals or []):
#             if "CAPITAL" not in _norm(sec.get("name","")): continue
#             v += _fv(sec.get(field))
#         return v

#     # Header row 1
#     bh1 = cur_row
#     for col in range(2, last_col+1):
#         c = ws.cell(bh1, col); c.fill = _FILL_HDR; c.font = _WHITE_BOLD
#         c.border = brd; c.alignment = _CENTER
#     ws.cell(bh1, 2, "Unit / Function")
#     _merge(ws, bh1, 2, bh1+1, 2)
#     ws.cell(bh1, c_sh_start, "Operating Expenses")
#     _merge(ws, bh1, c_sh_start, bh1, c_opex_tot)
#     ws.cell(bh1, c_capex, "Capex")
#     _merge(ws, bh1, c_capex, bh1+1, c_capex)
#     ws.cell(bh1, c_total, "Total")
#     _merge(ws, bh1, c_total, bh1+1, c_total)
#     cur_row += 1

#     # Header row 2
#     bh2 = cur_row
#     for col in range(2, last_col+1):
#         c = ws.cell(bh2, col); c.fill = _FILL_HDR2; c.font = _WHITE_BOLD
#         c.border = brd; c.alignment = _CENTER
#     for i, shname in enumerate(sub_head_names):
#         ws.cell(bh2, c_sh_start+i, shname)
#     ws.cell(bh2, c_opex_tot, "Total")
#     cur_row += 1

#     gt_sh_plan = {n: 0.0 for n in sub_head_names}
#     gt_sh_act  = {n: 0.0 for n in sub_head_names}
#     gt_opex_p = gt_opex_a = gt_cap_p = gt_cap_a = 0.0

#     for entry in main_entries:
#         act    = entry.get("actuals") or []
#         label  = (entry.get("label") or "").strip()

#         # Unit header row
#         for col in range(2, last_col+1):
#             c = ws.cell(cur_row, col); c.fill = _FILL_HEAD; c.border = brd
#             c.font = _BOLD; c.alignment = _LEFT if col==2 else _CENTER
#         ws.cell(cur_row, 2, label)
#         _merge(ws, cur_row, 3, cur_row, last_col)
#         cur_row += 1

#         # Plan row
#         for col in range(2, last_col+1):
#             c = ws.cell(cur_row, col); c.fill = _FILL_WHITE; c.border = brd; c.font = _NORMAL
#             c.alignment = _LEFT if col==2 else _RIGHT
#         ws.cell(cur_row, 2, f"  - {plan_label}")
#         opex_p = _opex_total(act, "ytd"); cap_p = _capex_total(act, "ytd")
#         for i, shname in enumerate(sub_head_names):
#             v = _sh_val(act, shname, "ytd")
#             ws.cell(cur_row, c_sh_start+i, v/1e7).number_format = NUM_FMT_CR
#             gt_sh_plan[shname] += v
#         ws.cell(cur_row, c_opex_tot, opex_p/1e7).number_format = NUM_FMT_CR
#         ws.cell(cur_row, c_capex,    cap_p/1e7).number_format  = NUM_FMT_CR
#         ws.cell(cur_row, c_total,   (opex_p+cap_p)/1e7).number_format = NUM_FMT_CR
#         gt_opex_p += opex_p; gt_cap_p += cap_p
#         cur_row += 1

#         # Actual row
#         for col in range(2, last_col+1):
#             c = ws.cell(cur_row, col); c.fill = PatternFill("solid", fgColor="FAFAFA")
#             c.border = brd; c.font = _NORMAL; c.alignment = _LEFT if col==2 else _RIGHT
#         ws.cell(cur_row, 2, f"  - {act_label}")
#         opex_a = _opex_total(act, "total_posted_amt_ytd"); cap_a = _capex_total(act, "total_posted_amt_ytd")
#         for i, shname in enumerate(sub_head_names):
#             v = _sh_val(act, shname, "total_posted_amt_ytd")
#             ws.cell(cur_row, c_sh_start+i, v/1e7).number_format = NUM_FMT_CR
#             gt_sh_act[shname] += v
#         ws.cell(cur_row, c_opex_tot, opex_a/1e7).number_format = NUM_FMT_CR
#         ws.cell(cur_row, c_capex,    cap_a/1e7).number_format  = NUM_FMT_CR
#         ws.cell(cur_row, c_total,   (opex_a+cap_a)/1e7).number_format = NUM_FMT_CR
#         gt_opex_a += opex_a; gt_cap_a += cap_a
#         cur_row += 1

#         # Spacer
#         for col in range(2, last_col+1):
#             c = ws.cell(cur_row, col)
#             c.fill = PatternFill("solid", fgColor="F4F6F8"); c.border = brd
#         cur_row += 1

#     # Grand Total label
#     for col in range(2, last_col+1):
#         c = ws.cell(cur_row, col); c.fill = _FILL_HTOTAL; c.border = brd
#         c.font = _BOLD; c.alignment = _LEFT if col==2 else _RIGHT
#     ws.cell(cur_row, 2, "Grand Total")
#     cur_row += 1

#     # GT Plan row
#     for col in range(2, last_col+1):
#         c = ws.cell(cur_row, col); c.fill = _FILL_GT_PLAN; c.border = brd
#         c.font = _NORMAL; c.alignment = _LEFT if col==2 else _RIGHT
#     ws.cell(cur_row, 2, f"  - {plan_label}")
#     for i, shname in enumerate(sub_head_names):
#         ws.cell(cur_row, c_sh_start+i, gt_sh_plan[shname]/1e7).number_format = NUM_FMT_CR
#     ws.cell(cur_row, c_opex_tot, gt_opex_p/1e7).number_format = NUM_FMT_CR
#     ws.cell(cur_row, c_capex,    gt_cap_p/1e7).number_format  = NUM_FMT_CR
#     ws.cell(cur_row, c_total,   (gt_opex_p+gt_cap_p)/1e7).number_format = NUM_FMT_CR
#     cur_row += 1

#     # GT Actual row
#     for col in range(2, last_col+1):
#         c = ws.cell(cur_row, col); c.fill = _FILL_GT_ACT; c.border = brd
#         c.font = _NORMAL; c.alignment = _LEFT if col==2 else _RIGHT
#     ws.cell(cur_row, 2, f"  - {act_label}")
#     for i, shname in enumerate(sub_head_names):
#         ws.cell(cur_row, c_sh_start+i, gt_sh_act[shname]/1e7).number_format = NUM_FMT_CR
#     ws.cell(cur_row, c_opex_tot, gt_opex_a/1e7).number_format = NUM_FMT_CR
#     ws.cell(cur_row, c_capex,    gt_cap_a/1e7).number_format  = NUM_FMT_CR
#     ws.cell(cur_row, c_total,   (gt_opex_a+gt_cap_a)/1e7).number_format = NUM_FMT_CR

#     _auto_col_widths(ws)
#     return ws


# # ═══════════════════════════════════════════════════════════════════════════════
# # SHEET: Headcount
# # ═══════════════════════════════════════════════════════════════════════════════

# def _sheet_headcount(wb, sheet_name, records, fy, org_name="Azim Premji Foundation"):
#     ws = wb.create_sheet(title=sheet_name[:31])
#     thin = Side(style="thin")
#     brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

#     def _fym(fy_str):
#         p = (fy_str or "").split("-")
#         return f"31-Mar-{p[1] if len(p)>1 else p[0][-2:]}"

#     def _fyl(fy_str):
#         p = (fy_str or "").split("-")
#         return f"FY{(p[0] or '')[-2:]}-{p[1] if len(p)>1 else ''}"

#     def _avg(a, b):
#         if a is None and b is None: return None
#         if a is None: return b
#         if b is None: return a
#         return (a + b) / 2

#     # transform
#     sorted_recs = sorted((records or []), key=lambda r: r.get("financial_year",""))
#     yrs = [r.get("financial_year","") for r in sorted_recs]
#     unit_map = {}
#     for rec in sorted_recs:
#         for u in (rec.get("units") or []):
#             uid = str(u.get("unit",""))
#             if uid not in unit_map:
#                 unit_map[uid] = {"id": uid, "desc": "", "hc": {}}
#             unit_map[uid]["hc"][rec["financial_year"]] = u.get("total_headcount") or 0
#             if rec["financial_year"] == yrs[-1]:
#                 unit_map[uid]["desc"] = (u.get("unit_description") or "").strip()
#     units = sorted(unit_map.values(), key=lambda u: int(u["id"]) if u["id"].isdigit() else 999)
#     totals = {r["financial_year"]: r.get("total_head_count",0) for r in sorted_recs}

#     # ── Sheet title ──
#     ws.append(["", org_name])
#     _merge(ws, 1, 2, 1, 2+len(yrs))
#     ws["B1"].font = Font(size=14, bold=True, name="Calibri", color="003B63"); ws["B1"].alignment = _LEFT
#     ws.append(["", f"Headcount – {fy}"])
#     _merge(ws, 2, 2, 2, 2+len(yrs))
#     ws["B2"].font = Font(size=12, bold=True, name="Calibri"); ws["B2"].alignment = _LEFT
#     ws.append([])

#     def _write_hc_section(title, get_val_fn, cur_row, fmt_fn=None):
#         ws.cell(cur_row, 2, title)
#         ws.cell(cur_row, 2).font = Font(bold=True, name="Calibri", size=11,
#                                         underline="single", color="003B63")
#         cur_row += 1
#         # header
#         hdr_cols = [_fym(y) for y in yrs]
#         for col, lbl in enumerate(["Unit"] + hdr_cols, start=2):
#             c = ws.cell(cur_row, col, lbl)
#             c.fill = _FILL_HDR; c.font = _WHITE_BOLD; c.border = brd; c.alignment = _CENTER
#         cur_row += 1
#         for u in units:
#             ws.cell(cur_row, 2, u["desc"]).alignment = _LEFT
#             ws.cell(cur_row, 2).border = brd; ws.cell(cur_row, 2).font = _NORMAL
#             for yi, y in enumerate(yrs):
#                 v = get_val_fn(u, yi, y)
#                 c = ws.cell(cur_row, 3+yi, round(v) if v is not None else None)
#                 c.border = brd; c.font = _NORMAL; c.alignment = _RIGHT
#                 if v is not None: c.number_format = "#,##0"
#             cur_row += 1
#         # total row
#         ws.cell(cur_row, 2, "Total").alignment = _LEFT
#         ws.cell(cur_row, 2).fill = _FILL_GRAND
#         ws.cell(cur_row, 2).font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
#         ws.cell(cur_row, 2).border = brd
#         for yi, y in enumerate(yrs):
#             v = totals.get(y)
#             avg_v = None
#             if yi == 0: avg_v = v
#             else:
#                 prev_v = totals.get(yrs[yi-1])
#                 if prev_v is not None and v is not None: avg_v = (prev_v + v)/2
#             val = get_val_fn({"hc": totals}, yi, y)
#             c = ws.cell(cur_row, 3+yi, round(val) if val is not None else None)
#             c.fill = _FILL_GRAND
#             c.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
#             c.border = brd; c.alignment = _RIGHT
#             if val is not None: c.number_format = "#,##0"
#         cur_row += 2
#         return cur_row

#     cur_row = 4

#     # Closing H/C
#     def closing_val(u, yi, y): return u["hc"].get(y)
#     cur_row = _write_hc_section("Closing H/C", closing_val, cur_row)

#     # Average H/C
#     def avg_val(u, yi, y):
#         if yi == 0: return u["hc"].get(yrs[0])
#         prev = u["hc"].get(yrs[yi-1]); curr = u["hc"].get(y)
#         if prev is None or curr is None: return None
#         return (prev + curr) / 2
#     cur_row = _write_hc_section("Average H/C", avg_val, cur_row)

#     # % Increase closing
#     ws.cell(cur_row, 2, "Increase in Closing H/C (%)")
#     ws.cell(cur_row, 2).font = Font(bold=True, name="Calibri", size=11,
#                                     underline="single", color="003B63")
#     cur_row += 1
#     if len(yrs) >= 2:
#         pairs = [(yrs[i-1], yrs[i]) for i in range(1, len(yrs))]
#         hdr_cols = [f"{_fym(p[0])} → {_fym(p[1])}" for p in pairs]
#         for col, lbl in enumerate(["Unit"] + hdr_cols, start=2):
#             c = ws.cell(cur_row, col, lbl)
#             c.fill = _FILL_HDR; c.font = _WHITE_BOLD; c.border = brd; c.alignment = _CENTER
#         cur_row += 1
#         for u in units:
#             ws.cell(cur_row, 2, u["desc"]).alignment = _LEFT
#             ws.cell(cur_row, 2).border = brd; ws.cell(cur_row, 2).font = _NORMAL
#             for pi, (yf, yt) in enumerate(pairs):
#                 a = u["hc"].get(yf); b = u["hc"].get(yt)
#                 if a and b: pct = ((b/a)-1)*100
#                 else: pct = None
#                 c = ws.cell(cur_row, 3+pi, round(pct,1) if pct is not None else None)
#                 c.border = brd; c.font = _NORMAL; c.alignment = _RIGHT
#                 if pct is not None: c.number_format = '+0.0%;-0.0%;0.0%'
#             cur_row += 1
#         ws.cell(cur_row, 2, "Total").fill = _FILL_GRAND
#         ws.cell(cur_row, 2).font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
#         ws.cell(cur_row, 2).alignment = _LEFT; ws.cell(cur_row, 2).border = brd
#         for pi, (yf, yt) in enumerate(pairs):
#             a = totals.get(yf); b = totals.get(yt)
#             if a and b: pct = ((b/a)-1)*100
#             else: pct = None
#             c = ws.cell(cur_row, 3+pi, round(pct,1) if pct is not None else None)
#             c.fill = _FILL_GRAND
#             c.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
#             c.border = brd; c.alignment = _RIGHT
#             if pct is not None: c.number_format = '+0.0%;-0.0%;0.0%'
#         cur_row += 2

#     _auto_col_widths(ws)
#     return ws


# # ═══════════════════════════════════════════════════════════════════════════════
# # SHEET: Annual Budget Consolidated
# # ═══════════════════════════════════════════════════════════════════════════════

# def _sheet_annual(wb, sheet_name, data, fy, org_name="Azim Premji Foundation"):
#     ws = wb.create_sheet(title=sheet_name[:31])

#     ws.append(["", org_name])
#     ws.merge_cells("B1:U1")
#     ws["B1"].font = Font(size=14, bold=True, name="Calibri", color="003B63")
#     ws["B1"].alignment = _LEFT

#     ws.append(["", f"Budget for the Financial Year {fy or ''}"])
#     ws.merge_cells("B2:U2")
#     ws["B2"].font = Font(size=12, bold=True, name="Calibri")
#     ws["B2"].alignment = _LEFT

#     ws.append([])

#     ws.append([
#         "", "Sl #", "HEAD OF EXPENSE", "TYPE OF EXPENSE",
#         "QUARTER I","","","QUARTER II","","",
#         "QUARTER III","","","QUARTER IV","","",
#         "QTR-1","QTR-2","QTR-3","QTR-4", f"YEAR {fy}",
#     ])
#     r1 = ws.max_row
#     ws.append([
#         "", "Sl #", "HEAD OF EXPENSE", "TYPE OF EXPENSE",
#         "Apr","May","Jun","Jul","Aug","Sep",
#         "Oct","Nov","Dec","Jan","Feb","Mar",
#         "QTR-1","QTR-2","QTR-3","QTR-4", f"YEAR {fy}",
#     ])
#     r2 = ws.max_row

#     _style_row(ws, r1, _FILL_HDR, _WHITE_BOLD, is_header=True)
#     _style_row(ws, r2, _FILL_HDR, _WHITE_BOLD, is_header=True)

#     ws.merge_cells(start_row=r1, start_column=5,  end_row=r1, end_column=7)
#     ws.merge_cells(start_row=r1, start_column=8,  end_row=r1, end_column=10)
#     ws.merge_cells(start_row=r1, start_column=11, end_row=r1, end_column=13)
#     ws.merge_cells(start_row=r1, start_column=14, end_row=r1, end_column=16)
#     ws.merge_cells(start_row=r1, start_column=2,  end_row=r2, end_column=2)
#     ws.merge_cells(start_row=r1, start_column=3,  end_row=r2, end_column=3)
#     ws.merge_cells(start_row=r1, start_column=4,  end_row=r2, end_column=4)
#     for col in range(17, 22):
#         ws.merge_cells(start_row=r1, start_column=col, end_row=r2, end_column=col)

#     ws.freeze_panes = "E6"

#     head_total_rows = []
#     head_counter    = 0

#     for head in (data or []):
#         head_counter += 1
#         alpha_index = chr(64 + head_counter)
#         head_name   = (head.get("name") or "").strip().upper()

#         if head_name == "COVID SUPPORT":
#             ws.append([])
#             item = head["items"][0] if head.get("items") else {}
#             r = ws.max_row + 1
#             ws.append([
#                 "", alpha_index, head["name"], item.get("name",""),
#                 *item.get("q1",[0,0,0]), *item.get("q2",[0,0,0]),
#                 *item.get("q3",[0,0,0]), *item.get("q4",[0,0,0]),
#                 *_qtr_formulas(r),
#             ])
#             _style_row(ws, ws.max_row, font=_NORMAL)
#             _fmt_numeric(ws, ws.max_row)
#             ws.cell(row=ws.max_row, column=COL_HEAD).font = _BOLD
#             continue

#         ws.append(["", alpha_index, head["name"]])
#         r_sec = ws.max_row
#         for col in range(COL_HEAD, COL_END + 1):
#             c = ws.cell(r_sec, col)
#             c.fill = _FILL_HEAD; c.font = Font(bold=True, name="Calibri", size=10, color="003B63")
#             c.border = _BORDER; c.alignment = _LEFT
#         ws.cell(r_sec, COL_SI).fill = _FILL_HEAD
#         ws.cell(r_sec, COL_SI).border = _BORDER
#         ws.merge_cells(start_row=r_sec, start_column=COL_HEAD, end_row=r_sec, end_column=COL_END)

#         if head_name == "OPERATING EXPENSES":
#             ws.append([])

#         sub_total_rows   = []
#         direct_item_rows = []

#         for item in head.get("items", []):
#             r = ws.max_row + 1
#             sub_val = (item.get("sub_head_of_expense") or "").strip()
#             ws.append([
#                 "", "", sub_val, item["name"],
#                 *item.get("q1",[0,0,0]), *item.get("q2",[0,0,0]),
#                 *item.get("q3",[0,0,0]), *item.get("q4",[0,0,0]),
#                 *_qtr_formulas(r),
#             ])
#             _style_row(ws, ws.max_row, font=_NORMAL)
#             _fmt_numeric(ws, ws.max_row)
#             direct_item_rows.append(ws.max_row)

#         sub_counter = 1
#         for sub in head.get("sub_heads", []):
#             roman_index = _to_roman(sub_counter)
#             ws.append(["", roman_index, sub["name"]])
#             r_sub = ws.max_row
#             for col in range(COL_HEAD, COL_END + 1):
#                 c = ws.cell(r_sub, col)
#                 c.fill = _FILL_SUBHEAD; c.font = Font(bold=True, name="Calibri", size=10, color="7A3B00")
#                 c.border = _BORDER; c.alignment = _LEFT
#             ws.cell(r_sub, COL_SI).fill = _FILL_SUBHEAD
#             ws.cell(r_sub, COL_SI).border = _BORDER
#             ws.merge_cells(start_row=r_sub, start_column=COL_HEAD, end_row=r_sub, end_column=COL_END)

#             sub_item_rows = []
#             for item in sub.get("items", []):
#                 r = ws.max_row + 1
#                 item_sub = (item.get("sub_head_of_expense") or "").strip()
#                 sub_nm   = (sub.get("name") or "").strip()
#                 head_display = item_sub if item_sub.lower() != sub_nm.lower() else ""
#                 ws.append([
#                     "", "", head_display, item["name"],
#                     *item.get("q1",[0,0,0]), *item.get("q2",[0,0,0]),
#                     *item.get("q3",[0,0,0]), *item.get("q4",[0,0,0]),
#                     *_qtr_formulas(r),
#                 ])
#                 _style_row(ws, ws.max_row, font=_NORMAL)
#                 _fmt_numeric(ws, ws.max_row)
#                 sub_item_rows.append(ws.max_row)

#             if sub_item_rows:
#                 ws.append(["", "", "", f"TOTAL - {sub['name']}", *_totals_from_rows(sub_item_rows)])
#                 _style_row(ws, ws.max_row, _FILL_SUBTOTAL, _BOLD)
#                 _fmt_numeric(ws, ws.max_row)
#                 sub_total_rows.append(ws.max_row)

#             sub_counter += 1

#         total_rows = sub_total_rows if sub_total_rows else direct_item_rows
#         if total_rows:
#             ws.append(["", "", "", f"TOTAL - {head['name']}", *_totals_from_rows(total_rows)])
#             _style_row(ws, ws.max_row, _FILL_HTOTAL, Font(bold=True, name="Calibri", size=10, color="003B63"))
#             _fmt_numeric(ws, ws.max_row)
#             head_total_rows.append(ws.max_row)
#             if head_name == "OPERATING EXPENSES":
#                 ws.append([])

#     # Clean up blank row before COVID SUPPORT
#     for r in range(ws.max_row, 1, -1):
#         if (ws.cell(r, COL_HEAD).value == "COVID SUPPORT" and
#                 ws.cell(r-1, COL_HEAD).value is None):
#             ws.delete_rows(r-1); break

#     if head_total_rows:
#         ws.append(["", "", "", "GRAND TOTAL", *_totals_from_rows(head_total_rows)])
#         _style_row(ws, ws.max_row, _FILL_GRAND,
#                    Font(bold=True, color="FFFFFF", name="Calibri", size=10))
#         _fmt_numeric(ws, ws.max_row)

#     _auto_col_widths(ws)
#     return ws


# # ═══════════════════════════════════════════════════════════════════════════════
# # SHEET: Estimate Consolidated
# # ═══════════════════════════════════════════════════════════════════════════════

# def _sheet_estimate(wb, sheet_name, data, fy, org_name="Azim Premji Foundation"):
#     MONTH_MAP = {
#         "q1": ["4","5","6"], "q2": ["7","8","9"],
#         "q3": ["10","11","12"], "q4": ["1","2","3"],
#     }

#     def _to_qlist(obj):
#         m = obj.get("months") or {}
#         return {qk: [float(m.get(k, 0) or 0) for k in keys]
#                 for qk, keys in MONTH_MAP.items()}

#     def _norm(obj):
#         ql = _to_qlist(obj); obj.update(ql)
#         for item in obj.get("items", []):
#             item.update(_to_qlist(item))
#         for sub in obj.get("sub_heads", []):
#             sub.update(_to_qlist(sub))
#             for item in sub.get("items", []):
#                 item.update(_to_qlist(item))
#         return obj

#     normalised = [_norm(dict(h)) for h in (data or [])]

#     ws = wb.create_sheet(title=sheet_name[:31])

#     ws.append(["", org_name])
#     ws.merge_cells("B1:U1")
#     ws["B1"].font = Font(size=14, bold=True, name="Calibri", color="003B63")
#     ws["B1"].alignment = _LEFT

#     ws.append(["", f"Estimate for the Financial Year {fy or ''}"])
#     ws.merge_cells("B2:U2")
#     ws["B2"].font = Font(size=12, bold=True, name="Calibri")
#     ws["B2"].alignment = _LEFT
#     ws.append([])

#     ws.append([
#         "", "Sl #", "HEAD OF EXPENSE", "TYPE OF EXPENSE",
#         "QUARTER I","","","QUARTER II","","",
#         "QUARTER III","","","QUARTER IV","","",
#         "QTR-1","QTR-2","QTR-3","QTR-4", f"YEAR {fy}",
#     ])
#     r1 = ws.max_row
#     ws.append([
#         "", "Sl #", "HEAD OF EXPENSE", "TYPE OF EXPENSE",
#         "Apr","May","Jun","Jul","Aug","Sep",
#         "Oct","Nov","Dec","Jan","Feb","Mar",
#         "QTR-1","QTR-2","QTR-3","QTR-4", f"YEAR {fy}",
#     ])
#     r2 = ws.max_row

#     _style_row(ws, r1, _FILL_HDR, _WHITE_BOLD, is_header=True)
#     _style_row(ws, r2, _FILL_HDR, _WHITE_BOLD, is_header=True)

#     ws.merge_cells(start_row=r1, start_column=5,  end_row=r1, end_column=7)
#     ws.merge_cells(start_row=r1, start_column=8,  end_row=r1, end_column=10)
#     ws.merge_cells(start_row=r1, start_column=11, end_row=r1, end_column=13)
#     ws.merge_cells(start_row=r1, start_column=14, end_row=r1, end_column=16)
#     ws.merge_cells(start_row=r1, start_column=2,  end_row=r2, end_column=2)
#     ws.merge_cells(start_row=r1, start_column=3,  end_row=r2, end_column=3)
#     ws.merge_cells(start_row=r1, start_column=4,  end_row=r2, end_column=4)
#     for col in range(17, 22):
#         ws.merge_cells(start_row=r1, start_column=col, end_row=r2, end_column=col)

#     ws.freeze_panes = "E6"

#     head_total_rows = []
#     head_counter    = 0

#     for head in normalised:
#         head_counter += 1
#         alpha_index = chr(64 + head_counter)
#         head_name   = (head.get("name") or "").strip().upper()

#         if head_name == "COVID SUPPORT":
#             ws.append([])
#             item = head["items"][0] if head.get("items") else {}
#             r = ws.max_row + 1
#             ws.append([
#                 "", alpha_index, head["name"], item.get("name",""),
#                 *item.get("q1",[0,0,0]), *item.get("q2",[0,0,0]),
#                 *item.get("q3",[0,0,0]), *item.get("q4",[0,0,0]),
#                 *_qtr_formulas(r),
#             ])
#             _style_row(ws, ws.max_row, font=_NORMAL)
#             _fmt_numeric(ws, ws.max_row)
#             ws.cell(row=ws.max_row, column=COL_HEAD).font = _BOLD
#             continue

#         ws.append(["", alpha_index, head["name"]])
#         r_sec = ws.max_row
#         for col in range(COL_HEAD, COL_END + 1):
#             c = ws.cell(r_sec, col)
#             c.fill = _FILL_HEAD; c.font = Font(bold=True, name="Calibri", size=10, color="003B63")
#             c.border = _BORDER; c.alignment = _LEFT
#         ws.cell(r_sec, COL_SI).fill = _FILL_HEAD
#         ws.cell(r_sec, COL_SI).border = _BORDER
#         ws.merge_cells(start_row=r_sec, start_column=COL_HEAD, end_row=r_sec, end_column=COL_END)

#         if head_name == "OPERATING EXPENSES":
#             ws.append([])

#         sub_total_rows = []
#         direct_item_rows = []

#         for item in head.get("items", []):
#             r = ws.max_row + 1
#             sub_val = (item.get("sub_head_of_expense") or "").strip()
#             ws.append([
#                 "", "", sub_val, item["name"],
#                 *item.get("q1",[0,0,0]), *item.get("q2",[0,0,0]),
#                 *item.get("q3",[0,0,0]), *item.get("q4",[0,0,0]),
#                 *_qtr_formulas(r),
#             ])
#             _style_row(ws, ws.max_row, font=_NORMAL)
#             _fmt_numeric(ws, ws.max_row)
#             direct_item_rows.append(ws.max_row)

#         sub_counter = 1
#         for sub in head.get("sub_heads", []):
#             roman_index = _to_roman(sub_counter)
#             ws.append(["", roman_index, sub["name"]])
#             r_sub = ws.max_row
#             for col in range(COL_HEAD, COL_END + 1):
#                 c = ws.cell(r_sub, col)
#                 c.fill = _FILL_SUBHEAD; c.font = Font(bold=True, name="Calibri", size=10, color="7A3B00")
#                 c.border = _BORDER; c.alignment = _LEFT
#             ws.cell(r_sub, COL_SI).fill = _FILL_SUBHEAD
#             ws.cell(r_sub, COL_SI).border = _BORDER
#             ws.merge_cells(start_row=r_sub, start_column=COL_HEAD, end_row=r_sub, end_column=COL_END)

#             sub_item_rows = []
#             for item in sub.get("items", []):
#                 r = ws.max_row + 1
#                 item_sub = (item.get("sub_head_of_expense") or "").strip()
#                 sub_nm   = (sub.get("name") or "").strip()
#                 head_display = item_sub if item_sub.lower() != sub_nm.lower() else ""
#                 ws.append([
#                     "", "", head_display, item["name"],
#                     *item.get("q1",[0,0,0]), *item.get("q2",[0,0,0]),
#                     *item.get("q3",[0,0,0]), *item.get("q4",[0,0,0]),
#                     *_qtr_formulas(r),
#                 ])
#                 _style_row(ws, ws.max_row, font=_NORMAL)
#                 _fmt_numeric(ws, ws.max_row)
#                 sub_item_rows.append(ws.max_row)

#             if sub_item_rows:
#                 ws.append(["", "", "", f"TOTAL - {sub['name']}", *_totals_from_rows(sub_item_rows)])
#                 _style_row(ws, ws.max_row, _FILL_SUBTOTAL, _BOLD)
#                 _fmt_numeric(ws, ws.max_row)
#                 sub_total_rows.append(ws.max_row)

#             sub_counter += 1

#         total_rows = sub_total_rows if sub_total_rows else direct_item_rows
#         if total_rows:
#             ws.append(["", "", "", f"TOTAL - {head['name']}", *_totals_from_rows(total_rows)])
#             _style_row(ws, ws.max_row, _FILL_HTOTAL,
#                        Font(bold=True, name="Calibri", size=10, color="003B63"))
#             _fmt_numeric(ws, ws.max_row)
#             head_total_rows.append(ws.max_row)
#             if head_name == "OPERATING EXPENSES":
#                 ws.append([])

#     for r in range(ws.max_row, 1, -1):
#         if (ws.cell(r, COL_HEAD).value == "COVID SUPPORT" and
#                 ws.cell(r-1, COL_HEAD).value is None):
#             ws.delete_rows(r-1); break

#     if head_total_rows:
#         ws.append(["", "", "", "GRAND TOTAL", *_totals_from_rows(head_total_rows)])
#         _style_row(ws, ws.max_row, _FILL_GRAND,
#                    Font(bold=True, color="FFFFFF", name="Calibri", size=10))
#         _fmt_numeric(ws, ws.max_row)

#     _auto_col_widths(ws)
#     return ws


# # ═══════════════════════════════════════════════════════════════════════════════
# # SHEET: Budget & Estimate
# # ═══════════════════════════════════════════════════════════════════════════════

def _sheet_be(wb, sheet_name, be_data, fy, plan_label, est_label,
              org_name="Azim Premji Foundation"):
    ws = wb.create_sheet(title=sheet_name[:31])
    entities = be_data or []
    n        = len(entities)
    thin = Side(style="thin")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Col layout (1-based): A unused, B=HEAD, C=TYPE, then 2*n entity cols, then 2 grand cols
    c_head     = 2
    c_type     = 3
    c_ent_start= 4              # first entity Plan col
    c_grand_p  = c_ent_start + n * 2       # Grand Total Plan
    c_grand_e  = c_grand_p + 1            # Grand Total Est
    last_col   = c_grand_e

    PF   = "ytd"
    EF   = "total_posted_amt_ytd"
    IF_E = "total_posted_amt"

    def _sec_v(entity, sname, f):
        for s in entity.get("actuals", []):
            if _is_grand_total_section(s): continue
            if s.get("name") == sname:
                return float(s.get(f, 0) or 0)
        return 0.0

    def _sub_v(entity, sname, subname, f):
        for s in entity.get("actuals", []):
            if _is_grand_total_section(s): continue
            if s.get("name") == sname:
                for sub in s.get("sub_heads", []):
                    if sub.get("name") == subname:
                        return float(sub.get(f, 0) or 0)
        return 0.0

    def _item_v(entity, iname, use_est=False):
        f = IF_E if use_est else PF
        for s in entity.get("actuals", []):
            if _is_grand_total_section(s): continue
            for item in s.get("items", []):
                if item.get("name") == iname:
                    return float(item.get(f, 0) or 0)
            for sub in s.get("sub_heads", []):
                for item in sub.get("items", []):
                    if item.get("name") == iname:
                        return float(item.get(f, 0) or 0)
        return 0.0

    def _grand_v(entity, f):
        # Sum all non-GRAND-TOTAL sections to avoid double-counting
        v = 0.0
        for s in entity.get("actuals", []):
            if _is_grand_total_section(s): continue
            v += float(s.get(f, 0) or 0)
        return v

    def _write_row(head_val, type_val, vp_list, ve_list, fill, font):
        ws.append(["", head_val, type_val] + [""] * (last_col - c_type))
        r = ws.max_row
        gp = ge = 0.0
        for ei in range(n):
            cp = ws.cell(r, c_ent_start + ei*2,     vp_list[ei] if vp_list[ei] else None)
            ce = ws.cell(r, c_ent_start + ei*2 + 1, ve_list[ei] if ve_list[ei] else None)
            cp.number_format = NUM_FMT_CR; ce.number_format = NUM_FMT_CR
            gp += vp_list[ei]; ge += ve_list[ei]
        ws.cell(r, c_grand_p, gp if gp else None).number_format = NUM_FMT_CR
        ws.cell(r, c_grand_e, ge if ge else None).number_format = NUM_FMT_CR
        for col in range(2, last_col+1):
            c = ws.cell(r, col)
            c.fill = fill; c.font = font; c.border = brd
            c.alignment = _LEFT if col <= c_type else _RIGHT

    # ── Sheet title ──
    ws.append(["", org_name])
    _merge(ws, 1, 2, 1, last_col)
    ws["B1"].font = Font(size=14, bold=True, name="Calibri", color="003B63")
    ws["B1"].alignment = _LEFT
    ws.append(["", f"Budget & Estimate – {fy}"])
    _merge(ws, 2, 2, 2, last_col)
    ws["B2"].font = Font(size=12, bold=True, name="Calibri"); ws["B2"].alignment = _LEFT
    ws.append([])

    # Note
    ws.cell(4, last_col, "₹ (Absolute)")
    ws.cell(4, last_col).font = _ITALIC; ws.cell(4, last_col).alignment = _RIGHT
    _merge(ws, 4, 2, 4, last_col)

    # ── Headers — style ALL cells BEFORE merging ──
    for col in range(2, last_col+1):
        for hr in [5, 6]:
            c = ws.cell(hr, col); c.fill = _FILL_HDR; c.font = _WHITE_BOLD
            c.border = brd; c.alignment = _CENTER

    ws.cell(5, c_head, "HEAD OF EXPENSE")
    ws.cell(5, c_type, "TYPE OF EXPENSE")
    _merge(ws, 5, c_head, 6, c_head)
    _merge(ws, 5, c_type, 6, c_type)

    for ei, entity in enumerate(entities):
        cs = c_ent_start + ei*2
        ws.cell(5, cs, (entity.get("label") or "").strip())
        _merge(ws, 5, cs, 5, cs+1)
        ws.cell(6, cs,   plan_label)
        ws.cell(6, cs+1, est_label)

    ws.cell(5, c_grand_p, "GRAND TOTAL")
    ws.cell(5, c_grand_p).fill = PatternFill("solid", fgColor="003B63")
    ws.cell(5, c_grand_e).fill = PatternFill("solid", fgColor="003B63")
    _merge(ws, 5, c_grand_p, 5, c_grand_e)
    ws.cell(6, c_grand_p, plan_label)
    ws.cell(6, c_grand_e, est_label)
    ws.cell(6, c_grand_p).fill = PatternFill("solid", fgColor="004F8B")
    ws.cell(6, c_grand_e).fill = PatternFill("solid", fgColor="004F8B")

    ws.freeze_panes = ws.cell(7, c_type+1).coordinate

    if not entities:
        _auto_col_widths(ws); return ws

    for sec in entities[0].get("actuals", []):
        if _is_grand_total_section(sec): continue
        sname = sec.get("name", "")

        # Section header
        ws.append(["", sname])
        r_sec = ws.max_row
        for col in range(2, last_col+1):
            c = ws.cell(r_sec, col)
            c.fill = _FILL_HEAD
            c.font = Font(bold=True, name="Calibri", size=10, color="003B63")
            c.border = brd; c.alignment = _LEFT
        _merge(ws, r_sec, 2, r_sec, last_col)

        for sub in sec.get("sub_heads", []):
            subname = sub.get("name","")
            vp = [_sub_v(e, sname, subname, PF)  for e in entities]
            ve = [_sub_v(e, sname, subname, EF)  for e in entities]
            _write_row(sname, subname, vp, ve, _FILL_SUBHEAD,
                       Font(bold=True, name="Calibri", size=10, color="7A3B00"))
            for item in sub.get("items", []):
                iname = item.get("name","")
                vp2 = [_item_v(e, iname, False) for e in entities]
                ve2 = [_item_v(e, iname, True)  for e in entities]
                _write_row("", iname, vp2, ve2, _FILL_WHITE, _NORMAL)

        for item in sec.get("items", []):
            iname = item.get("name","")
            vp = [_item_v(e, iname, False) for e in entities]
            ve = [_item_v(e, iname, True)  for e in entities]
            _write_row(sname, iname, vp, ve, _FILL_WHITE, _NORMAL)

        # Section total
        vp = [_sec_v(e, sname, PF) for e in entities]
        ve = [_sec_v(e, sname, EF) for e in entities]
        _write_row(f"TOTAL – {sname}", "", vp, ve, _FILL_HTOTAL,
                   Font(bold=True, name="Calibri", size=10, color="003B63"))

    # Grand total
    vp = [_grand_v(e, PF) for e in entities]
    ve = [_grand_v(e, EF) for e in entities]
    _write_row("GRAND TOTAL", "", vp, ve, _FILL_GRAND,
               Font(bold=True, color="FFFFFF", name="Calibri", size=10))

    _auto_col_widths(ws)
    return ws


# # ═══════════════════════════════════════════════════════════════════════════════
# # FY HELPERS
# # ═══════════════════════════════════════════════════════════════════════════════

# def _fy_labels(fy):
#     parts    = (fy or "2025-26").split("-")
#     start_yy = (parts[0] or "2025")[-2:]
#     end_yy   = (parts[1] if len(parts) > 1 else "26")[-2:]
#     prev_s   = str(int(start_yy) - 1).zfill(2)
#     prev_e   = str(int(end_yy)   - 1).zfill(2)
#     return {
#         "plan": f"FY{start_yy}-{end_yy} Plan",
#         "est":  f"FY{prev_s}-{prev_e} Estimate",
#     }

# def _prev_fy(fy):
#     parts = (fy or "2025-26").split("-")
#     s = int(parts[0] or 2025) - 1
#     e = int(parts[1] or 26)   - 1
#     return f"{s}-{str(e).zfill(2)}"


# # ═══════════════════════════════════════════════════════════════════════════════
# # WHITELISTED API ENDPOINTS
# # ═══════════════════════════════════════════════════════════════════════════════

# @frappe.whitelist()
# def export_ppt(financial_year, ppt_rows, prev_ppt_rows,
#                budget_label, est_label, prev_budget_label, prev_est_label):
#     fy      = financial_year or "2025-26"
#     prev_fy = _prev_fy(fy)
#     wb = Workbook(); wb.remove(wb.active)
#     _sheet_ppt_combined(
#         wb, "Foundation Metrics",
#         json.loads(ppt_rows),      budget_label,      est_label,
#         json.loads(prev_ppt_rows), prev_budget_label, prev_est_label,
#         fy=fy, prev_fy=prev_fy,
#     )
#     return {"filename": f"Foundation_Metrics_{fy}.xlsx", "data": _wb_to_b64(wb)}


# @frappe.whitelist()
# def export_summary_inr(financial_year, summary_data):
#     fy = financial_year or "2025-26"
#     wb = Workbook(); wb.remove(wb.active)
#     _sheet_summary_inr(wb, "Summary in INR", json.loads(summary_data), fy)
#     return {"filename": f"Summary_INR_{fy}.xlsx", "data": _wb_to_b64(wb)}


# @frappe.whitelist()
# def export_headcount(financial_year, headcount_data):
#     fy = financial_year or "2025-26"
#     wb = Workbook(); wb.remove(wb.active)
#     _sheet_headcount(wb, "Headcount", json.loads(headcount_data), fy)
#     return {"filename": f"Headcount_{fy}.xlsx", "data": _wb_to_b64(wb)}


# @frappe.whitelist()
# def export_annual(financial_year, annual_data):
#     fy = financial_year or "2025-26"
#     wb = Workbook(); wb.remove(wb.active)
#     _sheet_annual(wb, "Annual Budget Consolidated", json.loads(annual_data), fy)
#     return {"filename": f"Annual_Budget_Consolidated_{fy}.xlsx", "data": _wb_to_b64(wb)}


# @frappe.whitelist()
# def export_estimate(financial_year, estimate_data):
#     fy = financial_year or "2025-26"
#     wb = Workbook(); wb.remove(wb.active)
#     _sheet_estimate(wb, "Estimate Consolidated", json.loads(estimate_data), fy)
#     return {"filename": f"Estimate_Consolidated_{fy}.xlsx", "data": _wb_to_b64(wb)}


# @frappe.whitelist()
# def export_budget_estimate(financial_year, be_data):
#     fy     = financial_year or "2025-26"
#     labels = _fy_labels(fy)
#     wb = Workbook(); wb.remove(wb.active)
#     _sheet_be(wb, "Budget & Estimate", json.loads(be_data),
#               fy, labels["plan"], labels["est"])
#     return {"filename": f"Budget_and_Estimate_{fy}.xlsx", "data": _wb_to_b64(wb)}


# @frappe.whitelist()
# def export_all(financial_year,
#                ppt_rows, prev_ppt_rows,
#                budget_label, est_label, prev_budget_label, prev_est_label,
#                summary_data, headcount_data,
#                annual_data, estimate_data, be_data):
#     fy      = financial_year or "2025-26"
#     prev_fy = _prev_fy(fy)
#     labels  = _fy_labels(fy)
#     wb = Workbook(); wb.remove(wb.active)

#     _sheet_ppt_combined(
#         wb, "Foundation Metrics",
#         json.loads(ppt_rows),      budget_label,      est_label,
#         json.loads(prev_ppt_rows), prev_budget_label, prev_est_label,
#         fy=fy, prev_fy=prev_fy,
#     )
#     _sheet_summary_inr(wb, "Summary in INR",            json.loads(summary_data),    fy)
#     _sheet_headcount(  wb, "Headcount",                 json.loads(headcount_data),  fy)
#     _sheet_annual(     wb, "Annual Budget Consolidated", json.loads(annual_data),     fy)
#     _sheet_estimate(   wb, "Estimate Consolidated",      json.loads(estimate_data),   fy)
#     _sheet_be(         wb, "Budget & Estimate",          json.loads(be_data),
#               fy, labels["plan"], labels["est"])

#     return {
#         "filename": f"Foundation_Consolidated_Budget_{fy}.xlsx",
#         "data": _wb_to_b64(wb)
#     }



import io
import base64
import json

import frappe
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════════════════════════════
# SHARED STYLE CONSTANTS  — colours match JS CSS variables exactly
# ═══════════════════════════════════════════════════════════════════════════════

_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_RIGHT  = Alignment(horizontal="right",  vertical="center")

_BOLD       = Font(bold=True,  name="Calibri", size=10)
_NORMAL     = Font(bold=False, name="Calibri", size=10)
_WHITE_BOLD = Font(bold=True,  color="FFFFFF", name="Calibri", size=10)
_ITALIC     = Font(bold=False, italic=True,    name="Calibri", size=9, color="555555")

# Exact colours from JS --blue-mid, --orange, --blue-light, --orange-light, --blue-dark
_FILL_BLUE_MID    = PatternFill("solid", fgColor="0076B6")   # main header
_FILL_ORANGE      = PatternFill("solid", fgColor="F26B21")   # sub-header
_FILL_BLUE_LIGHT  = PatternFill("solid", fgColor="E9F4FB")   # cb-row-head
_FILL_ORANGE_LIGHT= PatternFill("solid", fgColor="FFF3E6")   # cb-row-sub
_FILL_BLUE_DARK   = PatternFill("solid", fgColor="003B63")   # cb-row-grand
_FILL_HTOTAL      = PatternFill("solid", fgColor="E8F0FA")   # ppt-total-row / sinr-total-row
_FILL_GT          = PatternFill("solid", fgColor="DDEAF7")   # sinr-gt-plan / sinr-gt-act
_FILL_WHITE       = PatternFill("solid", fgColor="FFFFFF")
_FILL_FAFAFA      = PatternFill("solid", fgColor="FAFAFA")   # sinr-brkdwn-act
_FILL_SPACER      = PatternFill("solid", fgColor="F4F6F8")   # sinr-spacer

_THIN   = Side(style="thin")
_MEDIUM = Side(style="medium")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_THICK  = Border(left=_THIN, right=_THIN, top=_MEDIUM, bottom=_THIN)

# Number formats — JS fmtCr divides by 1e7 and shows 2dp
NUM_FMT_CR  = '#,##0.00'   # Crore values
NUM_FMT_INT = '#,##0'      # Headcount whole numbers
NUM_FMT_PCT = '0.0%'       # percentage (stored as decimal, e.g. 0.357)
NUM_FMT_RAW = '#,##0.00'   # Annual/Estimate raw rupees


# ═══════════════════════════════════════════════════════════════════════════════
# LOW-LEVEL HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def _to_cr(v):
    """Raw rupees → Crores (÷ 1,00,00,000), matching JS fmtCr."""
    return round(float(v or 0) / 1e7, 2)


def _fv(v):
    return float(v or 0)


def _norm(s):
    return (s or "").replace("  ", " ").strip().upper()


def _is_gt(sec):
    nm = _norm(sec.get("name", ""))
    return sec.get("sequence_id") == 9999 or nm == "GRAND TOTAL"


def _is_consolidated(entry):
    return (entry.get("sequence_id") == 9999
            or (entry.get("table_name") or "").upper() == "CONSOLIDATED")


def _wb_to_b64(wb):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return base64.b64encode(buf.read()).decode("utf-8")


def _auto_widths(ws, min_w=8, max_w=55):
    for col in range(1, ws.max_column + 1):
        best = min_w
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row, col).value
            if v is not None:
                s = str(v)
                if s.startswith("="):
                    s = "999,999,999.00"
                best = max(best, len(s))
        ws.column_dimensions[get_column_letter(col)].width = min(best + 2, max_w)


def _cell(ws, row, col, value=None, fill=None, font=None, align=None, border=None, fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    if fill:   c.fill   = fill
    if font:   c.font   = font
    if align:  c.alignment = align
    if border: c.border = border
    if fmt:    c.number_format = fmt
    return c


def _row_style(ws, row, cols, fill=None, font=None, align=_RIGHT, border=_BORDER):
    for col in cols:
        c = ws.cell(row, col)
        if fill:  c.fill  = fill
        if font:  c.font  = font
        if align: c.alignment = align
        c.border = border


def _merge(ws, r1, c1, r2, c2):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)


def _fy_labels(fy):
    p  = (fy or "2025-26").split("-")
    sY = (p[0] or "2025")[-2:]
    eY = (p[1] if len(p) > 1 else "26")[-2:]
    ps = str(int(sY) - 1).zfill(2)
    pe = str(int(eY) - 1).zfill(2)
    return {"plan": f"FY{sY}-{eY} Plan", "est": f"FY{ps}-{pe} Estimate"}


def _prev_fy(fy):
    p = (fy or "2025-26").split("-")
    s = int(p[0] or 2025) - 1
    e = int(p[1] or 26) - 1
    return f"{s}-{str(e).zfill(2)}"


def _sheet_title(ws, org, subtitle, last_col):
    """Write 2 title rows at top of sheet."""
    c = ws.cell(1, 2, org)
    c.font = Font(size=13, bold=True, name="Calibri", color="003B63")
    c.alignment = _LEFT
    _merge(ws, 1, 2, 1, last_col)
    c2 = ws.cell(2, 2, subtitle)
    c2.font = Font(size=11, bold=True, name="Calibri")
    c2.alignment = _LEFT
    _merge(ws, 2, 2, 2, last_col)
    ws.append([])


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — Foundation Level / Overall Metrics  (PPT tab)
#
# JS display:
#   Two tables stacked: current FY and previous FY
#   Columns: Unit | Budget(Opex Capex Total) | Actual(Opex Capex Total)
#   Total uses GRAND TOTAL section (seq 9999) directly for the Total column
#   Opex / Capex come from OPERATING/CAPITAL named sections
#   Last row = "Total" with blue background
# ═══════════════════════════════════════════════════════════════════════════════

def _extract_opex_capex(sections, field):
    """
    Match JS extractVals: sum OPERATING→opex, CAPITAL→capex.
    Fallback to GRAND TOTAL section if no breakdown found.
    """
    opex = capex = 0.0
    has_breakdown = False
    for sec in (sections or []):
        nm = _norm(sec.get("name", ""))
        if sec.get("sequence_id") == 9999 or nm == "GRAND TOTAL":
            continue
        if "OPERATING" in nm:
            opex += _fv(sec.get(field)); has_breakdown = True
        elif "CAPITAL" in nm:
            capex += _fv(sec.get(field)); has_breakdown = True
    if not has_breakdown:
        for sec in (sections or []):
            nm = _norm(sec.get("name", ""))
            if sec.get("sequence_id") == 9999 or nm == "GRAND TOTAL":
                opex += _fv(sec.get(field))
        if not opex:
            for sec in (sections or []):
                nm = _norm(sec.get("name", ""))
                if sec.get("sequence_id") != 9999 and nm != "GRAND TOTAL":
                    opex += _fv(sec.get(field))
    return opex, capex


def _extract_total(sections, field):
    """
    Match JS extractTotal: use GRAND TOTAL section (seq 9999) directly.
    Fallback: sum all sections.
    """
    gt = 0.0
    found = False
    for sec in (sections or []):
        nm = _norm(sec.get("name", ""))
        if sec.get("sequence_id") == 9999 or nm == "GRAND TOTAL":
            gt += _fv(sec.get(field)); found = True
    if not found:
        for sec in (sections or []):
            gt += _fv(sec.get(field))
    return gt


def _sheet_ppt(wb, ppt_rows, budget_label, est_label,
               prev_rows, prev_budget_label, prev_est_label, fy, prev_fy,
               org="Azim Premji Foundation"):
    ws = wb.create_sheet("Foundation Metrics")
    COLS = list(range(2, 9))   # B–H  (7 data cols)

    _sheet_title(ws, org, f"Foundation Level / Overall Metrics – {fy}", 8)

    def _write_block(start_row, rows, b_lbl, e_lbl, title):
        # Title bar
        ws.cell(start_row, 2, title)
        for col in COLS:
            c = ws.cell(start_row, col)
            c.fill = _FILL_BLUE_LIGHT
            c.font = Font(bold=True, name="Calibri", size=11, color="003B63")
            c.border = _BORDER
            c.alignment = _LEFT
        _merge(ws, start_row, 2, start_row, 8)

        # Currency note
        nr = start_row + 1
        ws.cell(nr, 8, "₹ Cr.").font = _ITALIC
        ws.cell(nr, 8).alignment = Alignment(horizontal="right")

        # Header row 1 (blue)
        h1 = start_row + 2
        for col in COLS:
            c = ws.cell(h1, col)
            c.fill = _FILL_BLUE_MID; c.font = _WHITE_BOLD
            c.border = _BORDER; c.alignment = _CENTER
        ws.cell(h1, 2, "Unit")
        ws.cell(h1, 3, b_lbl)
        ws.cell(h1, 6, e_lbl)
        _merge(ws, h1, 2, h1 + 1, 2)
        _merge(ws, h1, 3, h1, 5)
        _merge(ws, h1, 6, h1, 8)

        # Header row 2 (orange)
        h2 = start_row + 3
        for col in COLS:
            c = ws.cell(h2, col)
            c.fill = _FILL_ORANGE; c.font = _WHITE_BOLD
            c.border = _BORDER; c.alignment = _CENTER
        for col, lbl in zip(range(3, 9), ["Opex", "Capex", "Total", "Opex", "Capex", "Total"]):
            ws.cell(h2, col, lbl)

        # Data rows
        dr = start_row + 4
        for row in rows:
            is_total = row.get("is_total", False)
            fill = _FILL_HTOTAL if is_total else _FILL_WHITE
            font = _BOLD if is_total else _NORMAL
            bO = _to_cr(row.get("bOpex",  0))
            bC = _to_cr(row.get("bCapex", 0))
            bT = _to_cr(row.get("bTotal", 0)) if row.get("bTotal") else _to_cr(_fv(row.get("bOpex", 0)) + _fv(row.get("bCapex", 0)))
            eO = _to_cr(row.get("eOpex",  0))
            eC = _to_cr(row.get("eCapex", 0))
            eT = _to_cr(row.get("eTotal", 0)) if row.get("eTotal") else _to_cr(_fv(row.get("eOpex", 0)) + _fv(row.get("eCapex", 0)))
            for col, val in enumerate([row.get("label", ""), bO or None, bC or None, bT or None,
                                        eO or None, eC or None, eT or None], start=2):
                c = ws.cell(dr, col, val)
                c.fill = fill; c.font = font; c.border = _BORDER
                c.alignment = _LEFT if col == 2 else _RIGHT
                if col > 2:
                    c.number_format = NUM_FMT_CR
            dr += 1
        return dr

    next_r = _write_block(4, ppt_rows, budget_label, est_label,
                          f"Overall Foundation – {fy} Budget vs {prev_fy} Actual")
    _write_block(next_r + 1, prev_rows, prev_budget_label, prev_est_label,
                 f"Overall Foundation – {prev_fy} Budget vs {prev_fy} Actual")

    _auto_widths(ws)
    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — Summary in INR  (SummaryINR tab)
#
# JS display:
#   Table A: Unit / Function | Plan(Opex Capex Total) | Act(Opex Capex Total)
#     - Each unit row from API
#     - "Grand Total" row from CONSOLIDATED TOTAL entry (seq 9999)
#     - Values in Crores (÷ 1e7)
#
#   Table B: Breakdown — rows per unit, plan + actual rows
#     - Columns: Unit | Opex sub-heads… | Opex Total | Capex | Total
#     - Grand Total uses CONSOLIDATED TOTAL values
# ═══════════════════════════════════════════════════════════════════════════════

def _sinr_extract_a(actuals):
    """Match JS extractA: sum OPERATING→opex_plan/act, CAPITAL→capex_plan/act."""
    op_p = cp_p = op_a = cp_a = 0.0
    for sec in (actuals or []):
        nm = _norm(sec.get("name", ""))
        if "OPERATING" in nm:
            op_p += _fv(sec.get("ytd"))
            op_a += _fv(sec.get("total_posted_amt_ytd"))
        elif "CAPITAL" in nm:
            cp_p += _fv(sec.get("ytd"))
            cp_a += _fv(sec.get("total_posted_amt_ytd"))
    return op_p, cp_p, op_a, cp_a


def _sinr_get_consolidated(data):
    """
    Match JS getConsolidatedTotals: find seq_id 9999 / table CONSOLIDATED entry.
    Returns dict with opex_plan, capex_plan, total_plan, opex_act, capex_act, total_act.
    """
    for e in (data or []):
        if _is_consolidated(e):
            r = {"opex_plan": 0, "capex_plan": 0, "total_plan": 0,
                 "opex_act":  0, "capex_act":  0, "total_act":  0}
            for a in (e.get("actuals") or []):
                nm = _norm(a.get("name", ""))
                if nm == "OPEX TOTAL":
                    r["opex_plan"] += _fv(a.get("ytd"))
                    r["opex_act"]  += _fv(a.get("total_posted_amt_ytd"))
                elif nm == "CAPEX TOTAL":
                    r["capex_plan"] += _fv(a.get("ytd"))
                    r["capex_act"]  += _fv(a.get("total_posted_amt_ytd"))
                elif nm == "OVERALL GRAND TOTAL":
                    r["total_plan"] = _fv(a.get("ytd"))
                    r["total_act"]  = _fv(a.get("total_posted_amt_ytd"))
            if not r["total_plan"] and not r["total_act"]:
                r["total_plan"] = r["opex_plan"] + r["capex_plan"]
                r["total_act"]  = r["opex_act"]  + r["capex_act"]
            return r
    return None


def _sheet_summary_inr(wb, api_data, fy, org="Azim Premji Foundation"):
    ws = wb.create_sheet("Summary in INR")
    fp = (fy or "2025-26").split("-")
    p_lbl = f"{fy} Budget"
    a_lbl = f"{int(fp[0])-1}-{str(int(fp[1])-1).zfill(2)} Est"

    _sheet_title(ws, org, f"Summary in INR – {fy}", 8)

    sorted_data = sorted((api_data or []), key=lambda e: e.get("sequence_id", 0))
    # Exclude consolidated total entry from unit rows
    unit_rows  = [e for e in sorted_data if not _is_consolidated(e)]
    normal     = [e for e in unit_rows if "covid" not in (e.get("label") or "").lower()]
    covid_rows = [e for e in unit_rows if "covid"     in (e.get("label") or "").lower()]
    ct         = _sinr_get_consolidated(api_data)

    cur = 4  # current row

    # ── Table A header ──────────────────────────────────────────────────────
    ws.cell(cur, 2, "A. Unit Wise Plan")
    ws.cell(cur, 2).font = Font(bold=True, name="Calibri", size=11, underline="single")
    cur += 1
    ws.cell(cur, 8, "₹ Cr."); ws.cell(cur, 8).font = _ITALIC
    ws.cell(cur, 8).alignment = Alignment(horizontal="right"); cur += 1

    h1 = cur
    for col in range(2, 9):
        c = ws.cell(h1, col)
        c.fill = _FILL_BLUE_MID; c.font = _WHITE_BOLD; c.border = _BORDER; c.alignment = _CENTER
    ws.cell(h1, 2, "Unit / Function"); ws.cell(h1, 3, p_lbl); ws.cell(h1, 6, a_lbl)
    _merge(ws, h1, 2, h1 + 1, 2); _merge(ws, h1, 3, h1, 5); _merge(ws, h1, 6, h1, 8)
    cur += 1

    h2 = cur
    for col in range(2, 9):
        c = ws.cell(h2, col)
        c.fill = _FILL_ORANGE; c.font = _WHITE_BOLD; c.border = _BORDER; c.alignment = _CENTER
    for col, lbl in zip(range(3, 9), ["Opex", "Capex", "Total", "Opex", "Capex", "Total"]):
        ws.cell(h2, col, lbl)
    cur += 1

    def _write_unit_row_a(entry, fill=_FILL_WHITE, font=_NORMAL, label_override=None):
        nonlocal cur
        op_p, cp_p, op_a, cp_a = _sinr_extract_a(entry.get("actuals", []))
        is_sub = entry.get("is_this_sub_item") == 1
        lbl    = label_override or (entry.get("label") or "").strip()
        indent = "    " if is_sub else ""
        for col, val in enumerate([indent + lbl,
                                    _to_cr(op_p), _to_cr(cp_p), _to_cr(op_p + cp_p),
                                    _to_cr(op_a), _to_cr(cp_a), _to_cr(op_a + cp_a)], start=2):
            c = ws.cell(cur, col, val if (col == 2 or val != 0.0) else None)
            c.fill = fill; c.font = font; c.border = _BORDER
            c.alignment = _LEFT if col == 2 else _RIGHT
            if col > 2: c.number_format = NUM_FMT_CR
        cur += 1

    for e in normal:
        _write_unit_row_a(e)

    for e in covid_rows:
        _write_unit_row_a(e)

    # Grand Total row (from CONSOLIDATED TOTAL entry)
    if ct:
        op_p = ct["opex_plan"]; cp_p = ct["capex_plan"]; tp = ct["total_plan"]
        op_a = ct["opex_act"];  cp_a = ct["capex_act"];  ta = ct["total_act"]
    else:
        op_p = cp_p = op_a = cp_a = 0.0
        for e in unit_rows:
            if e.get("is_this_sub_item") != 1:
                o, c2, oa, ca = _sinr_extract_a(e.get("actuals", []))
                op_p += o; cp_p += c2; op_a += oa; cp_a += ca
        tp = op_p + cp_p; ta = op_a + cp_a

    for col, val in enumerate(["Grand Total",
                                _to_cr(op_p), _to_cr(cp_p), _to_cr(tp),
                                _to_cr(op_a), _to_cr(cp_a), _to_cr(ta)], start=2):
        c = ws.cell(cur, col, val)
        c.fill = _FILL_BLUE_DARK
        c.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        c.border = _BORDER
        c.alignment = _LEFT if col == 2 else _RIGHT
        if col > 2: c.number_format = NUM_FMT_CR
    cur += 2

    # ── Table B header ──────────────────────────────────────────────────────
    ws.cell(cur, 2, "B. Breakdown of Unit Wise Plan")
    ws.cell(cur, 2).font = Font(bold=True, name="Calibri", size=11, underline="single")
    cur += 1
    ws.cell(cur, 8, "₹ Cr."); ws.cell(cur, 8).font = _ITALIC
    ws.cell(cur, 8).alignment = Alignment(horizontal="right"); cur += 1

    # Collect sub-head names from OPERATING section (match JS getSubNames)
    main_entries = [e for e in unit_rows if e.get("is_this_sub_item") != 1]
    seen_sh = {}; sub_head_names = []
    for e in main_entries:
        for sec in (e.get("actuals") or []):
            if "OPERATING" not in _norm(sec.get("name", "")): continue
            for sh in (sec.get("sub_heads") or []):
                n = (sh.get("name") or "").strip()
                if n and n not in seen_sh: seen_sh[n] = True; sub_head_names.append(n)

    n_sh    = len(sub_head_names)
    c_start = 3                     # first sub-head col
    c_opx_t = c_start + n_sh        # Opex Total col
    c_capex = c_opx_t + 1           # Capex col
    c_total = c_capex + 1           # Total col
    last_c  = c_total

    # Header row 1
    bh1 = cur
    for col in range(2, last_c + 1):
        c = ws.cell(bh1, col)
        c.fill = _FILL_BLUE_MID; c.font = _WHITE_BOLD; c.border = _BORDER; c.alignment = _CENTER
    ws.cell(bh1, 2, "Unit / Function"); _merge(ws, bh1, 2, bh1 + 1, 2)
    ws.cell(bh1, c_start, "Operating Expenses")
    _merge(ws, bh1, c_start, bh1, c_opx_t)
    ws.cell(bh1, c_capex, "Capex"); _merge(ws, bh1, c_capex, bh1 + 1, c_capex)
    ws.cell(bh1, c_total, "Total"); _merge(ws, bh1, c_total, bh1 + 1, c_total)
    cur += 1

    # Header row 2
    bh2 = cur
    for col in range(2, last_c + 1):
        c = ws.cell(bh2, col)
        c.fill = _FILL_ORANGE; c.font = _WHITE_BOLD; c.border = _BORDER; c.alignment = _CENTER
    for i, shn in enumerate(sub_head_names):
        ws.cell(bh2, c_start + i, shn)
    ws.cell(bh2, c_opx_t, "Total")
    cur += 1

    def _sh_val(actuals, shname, field):
        v = 0.0
        for sec in (actuals or []):
            if "OPERATING" not in _norm(sec.get("name", "")): continue
            for sh in (sec.get("sub_heads") or []):
                if sh.get("name", "").strip() == shname: v += _fv(sh.get(field))
        return v

    def _opex_total(actuals, field):
        v = 0.0
        for sec in (actuals or []):
            if "OPERATING" not in _norm(sec.get("name", "")): continue
            v += _fv(sec.get(field))
        return v

    def _capex_total(actuals, field):
        v = 0.0
        for sec in (actuals or []):
            if "CAPITAL" not in _norm(sec.get("name", "")): continue
            v += _fv(sec.get(field))
        return v

    gt_sh_p = {n: 0.0 for n in sub_head_names}
    gt_sh_a = {n: 0.0 for n in sub_head_names}
    gt_op_p = gt_op_a = gt_cp_p = gt_cp_a = 0.0

    for entry in main_entries:
        act = entry.get("actuals") or []
        lbl = (entry.get("label") or "").strip()

        # Unit header row — sinr-unit-hdr (blue light)
        for col in range(2, last_c + 1):
            c = ws.cell(cur, col)
            c.fill = _FILL_BLUE_LIGHT; c.font = _BOLD; c.border = _BORDER
            c.alignment = _LEFT if col == 2 else _CENTER
        ws.cell(cur, 2, lbl); _merge(ws, cur, 3, cur, last_c)
        cur += 1

        # Plan row — sinr-brkdwn-plan
        op_p = _opex_total(act, "ytd"); cp_p = _capex_total(act, "ytd")
        for col in range(2, last_c + 1):
            c = ws.cell(cur, col)
            c.fill = _FILL_WHITE; c.font = _NORMAL; c.border = _BORDER
            c.alignment = _LEFT if col == 2 else _RIGHT
        ws.cell(cur, 2, f"  - {p_lbl}")
        for i, shn in enumerate(sub_head_names):
            v = _sh_val(act, shn, "ytd")
            gt_sh_p[shn] += v
            c = ws.cell(cur, c_start + i, _to_cr(v) or None)
            c.number_format = NUM_FMT_CR
        ws.cell(cur, c_opx_t, _to_cr(op_p) or None).number_format = NUM_FMT_CR
        ws.cell(cur, c_capex, _to_cr(cp_p) or None).number_format = NUM_FMT_CR
        ws.cell(cur, c_total, _to_cr(op_p + cp_p) or None).number_format = NUM_FMT_CR
        gt_op_p += op_p; gt_cp_p += cp_p
        cur += 1

        # Actual row — sinr-brkdwn-act
        op_a = _opex_total(act, "total_posted_amt_ytd"); cp_a = _capex_total(act, "total_posted_amt_ytd")
        for col in range(2, last_c + 1):
            c = ws.cell(cur, col)
            c.fill = _FILL_FAFAFA; c.font = _NORMAL; c.border = _BORDER
            c.alignment = _LEFT if col == 2 else _RIGHT
        ws.cell(cur, 2, f"  - {a_lbl}")
        for i, shn in enumerate(sub_head_names):
            v = _sh_val(act, shn, "total_posted_amt_ytd")
            gt_sh_a[shn] += v
            c = ws.cell(cur, c_start + i, _to_cr(v) or None)
            c.number_format = NUM_FMT_CR
        ws.cell(cur, c_opx_t, _to_cr(op_a) or None).number_format = NUM_FMT_CR
        ws.cell(cur, c_capex, _to_cr(cp_a) or None).number_format = NUM_FMT_CR
        ws.cell(cur, c_total, _to_cr(op_a + cp_a) or None).number_format = NUM_FMT_CR
        gt_op_a += op_a; gt_cp_a += cp_a
        cur += 1

        # Spacer
        for col in range(2, last_c + 1):
            ws.cell(cur, col).fill = _FILL_SPACER; ws.cell(cur, col).border = _BORDER
        cur += 1

    # Grand Total label row — cb-row-grand (dark blue)
    for col in range(2, last_c + 1):
        c = ws.cell(cur, col)
        c.fill = _FILL_BLUE_DARK
        c.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        c.border = _BORDER; c.alignment = _LEFT if col == 2 else _CENTER
    ws.cell(cur, 2, "Grand Total"); _merge(ws, cur, 3, cur, last_c)
    cur += 1

    # Use CONSOLIDATED TOTAL values for GT if available (match JS finalOP etc.)
    fin_op_p = ct["opex_plan"]  if ct else gt_op_p
    fin_op_a = ct["opex_act"]   if ct else gt_op_a
    fin_cp_p = ct["capex_plan"] if ct else gt_cp_p
    fin_cp_a = ct["capex_act"]  if ct else gt_cp_a
    fin_tp   = ct["total_plan"] if ct else (gt_op_p + gt_cp_p)
    fin_ta   = ct["total_act"]  if ct else (gt_op_a + gt_cp_a)

    # GT Plan row — sinr-gt-plan
    for col in range(2, last_c + 1):
        c = ws.cell(cur, col); c.fill = _FILL_GT; c.font = _NORMAL
        c.border = _BORDER; c.alignment = _LEFT if col == 2 else _RIGHT
    ws.cell(cur, 2, f"  - {p_lbl}")
    for i, shn in enumerate(sub_head_names):
        ws.cell(cur, c_start + i, _to_cr(gt_sh_p[shn]) or None).number_format = NUM_FMT_CR
    ws.cell(cur, c_opx_t, _to_cr(fin_op_p) or None).number_format = NUM_FMT_CR
    ws.cell(cur, c_capex, _to_cr(fin_cp_p) or None).number_format = NUM_FMT_CR
    ws.cell(cur, c_total, _to_cr(fin_tp)   or None).number_format = NUM_FMT_CR
    cur += 1

    # GT Actual row — sinr-gt-act
    for col in range(2, last_c + 1):
        c = ws.cell(cur, col); c.fill = _FILL_GT; c.font = _NORMAL
        c.border = _BORDER; c.alignment = _LEFT if col == 2 else _RIGHT
    ws.cell(cur, 2, f"  - {a_lbl}")
    for i, shn in enumerate(sub_head_names):
        ws.cell(cur, c_start + i, _to_cr(gt_sh_a[shn]) or None).number_format = NUM_FMT_CR
    ws.cell(cur, c_opx_t, _to_cr(fin_op_a) or None).number_format = NUM_FMT_CR
    ws.cell(cur, c_capex, _to_cr(fin_cp_a) or None).number_format = NUM_FMT_CR
    ws.cell(cur, c_total, _to_cr(fin_ta)   or None).number_format = NUM_FMT_CR

    _auto_widths(ws)
    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — Headcount
#
# JS display — 5 sections (matching renderSummary + 4 tables):
#   1. Headcount Summary  — Unit | Avg H/C y[-2] | Avg H/C y[-1] | % | Est Opex | Plan Opex | %
#   2. Closing H/C        — Unit | [hc per year straight from API]
#   3. Average H/C        — Unit | [(prev+curr)/2 per year;  i=0 → hc[0]/2]
#   4. Increase Closing % — transitions (curr/prev)-1
#   5. Increase Avg H/C % — transitions using avg values
# ═══════════════════════════════════════════════════════════════════════════════

def _sheet_headcount(wb, payload, fy, org="Azim Premji Foundation"):
    ws = wb.create_sheet("Headcount")

    if isinstance(payload, dict):
        records   = payload.get("headcount_data") or []
        plan_data = payload.get("plan_data") or []
    else:
        records   = payload or []
        plan_data = []

    # Build opex map from plan_data — match JS buildOpexMap
    def _norm_lbl(s): return (s or "").lower().replace("  ", " ").strip()

    opex_map = {}
    for p in (plan_data or []):
        lbl = _norm_lbl(p.get("label", ""))
        op  = None
        for a in (p.get("actuals") or []):
            nm = (a.get("name") or "").strip()
            if nm in ("OPERATING  EXPENSES", "OPERATING EXPENSES"):
                op = a; break
        opex_map[lbl] = {
            "est" : _fv(op.get("total_posted_amt_ytd", 0) if op else 0) / 1e7,
            "plan": _fv(op.get("ytd", 0)                if op else 0) / 1e7,
        }

    # Transform records — match JS transform (straight from API)
    sorted_recs = sorted((records or []), key=lambda r: r.get("financial_year", ""))
    yrs = [r.get("financial_year", "") for r in sorted_recs]

    um = {}
    for rec in sorted_recs:
        for u in (rec.get("units") or []):
            uid = str(u.get("unit", ""))
            if uid not in um:
                um[uid] = {"id": uid, "desc": "", "hc": {}, "seq": int(uid) if uid.isdigit() else 999}
            um[uid]["hc"][rec["financial_year"]] = _fv(u.get("total_headcount") or 0)
            if rec["financial_year"] == yrs[-1]:
                um[uid]["desc"] = (u.get("unit_description") or u.get("description") or "").strip()
    units  = sorted(um.values(), key=lambda u: u["seq"])
    totals = {r["financial_year"]: _fv(r.get("total_head_count") or r.get("total_headcount") or 0)
              for r in sorted_recs}

    # avgHC: i=0 → hc[0]/2; i>0 → (prev+curr)/2 — match JS
    def _avg_hc(hc, i):
        c = hc.get(yrs[i])
        if i == 0:
            return c / 2 if c else None
        p = hc.get(yrs[i - 1])
        return (p + c) / 2 if (p is not None and c is not None) else None

    def _avg_tot(i):
        c = totals.get(yrs[i])
        if i == 0:
            return c / 2 if c else None
        p = totals.get(yrs[i - 1])
        return (p + c) / 2 if (p is not None and c is not None) else None

    def _fym(fy_str):
        p = (fy_str or "").split("-")
        return f"31-Mar-{p[1][-2:] if len(p) > 1 else (p[0] or '')[-2:]}"

    max_col = max(2 + len(yrs), 8)
    _sheet_title(ws, org, f"Headcount – {fy}", max_col)
    cur = 4

    def _section_hdr(title):
        nonlocal cur
        c = ws.cell(cur, 2, title)
        c.font = Font(bold=True, name="Calibri", size=11, underline="single", color="003B63")
        cur += 1

    def _col_hdr(labels):
        nonlocal cur
        for ci, lbl in enumerate(["Unit"] + labels, start=2):
            c = ws.cell(cur, ci, lbl)
            c.fill = _FILL_BLUE_MID; c.font = _WHITE_BOLD
            c.border = _BORDER; c.alignment = _CENTER
        cur += 1

    def _data_row(label, values, fill=_FILL_WHITE, font=_NORMAL, fmt=NUM_FMT_INT, is_grand=False):
        nonlocal cur
        gf = _FILL_BLUE_DARK
        gfont = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        ws.cell(cur, 2, label).fill = gf if is_grand else fill
        ws.cell(cur, 2).font = gfont if is_grand else font
        ws.cell(cur, 2).border = _BORDER; ws.cell(cur, 2).alignment = _LEFT
        for ci, val in enumerate(values, start=3):
            c = ws.cell(cur, ci)
            c.fill = gf if is_grand else fill
            c.font = gfont if is_grand else font
            c.border = _BORDER; c.alignment = _RIGHT
            if val is not None:
                if fmt == NUM_FMT_PCT:
                    c.value = val / 100   # store as decimal for % format
                    c.number_format = NUM_FMT_PCT
                else:
                    c.value = round(val) if fmt == NUM_FMT_INT else val
                    if fmt: c.number_format = fmt
        cur += 1

    # ── 1. Headcount Summary ────────────────────────────────────────────────
    _section_hdr("Headcount Summary")
    if len(yrs) >= 2:
        i1 = len(yrs) - 2; i2 = len(yrs) - 1
        _col_hdr([yrs[i1], yrs[i2], "%", f"{yrs[i1]} Est", f"{yrs[i2]} Plan", "%"])
        tot_est = tot_plan = 0.0
        for u in units:
            a1 = _avg_hc(u["hc"], i1); a2 = _avg_hc(u["hc"], i2)
            o  = opex_map.get(_norm_lbl(u["desc"]), {"est": 0.0, "plan": 0.0})
            tot_est += o["est"]; tot_plan += o["plan"]
            pct_hc  = round(((a2 / a1) - 1) * 100, 1) if (a1 and a2) else None
            pct_opx = round(((o["plan"] / o["est"]) - 1) * 100, 1) if (o["est"] and o["plan"]) else None
            _data_row(u["desc"], [
                round(a1) if a1 is not None else None,
                round(a2) if a2 is not None else None,
                pct_hc,
                round(o["est"], 2), round(o["plan"], 2), pct_opx,
            ])
        ta1 = _avg_tot(i1); ta2 = _avg_tot(i2)
        pct_t = round(((ta2 / ta1) - 1) * 100, 1) if (ta1 and ta2) else None
        pct_o = round(((tot_plan / tot_est) - 1) * 100, 1) if (tot_est and tot_plan) else None
        _data_row("Total", [
            round(ta1) if ta1 else None, round(ta2) if ta2 else None, pct_t,
            round(tot_est, 2), round(tot_plan, 2), pct_o,
        ], is_grand=True)
    cur += 1

    # ── 2. Closing H/C ──────────────────────────────────────────────────────
    _section_hdr("Closing H/C")
    _col_hdr([_fym(y) for y in yrs])
    for u in units:
        _data_row(u["desc"], [u["hc"].get(y) for y in yrs])
    _data_row("Total", [totals.get(y) for y in yrs], is_grand=True)
    cur += 1

    # ── 3. Average H/C ──────────────────────────────────────────────────────
    _section_hdr("Average H/C")
    _col_hdr([_fym(y) for y in yrs])
    for u in units:
        _data_row(u["desc"], [_avg_hc(u["hc"], i) for i in range(len(yrs))])
    _data_row("Total", [_avg_tot(i) for i in range(len(yrs))], is_grand=True)
    cur += 1

    # ── 4. Increase in Closing H/C (%) ─────────────────────────────────────
    if len(yrs) >= 2:
        _section_hdr("Increase in Closing H/C (%)")
        pairs = [(yrs[i - 1], yrs[i]) for i in range(1, len(yrs))]
        _col_hdr([f"{_fym(yf)} → {_fym(yt)}" for yf, yt in pairs])
        for u in units:
            vals = []
            for yf, yt in pairs:
                a = u["hc"].get(yf); b = u["hc"].get(yt)
                vals.append(round(((b / a) - 1) * 100, 1) if (a and b) else None)
            _data_row(u["desc"], vals, fmt=NUM_FMT_PCT)
        tot_vals = []
        for yf, yt in pairs:
            a = totals.get(yf); b = totals.get(yt)
            tot_vals.append(round(((b / a) - 1) * 100, 1) if (a and b) else None)
        _data_row("Total", tot_vals, fmt=NUM_FMT_PCT, is_grand=True)
        cur += 1

    # ── 5. Increase in Average H/C (%) ─────────────────────────────────────
    if len(yrs) >= 2:
        _section_hdr("Increase in Average H/C (%)")
        pair_idxs = [(i - 1, i) for i in range(1, len(yrs))]
        _col_hdr([f"{yrs[i1]} → {yrs[i2]}" for i1, i2 in pair_idxs])
        for u in units:
            vals = []
            for i1, i2 in pair_idxs:
                prev = _avg_hc(u["hc"], i1); curr = _avg_hc(u["hc"], i2)
                vals.append(round(((curr / prev) - 1) * 100, 1) if (prev and curr) else None)
            _data_row(u["desc"], vals, fmt=NUM_FMT_PCT)
        tot_vals = []
        for i1, i2 in pair_idxs:
            prev = _avg_tot(i1); curr = _avg_tot(i2)
            tot_vals.append(round(((curr / prev) - 1) * 100, 1) if (prev and curr) else None)
        _data_row("Total", tot_vals, fmt=NUM_FMT_PCT, is_grand=True)

    _auto_widths(ws)
    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — Annual Budget Consolidated
#
# JS display: Expense Head rows (blue), Sub-head rows (orange), Item rows,
#   Quarter columns (Q1-Q4 monthly + QTR totals + YEAR total)
#   Values in raw rupees (formatINR — whole number Indian grouping)
# ═══════════════════════════════════════════════════════════════════════════════

# Column layout (1-based)
_C_SI    = 2    # Sl #
_C_HEAD  = 3    # HEAD OF EXPENSE
_C_TYPE  = 4    # TYPE OF EXPENSE
_C_START = 5    # Apr (first data col)
_C_END   = 21   # YEAR total
# 5-7=Q1 months, 8-10=Q2, 11-13=Q3, 14-16=Q4, 17=QTR1, 18=QTR2, 19=QTR3, 20=QTR4, 21=YEAR


def _to_roman(n):
    vals = [1000,900,500,400,100,90,50,40,10,9,5,4,1]
    syms = ["M","CM","D","CD","C","XC","L","XL","X","IX","V","IV","I"]
    out = ""; i = 0
    while n > 0:
        for _ in range(n // vals[i]):
            out += syms[i]; n -= vals[i]
        i += 1
    return out


def _qtr_fmls(r):
    return [f"=SUM(E{r}:G{r})", f"=SUM(H{r}:J{r})", f"=SUM(K{r}:M{r})",
            f"=SUM(N{r}:P{r})", f"=SUM(Q{r}:T{r})"]


def _sum_fml(col_letter, rows):
    if not rows: return 0
    return "=" + "+".join(f"{col_letter}{r}" for r in rows)


def _totals_fmls(rows):
    return [_sum_fml(c, rows) for c in list("EFGHIJKLMNOPQRSTU")]


def _style_annual_row(ws, row, fill=None, font=None, is_hdr=False):
    for col in range(_C_SI, _C_END + 1):
        c = ws.cell(row, col)
        c.border = _BORDER
        if is_hdr:
            c.alignment = _CENTER
        else:
            if col == _C_SI:                   c.alignment = _CENTER
            elif col in (_C_HEAD, _C_TYPE):    c.alignment = _LEFT
            else:                              c.alignment = _RIGHT
        if fill: c.fill = fill
        if font: c.font = font


def _fmt_annual_row(ws, row):
    for col in range(_C_START, _C_END + 1):
        ws.cell(row, col).number_format = NUM_FMT_RAW


def _write_annual_headers(ws, fy, r1_label="Annual Budget Consolidated"):
    ws.append([
        "", "Sl #", "HEAD OF EXPENSE", "TYPE OF EXPENSE",
        "QUARTER I", "", "", "QUARTER II", "", "",
        "QUARTER III", "", "", "QUARTER IV", "", "",
        "QTR-1", "QTR-2", "QTR-3", "QTR-4", f"YEAR {fy}",
    ])
    r1 = ws.max_row
    ws.append([
        "", "Sl #", "HEAD OF EXPENSE", "TYPE OF EXPENSE",
        "Apr", "May", "Jun", "Jul", "Aug", "Sep",
        "Oct", "Nov", "Dec", "Jan", "Feb", "Mar",
        "QTR-1", "QTR-2", "QTR-3", "QTR-4", f"YEAR {fy}",
    ])
    r2 = ws.max_row
    _style_annual_row(ws, r1, _FILL_BLUE_MID, _WHITE_BOLD, is_hdr=True)
    _style_annual_row(ws, r2, _FILL_BLUE_MID, _WHITE_BOLD, is_hdr=True)
    for c1, c2 in [(5,7),(8,10),(11,13),(14,16)]:
        _merge(ws, r1, c1, r1, c2)
    for col in [2, 3, 4]:
        _merge(ws, r1, col, r2, col)
    for col in range(17, 22):
        _merge(ws, r1, col, r2, col)
    ws.freeze_panes = "E7"


def _sheet_annual(wb, data, fy, org="Azim Premji Foundation"):
    ws = wb.create_sheet("Annual Budget Consolidated")
    _sheet_title(ws, org, f"Annual Budget Consolidated – {fy}", 21)
    _write_annual_headers(ws, fy)

    head_total_rows = []
    head_counter    = 0

    for head in (data or []):
        head_counter += 1
        alpha = chr(64 + head_counter)
        head_name = _norm(head.get("name") or "")

        if "COVID" in head_name:
            ws.append([])
            item = (head.get("items") or [{}])[0]
            r = ws.max_row + 1
            ws.append(["", alpha, head.get("name", ""), item.get("name", ""),
                        *item.get("q1", [0, 0, 0]), *item.get("q2", [0, 0, 0]),
                        *item.get("q3", [0, 0, 0]), *item.get("q4", [0, 0, 0]),
                        *_qtr_fmls(r)])
            _style_annual_row(ws, ws.max_row, font=_NORMAL)
            _fmt_annual_row(ws, ws.max_row)
            continue

        # Section header row
        ws.append(["", alpha, head.get("name", "")])
        r_sec = ws.max_row
        for col in range(_C_HEAD, _C_END + 1):
            c = ws.cell(r_sec, col)
            c.fill = _FILL_BLUE_LIGHT
            c.font = Font(bold=True, name="Calibri", size=10, color="003B63")
            c.border = _BORDER; c.alignment = _LEFT
        ws.cell(r_sec, _C_SI).fill = _FILL_BLUE_LIGHT
        ws.cell(r_sec, _C_SI).border = _BORDER
        _merge(ws, r_sec, _C_HEAD, r_sec, _C_END)

        if "OPERATING" in head_name:
            ws.append([])

        sub_total_rows = []; direct_item_rows = []

        # Direct items (e.g. CAPEX items)
        for item in (head.get("items") or []):
            r = ws.max_row + 1
            ws.append(["", "", (item.get("sub_head_of_expense") or "").strip(), item.get("name", ""),
                        *item.get("q1", [0, 0, 0]), *item.get("q2", [0, 0, 0]),
                        *item.get("q3", [0, 0, 0]), *item.get("q4", [0, 0, 0]),
                        *_qtr_fmls(r)])
            _style_annual_row(ws, ws.max_row, font=_NORMAL)
            _fmt_annual_row(ws, ws.max_row)
            direct_item_rows.append(ws.max_row)

        # Sub-heads
        sub_ctr = 1
        for sub in (head.get("sub_heads") or []):
            roman = _to_roman(sub_ctr)
            ws.append(["", roman, sub.get("name", "")])
            r_sub = ws.max_row
            for col in range(_C_HEAD, _C_END + 1):
                c = ws.cell(r_sub, col)
                c.fill = _FILL_ORANGE_LIGHT
                c.font = Font(bold=True, name="Calibri", size=10, color="7A3B00")
                c.border = _BORDER; c.alignment = _LEFT
            ws.cell(r_sub, _C_SI).fill = _FILL_ORANGE_LIGHT
            ws.cell(r_sub, _C_SI).border = _BORDER
            _merge(ws, r_sub, _C_HEAD, r_sub, _C_END)

            sub_item_rows = []
            for item in (sub.get("items") or []):
                r = ws.max_row + 1
                ws.append(["", "", sub.get("name", ""), item.get("name", ""),
                            *item.get("q1", [0, 0, 0]), *item.get("q2", [0, 0, 0]),
                            *item.get("q3", [0, 0, 0]), *item.get("q4", [0, 0, 0]),
                            *_qtr_fmls(r)])
                _style_annual_row(ws, ws.max_row, font=_NORMAL)
                _fmt_annual_row(ws, ws.max_row)
                sub_item_rows.append(ws.max_row)

            if sub_item_rows:
                ws.append(["", "", "", f"TOTAL - {sub.get('name', '')}",
                            *_totals_fmls(sub_item_rows)])
                _style_annual_row(ws, ws.max_row,
                                   PatternFill("solid", fgColor="EBF5FB"), _BOLD)
                _fmt_annual_row(ws, ws.max_row)
                sub_total_rows.append(ws.max_row)
            sub_ctr += 1

        total_rows = sub_total_rows if sub_total_rows else direct_item_rows
        if total_rows:
            ws.append(["", "", "", f"TOTAL - {head.get('name', '')}",
                        *_totals_fmls(total_rows)])
            _style_annual_row(ws, ws.max_row,
                               PatternFill("solid", fgColor="D4E6F1"),
                               Font(bold=True, name="Calibri", size=10, color="003B63"))
            _fmt_annual_row(ws, ws.max_row)
            head_total_rows.append(ws.max_row)
            if "OPERATING" in head_name:
                ws.append([])

    if head_total_rows:
        ws.append(["", "", "", "GRAND TOTAL", *_totals_fmls(head_total_rows)])
        _style_annual_row(ws, ws.max_row, _FILL_BLUE_DARK,
                           Font(bold=True, color="FFFFFF", name="Calibri", size=10))
        _fmt_annual_row(ws, ws.max_row)

    _auto_widths(ws)
    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 5 — Estimate Consolidated
#
# Same format as Annual but data comes from get_grouped_actuals… API.
# API shape: items have Q1/Q2/Q3/Q4 keys (float totals) and months dict.
# ═══════════════════════════════════════════════════════════════════════════════

def _sheet_estimate(wb, data, fy, org="Azim Premji Foundation"):
    # JS uses Q1/Q2/Q3/Q4 keys directly (pre-computed by API).
    # months dict is only used when quarters are expanded (individual month cols).
    # We MUST use Q1/Q2/Q3/Q4 as the source of truth to match JS totals exactly.
    MONTH_KEYS = ["4", "5", "6", "7", "8", "9", "10", "11", "12", "1", "2", "3"]
    Q_MONTH_IDX = {"q1": [0,1,2], "q2": [3,4,5], "q3": [6,7,8], "q4": [9,10,11]}

    def _get_quarters(obj):
        """
        Return q1/q2/q3/q4 as 3-element monthly lists.
        Source: months dict for individual values.
        But quarterly TOTAL must equal Q1/Q2/Q3/Q4 keys (JS qTot).
        If months sum != Q key, distribute Q key evenly across 3 months.
        """
        m    = obj.get("months") or {}
        mths = [_fv(m.get(k, 0)) for k in MONTH_KEYS]
        result = {}
        for qk, idxs in Q_MONTH_IDX.items():
            q_key  = qk.upper()   # Q1/Q2/Q3/Q4
            q_val  = _fv(obj.get(q_key, 0))   # authoritative quarter total from API
            m_vals = [mths[i] for i in idxs]
            m_sum  = sum(m_vals)
            if abs(m_sum - q_val) < 0.01:
                # months already add up correctly
                result[qk] = m_vals
            else:
                # months don't match Q key — use Q key split evenly (preserves total)
                # This ensures Excel SUM(monthly) = Q key = matches JS display
                each = q_val / 3.0
                result[qk] = [each, each, q_val - 2 * each]
        return result

    def _norm_obj(obj):
        ql = _get_quarters(obj)
        obj.update(ql)
        for item in (obj.get("items") or []):
            item.update(_get_quarters(item))
        for sub in (obj.get("sub_heads") or []):
            sub.update(_get_quarters(sub))
            for item in (sub.get("items") or []):
                item.update(_get_quarters(item))
        return obj

    normalised = [_norm_obj(dict(h)) for h in (data or [])]

    ws = wb.create_sheet("Estimate Consolidated")
    _sheet_title(ws, org, f"Estimate Consolidated – {fy}", 21)
    _write_annual_headers(ws, fy)

    head_total_rows = []
    head_counter    = 0

    for head in normalised:
        head_counter += 1
        alpha     = chr(64 + head_counter)
        head_name = _norm(head.get("name") or "")

        if "COVID" in head_name:
            ws.append([])
            item = (head.get("items") or [{}])[0]
            r = ws.max_row + 1
            ws.append(["", alpha, head.get("name", ""), item.get("name", ""),
                        *item.get("q1", [0, 0, 0]), *item.get("q2", [0, 0, 0]),
                        *item.get("q3", [0, 0, 0]), *item.get("q4", [0, 0, 0]),
                        *_qtr_fmls(r)])
            _style_annual_row(ws, ws.max_row, font=_NORMAL)
            _fmt_annual_row(ws, ws.max_row)
            # COVID row IS included in GRAND TOTAL (JS sums all heads incl. COVID)
            head_total_rows.append(ws.max_row)
            continue

        ws.append(["", alpha, head.get("name", "")])
        r_sec = ws.max_row
        for col in range(_C_HEAD, _C_END + 1):
            c = ws.cell(r_sec, col)
            c.fill = _FILL_BLUE_LIGHT
            c.font = Font(bold=True, name="Calibri", size=10, color="003B63")
            c.border = _BORDER; c.alignment = _LEFT
        ws.cell(r_sec, _C_SI).fill = _FILL_BLUE_LIGHT
        ws.cell(r_sec, _C_SI).border = _BORDER
        _merge(ws, r_sec, _C_HEAD, r_sec, _C_END)

        if "OPERATING" in head_name:
            ws.append([])

        sub_total_rows = []; direct_item_rows = []

        for item in (head.get("items") or []):
            r = ws.max_row + 1
            ws.append(["", "", (item.get("sub_head_of_expense") or "").strip(), item.get("name", ""),
                        *item.get("q1", [0, 0, 0]), *item.get("q2", [0, 0, 0]),
                        *item.get("q3", [0, 0, 0]), *item.get("q4", [0, 0, 0]),
                        *_qtr_fmls(r)])
            _style_annual_row(ws, ws.max_row, font=_NORMAL)
            _fmt_annual_row(ws, ws.max_row)
            direct_item_rows.append(ws.max_row)

        sub_ctr = 1
        for sub in (head.get("sub_heads") or []):
            roman = _to_roman(sub_ctr)
            ws.append(["", roman, sub.get("name", "")])
            r_sub = ws.max_row
            for col in range(_C_HEAD, _C_END + 1):
                c = ws.cell(r_sub, col)
                c.fill = _FILL_ORANGE_LIGHT
                c.font = Font(bold=True, name="Calibri", size=10, color="7A3B00")
                c.border = _BORDER; c.alignment = _LEFT
            ws.cell(r_sub, _C_SI).fill = _FILL_ORANGE_LIGHT
            ws.cell(r_sub, _C_SI).border = _BORDER
            _merge(ws, r_sub, _C_HEAD, r_sub, _C_END)

            sub_item_rows = []
            for item in (sub.get("items") or []):
                r = ws.max_row + 1
                ws.append(["", "", sub.get("name", ""), item.get("name", ""),
                            *item.get("q1", [0, 0, 0]), *item.get("q2", [0, 0, 0]),
                            *item.get("q3", [0, 0, 0]), *item.get("q4", [0, 0, 0]),
                            *_qtr_fmls(r)])
                _style_annual_row(ws, ws.max_row, font=_NORMAL)
                _fmt_annual_row(ws, ws.max_row)
                sub_item_rows.append(ws.max_row)

            if sub_item_rows:
                ws.append(["", "", "", f"TOTAL - {sub.get('name', '')}",
                            *_totals_fmls(sub_item_rows)])
                _style_annual_row(ws, ws.max_row,
                                   PatternFill("solid", fgColor="EBF5FB"), _BOLD)
                _fmt_annual_row(ws, ws.max_row)
                sub_total_rows.append(ws.max_row)
            sub_ctr += 1

        total_rows = sub_total_rows if sub_total_rows else direct_item_rows
        if total_rows:
            ws.append(["", "", "", f"TOTAL - {head.get('name', '')}",
                        *_totals_fmls(total_rows)])
            _style_annual_row(ws, ws.max_row,
                               PatternFill("solid", fgColor="D4E6F1"),
                               Font(bold=True, name="Calibri", size=10, color="003B63"))
            _fmt_annual_row(ws, ws.max_row)
            head_total_rows.append(ws.max_row)
            if "OPERATING" in head_name:
                ws.append([])

    if head_total_rows:
        ws.append(["", "", "", "GRAND TOTAL", *_totals_fmls(head_total_rows)])
        _style_annual_row(ws, ws.max_row, _FILL_BLUE_DARK,
                           Font(bold=True, color="FFFFFF", name="Calibri", size=10))
        _fmt_annual_row(ws, ws.max_row)

    _auto_widths(ws)
    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 6 — Budget & Estimate
#
# JS display: one table, rows = Expense Head / Sub-head / Line Item
#   Columns: Expense Head/Line Item | [per unit: Plan | Est] | Grand Total Plan | Grand Total Est
#   Values in raw rupees (formatINR whole number)
#   Sections from rawData[0].actuals (structure only, values summed across all entities)
#   Total column = GRAND TOTAL section (seq 9999) from each entity
#
# Export format:
#   Row 1 = unit column headers (plan + est per unit + Grand Total plan + Grand Total est)
#   Section rows (blue), sub-head rows (orange), item rows (white)
#   Grand Total row (dark blue)
# ═══════════════════════════════════════════════════════════════════════════════

def _sheet_be(wb, be_data, fy, plan_label, est_label, org="Azim Premji Foundation"):
    ws = wb.create_sheet("Budget & Estimate")
    entities = [e for e in (be_data or []) if not _is_consolidated(e)]

    if not entities:
        ws.cell(1, 1, "No data")
        return ws

    # Build structure from first entity's actuals
    def _build_struct(entity):
        struct = []
        for sec in (entity.get("actuals") or []):
            if _is_gt(sec): continue
            struct.append({
                "name": sec.get("name", ""),
                "sub_heads": [
                    {"name": sub.get("name", ""),
                     "items": [{"name": i.get("name", "")} for i in (sub.get("items") or [])]}
                    for sub in (sec.get("sub_heads") or [])
                ],
                "items": [{"name": i.get("name", "")} for i in (sec.get("items") or [])],
            })
        return struct

    struct = _build_struct(entities[0])

    # Value extraction functions — match JS secVal/subVal/itemVal/grandVal
    def _sec_val(entity, sn, field):
        v = 0.0
        for s in (entity.get("actuals") or []):
            if _is_gt(s) or s.get("name") != sn: continue
            v += _fv(s.get("ytd" if field == "plan" else "total_posted_amt_ytd"))
        return v

    def _sub_val(entity, sn, subn, field):
        v = 0.0
        for s in (entity.get("actuals") or []):
            if _is_gt(s) or s.get("name") != sn: continue
            for sub in (s.get("sub_heads") or []):
                if sub.get("name") != subn: continue
                v += _fv(sub.get("ytd" if field == "plan" else "total_posted_amt_ytd"))
        return v

    def _item_val(entity, nm, field):
        v = 0.0
        fld = "ytd" if field == "plan" else "total_posted_amt"
        for s in (entity.get("actuals") or []):
            if _is_gt(s): continue
            for i in (s.get("items") or []):
                if i.get("name") == nm: v += _fv(i.get(fld))
            for sub in (s.get("sub_heads") or []):
                for i in (sub.get("items") or []):
                    if i.get("name") == nm: v += _fv(i.get(fld))
        return v

    def _grand_val(entity, field):
        # Use GRAND TOTAL section directly (match JS grandVal)
        gt = 0.0; found = False
        fld = "ytd" if field == "plan" else "total_posted_amt_ytd"
        for s in (entity.get("actuals") or []):
            if _is_gt(s):
                gt += _fv(s.get(fld)); found = True
        if not found:
            for s in (entity.get("actuals") or []):
                gt += _fv(s.get(fld))
        return gt

    # Column layout:
    # Col 2 = Expense Head / Line Item
    # Col 3+ = per entity: Plan, Est
    # Last 2 = Grand Total Plan, Grand Total Est
    n_ents   = len(entities)
    c_items  = 2
    c_data   = 3                           # first entity plan col
    c_gt_p   = c_data + n_ents * 2         # Grand Total Plan
    c_gt_e   = c_gt_p + 1                  # Grand Total Est
    last_col = c_gt_e

    _sheet_title(ws, org, f"Budget & Estimate – {fy}", last_col)

    cur = 4

    # ── Header row 1 — entity names ─────────────────────────────────────────
    for col in range(c_items, last_col + 1):
        c = ws.cell(cur, col)
        c.fill = _FILL_BLUE_MID; c.font = _WHITE_BOLD
        c.border = _BORDER; c.alignment = _CENTER

    ws.cell(cur, c_items, "Expense Head / Line Item")
    _merge(ws, cur, c_items, cur + 1, c_items)

    for i, e in enumerate(entities):
        col = c_data + i * 2
        ws.cell(cur, col, (e.get("label") or "").strip())
        _merge(ws, cur, col, cur, col + 1)

    ws.cell(cur, c_gt_p, "Grand Total")
    _merge(ws, cur, c_gt_p, cur, c_gt_e)
    cur += 1

    # ── Header row 2 — Plan / Est per entity ────────────────────────────────
    for col in range(c_items, last_col + 1):
        c = ws.cell(cur, col)
        c.fill = _FILL_ORANGE; c.font = _WHITE_BOLD
        c.border = _BORDER; c.alignment = _CENTER

    for i in range(n_ents):
        col = c_data + i * 2
        ws.cell(cur, col,     plan_label)
        ws.cell(cur, col + 1, est_label)

    ws.cell(cur, c_gt_p, plan_label)
    ws.cell(cur, c_gt_e, est_label)
    cur += 1

    # ── Currency note ────────────────────────────────────────────────────────
    ws.cell(cur, last_col, "Raw ₹")
    ws.cell(cur, last_col).font = _ITALIC
    ws.cell(cur, last_col).alignment = Alignment(horizontal="right")
    cur += 1

    def _fill_data_cols(row, plan_vals, est_vals, gt_p, gt_e,
                        sec_fill=None, sec_font=None):
        for col in range(c_items, last_col + 1):
            c = ws.cell(row, col)
            c.border = _BORDER
            c.alignment = _LEFT if col == c_items else _RIGHT
            if sec_fill: c.fill = sec_fill
            if sec_font: c.font = sec_font
            c.number_format = NUM_FMT_RAW
        for i, (pv, ev) in enumerate(zip(plan_vals, est_vals)):
            col = c_data + i * 2
            ws.cell(row, col,     pv or None).number_format = NUM_FMT_RAW
            ws.cell(row, col + 1, ev or None).number_format = NUM_FMT_RAW
        ws.cell(row, c_gt_p, gt_p or None).number_format = NUM_FMT_RAW
        ws.cell(row, c_gt_e, gt_e or None).number_format = NUM_FMT_RAW

    # ── Data rows ────────────────────────────────────────────────────────────
    for sec in struct:
        sn = sec["name"]

        # Section row — cb-row-head (blue light)
        sec_p_vals = [_sec_val(e, sn, "plan") for e in entities]
        sec_e_vals = [_sec_val(e, sn, "est")  for e in entities]
        gt_sec_p   = sum(sec_p_vals); gt_sec_e = sum(sec_e_vals)

        ws.cell(cur, c_items, sn)
        _fill_data_cols(cur, sec_p_vals, sec_e_vals, gt_sec_p, gt_sec_e,
                        _FILL_BLUE_LIGHT,
                        Font(bold=True, name="Calibri", size=10, color="003B63"))
        cur += 1

        # Sub-head rows — cb-row-sub (orange light)
        for sub in sec.get("sub_heads", []):
            subn = sub["name"]
            sub_p = [_sub_val(e, sn, subn, "plan") for e in entities]
            sub_e = [_sub_val(e, sn, subn, "est")  for e in entities]
            gt_sp = sum(sub_p); gt_se = sum(sub_e)

            ws.cell(cur, c_items, f"  {subn}")
            _fill_data_cols(cur, sub_p, sub_e, gt_sp, gt_se,
                            _FILL_ORANGE_LIGHT,
                            Font(bold=True, name="Calibri", size=10, color="7A3B00"))
            cur += 1

            # Item rows
            for item in sub.get("items", []):
                nm = item["name"]
                ip = [_item_val(e, nm, "plan") for e in entities]
                ie = [_item_val(e, nm, "est")  for e in entities]
                gt_ip = sum(ip); gt_ie = sum(ie)
                ws.cell(cur, c_items, f"    {nm}")
                _fill_data_cols(cur, ip, ie, gt_ip, gt_ie, _FILL_WHITE, _NORMAL)
                cur += 1

        # Direct items on section
        for item in sec.get("items", []):
            nm = item["name"]
            ip = [_item_val(e, nm, "plan") for e in entities]
            ie = [_item_val(e, nm, "est")  for e in entities]
            gt_ip = sum(ip); gt_ie = sum(ie)
            ws.cell(cur, c_items, f"  {nm}")
            _fill_data_cols(cur, ip, ie, gt_ip, gt_ie, _FILL_WHITE, _NORMAL)
            cur += 1

    # Grand Total row — use GRAND TOTAL section (seq 9999) per entity
    gt_p_vals = [_grand_val(e, "plan") for e in entities]
    gt_e_vals = [_grand_val(e, "est")  for e in entities]
    all_gp    = sum(gt_p_vals); all_ge = sum(gt_e_vals)

    ws.cell(cur, c_items, "GRAND TOTAL")
    for col in range(c_items, last_col + 1):
        c = ws.cell(cur, col)
        c.fill = _FILL_BLUE_DARK
        c.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        c.border = _BORDER; c.alignment = _LEFT if col == c_items else _RIGHT
        c.number_format = NUM_FMT_RAW
    for i, (pv, ev) in enumerate(zip(gt_p_vals, gt_e_vals)):
        col = c_data + i * 2
        ws.cell(cur, col,     pv or None).number_format = NUM_FMT_RAW
        ws.cell(cur, col + 1, ev or None).number_format = NUM_FMT_RAW
    ws.cell(cur, c_gt_p, all_gp or None).number_format = NUM_FMT_RAW
    ws.cell(cur, c_gt_e, all_ge or None).number_format = NUM_FMT_RAW

    _auto_widths(ws)
    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# WHITELISTED API ENDPOINTS
# ═══════════════════════════════════════════════════════════════════════════════

@frappe.whitelist()
def export_ppt(financial_year, ppt_rows, prev_ppt_rows,
               budget_label, est_label, prev_budget_label, prev_est_label):
    fy = financial_year or "2025-26"
    wb = Workbook(); wb.remove(wb.active)
    _sheet_ppt(wb,
               json.loads(ppt_rows),      budget_label,      est_label,
               json.loads(prev_ppt_rows), prev_budget_label, prev_est_label,
               fy, _prev_fy(fy))
    return {"filename": f"Foundation_Metrics_{fy}.xlsx", "data": _wb_to_b64(wb)}


@frappe.whitelist()
def export_summary_inr(financial_year, summary_data):
    fy = financial_year or "2025-26"
    wb = Workbook(); wb.remove(wb.active)
    _sheet_summary_inr(wb, json.loads(summary_data), fy)
    return {"filename": f"Summary_INR_{fy}.xlsx", "data": _wb_to_b64(wb)}


@frappe.whitelist()
def export_headcount(financial_year, headcount_data):
    fy  = financial_year or "2025-26"
    raw = json.loads(headcount_data)
    wb  = Workbook(); wb.remove(wb.active)
    _sheet_headcount(wb, raw, fy)
    return {"filename": f"Headcount_{fy}.xlsx", "data": _wb_to_b64(wb)}


@frappe.whitelist()
def export_annual(financial_year, annual_data):
    fy = financial_year or "2025-26"
    wb = Workbook(); wb.remove(wb.active)
    _sheet_annual(wb, json.loads(annual_data), fy)
    return {"filename": f"Annual_Budget_Consolidated_{fy}.xlsx", "data": _wb_to_b64(wb)}


@frappe.whitelist()
def export_estimate(financial_year, estimate_data):
    fy = financial_year or "2025-26"
    wb = Workbook(); wb.remove(wb.active)
    _sheet_estimate(wb, json.loads(estimate_data), fy)
    return {"filename": f"Estimate_Consolidated_{fy}.xlsx", "data": _wb_to_b64(wb)}


@frappe.whitelist()
def export_budget_estimate(financial_year, be_data):
    fy     = financial_year or "2025-26"
    labels = _fy_labels(fy)
    wb     = Workbook(); wb.remove(wb.active)
    _sheet_be(wb, json.loads(be_data), fy, labels["plan"], labels["est"])
    return {"filename": f"Budget_and_Estimate_{fy}.xlsx", "data": _wb_to_b64(wb)}


@frappe.whitelist()
def export_all(financial_year,
               ppt_rows, prev_ppt_rows,
               budget_label, est_label, prev_budget_label, prev_est_label,
               summary_data, headcount_data,
               annual_data, estimate_data, be_data):
    fy     = financial_year or "2025-26"
    labels = _fy_labels(fy)
    wb     = Workbook(); wb.remove(wb.active)

    _sheet_ppt(wb,
               json.loads(ppt_rows),      budget_label,      est_label,
               json.loads(prev_ppt_rows), prev_budget_label, prev_est_label,
               fy, _prev_fy(fy))
    _sheet_summary_inr(wb, json.loads(summary_data),   fy)
    _sheet_headcount(  wb, json.loads(headcount_data),  fy)
    _sheet_annual(     wb, json.loads(annual_data),     fy)
    _sheet_estimate(   wb, json.loads(estimate_data),   fy)
    _sheet_be(         wb, json.loads(be_data),         fy, labels["plan"], labels["est"])

    return {
        "filename": f"Foundation_Consolidated_Budget_{fy}.xlsx",
        "data"    : _wb_to_b64(wb),
    }


# * ==============================================================  Excel export for the ERP Actuals page  =====================================================================================
# Maps the page's API picker labels to the actual functions — kept as an
# explicit allowlist so the export endpoint can't be pointed at arbitrary
# server functions via method_key.
ERP_ACTUALS_EXPORT_METHODS = {
    "get_actuals_from_erp_prod": get_actuals_from_erp_prod,
    "get_actuals_from_erp_month_wise": get_actuals_from_erp_month_wise,
}


@frappe.whitelist()
def export_erp_actuals_excel(fiscal_year, accounting_period, rows=None, method_key=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    # `rows` carries exactly what's on screen (post column-filter) as JSON, sent
    # from erp_actuals.js so the export matches whatever the user filtered to.
    # method_key is kept only as a fallback for old cached JS / direct API callers.
    if rows is not None:
        rows = frappe.parse_json(rows) if isinstance(rows, str) else rows
    else:
        fn = ERP_ACTUALS_EXPORT_METHODS.get(method_key)
        if not fn:
            frappe.throw("Invalid export method")

        result = fn(fiscal_year=fiscal_year, accounting_period=accounting_period)

        if result.get("status") != "success":
            frappe.throw(result.get("error") or "ERP request failed")

        rows = result.get("data") or []

    # fiscal_year is dropped — same as the on-screen table, it just repeats
    # the fiscal_year/accounting_period already shown for the whole export.
    columns = []
    for row in rows:
        for key in (row or {}).keys():
            if key != "fiscal_year" and key not in columns:
                columns.append(key)

    wb = Workbook()
    ws = wb.active
    ws.title = "ERP Actuals"

    header_fill = PatternFill("solid", fgColor="0076B6")
    header_font = Font(bold=True, color="FFFFFF")

    ws.append(columns)
    for col_idx in range(1, len(columns) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font

    # Amount columns get real Excel numbers (right-aligned, thousands separator);
    # code-like columns (account, deptid, accounting_period, ...) stay as text so
    # leading zeros / exact codes aren't mangled by Excel's numeric coercion.
    AMOUNT_COLUMNS = {"posted_total_amt"}
    AMOUNT_FORMAT = "#,##0.00;[Red]-#,##0.00"

    for row in rows:
        values = []
        for col in columns:
            value = row.get(col)
            if col == "is_adjustment":
                value = "True" if str(value).strip().lower() in ("1", "true", "yes") else "False"
            elif col in AMOUNT_COLUMNS:
                value = frappe.utils.flt(value)
            values.append(value)
        ws.append(values)
        if AMOUNT_COLUMNS:
            for col_idx, col in enumerate(columns, start=1):
                if col in AMOUNT_COLUMNS:
                    ws.cell(row=ws.max_row, column=col_idx).number_format = AMOUNT_FORMAT

    if "posted_total_amt" in columns:
        total = sum(frappe.utils.flt(row.get("posted_total_amt")) for row in rows)
        total_row = ["" for _ in columns]
        total_row[0] = "Total"
        total_row[columns.index("posted_total_amt")] = total
        ws.append(total_row)

        total_fill = PatternFill("solid", fgColor="EAF3FA")
        total_font = Font(bold=True, color="0076B6")
        for col_idx, col in enumerate(columns, start=1):
            cell = ws.cell(row=ws.max_row, column=col_idx)
            cell.fill = total_fill
            cell.font = total_font
            if col in AMOUNT_COLUMNS:
                cell.number_format = AMOUNT_FORMAT

    for col_idx, col in enumerate(columns, start=1):
        max_len = max([len(col)] + [len(str(r.get(col, "") or "")) for r in rows]) if rows else len(col)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

    stream = io.BytesIO()
    wb.save(stream)

    frappe.response["filename"] = f"ERP_Actuals_{fiscal_year}_P{accounting_period}.xlsx"
    frappe.response["filecontent"] = stream.getvalue()
    frappe.response["type"] = "binary"