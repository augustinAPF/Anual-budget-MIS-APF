# =============================================================================
# File: apps/annual_budget/annual_budget/api/monthly_mis.py
# Endpoint: /api/method/annual_budget.api.monthly_mis.export_monthly_mis
# =============================================================================

import frappe
import io
from frappe.utils import flt


@frappe.whitelist()
def export_monthly_mis(financial_year, month, export_format="excel"):
    """Export full Monthly MIS as Excel or PDF."""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    except ImportError:
        frappe.throw("openpyxl required. Run: bench pip install openpyxl")

    from annual_budget.api.foundation_consolidated_report import (
        get_unit_wise_plan,
        get_monthly_mis_break_up,
    )

    # ── Palette ───────────────────────────────────────────────────────────────
    C_BLUE   = "1565C0"
    C_ORANGE = "F26B21"
    C_LTBLUE = "E3F2FD"
    C_WHITE  = "FFFFFF"
    C_GREY   = "F0F4FF"
    C_SUBROW = "F8F9FA"
    C_BLACK  = "000000"

    # ── Style cache (Fix 3: create once, reuse — much faster) ─────────────────
    _fill_cache  = {}
    _border_cache = {}
    _font_cache  = {}

    def _fill(c):
        if c not in _fill_cache:
            _fill_cache[c] = PatternFill("solid", fgColor=c)
        return _fill_cache[c]

    def _font(bold=False, color=C_BLACK, size=9, italic=False, underline=None):
        key = (bold, color, size, italic, underline)
        if key not in _font_cache:
            _font_cache[key] = Font(name="Arial", bold=bold, color=color,
                                     size=size, italic=italic,
                                     underline=underline or "none" if underline else None)
        return _font_cache[key]

    def _border(color=C_BLACK):
        if color not in _border_cache:
            s = Side(style="thin", color=color)
            _border_cache[color] = Border(left=s, right=s, top=s, bottom=s)
        return _border_cache[color]

    def _align(h="left", v="center"):
        return Alignment(horizontal=h, vertical=v, wrap_text=False)

    def prev_fy(fy):
        p = fy.split("-")
        return f"{int(p[0])-1}-{str(int(p[1])-1).zfill(2)}"

    def to_cr(paisa):
        if not paisa: return None
        v = round(flt(paisa) / 10_000_000, 1)
        return v if v else None

    def pct_val(act, bud):
        if not bud: return None
        p = round(flt(act) / flt(bud) * 100, 1)
        return p if p else None

    PREV_FY     = prev_fy(financial_year)
    CUR_LBL     = f"Current Year YTD  {financial_year}"
    PREV_LBL    = f"Last Year YTD  {PREV_FY}"
    GRANTS_NAME = "Grants & Donations"

    # ── Fetch ─────────────────────────────────────────────────────────────────
    cur_raw  = get_unit_wise_plan(financial_year, month,
                                   table_name_filter="Monthly MIS Capex & Opex")
    prev_raw = get_unit_wise_plan(PREV_FY, month,
                                   table_name_filter="Monthly MIS Capex & Opex")
    BREAKUP_LABELS = (
        "Education - District Institutes,Education- Azim Premji Schools,"
        "Azim Premji University (Bangalore Campus),"
        "Azim Premji University (Bhopal Campus),"
        "Azim Premji University (Ranchi Campus),"
        "Enablers,Livelihoods,"
        "Urban Primary care work,Rural Primary care work,"
        "Central Initiatives,Hospital,Health Programs Team & Enablers"
    )
    breakup_raw = get_monthly_mis_break_up(
        financial_year, month,
        table_name_filter=BREAKUP_LABELS,
    ) or {}

    def parse_main(raw):
        out, order, seen = {}, [], set()
        for row in sorted(raw or [], key=lambda x: x.get("sequence_id", 0)):
            lbl = (row.get("label") or "").strip()
            if not lbl or row.get("sequence_id") == 9999: continue
            r = dict(ob=0, oa=0, cb=0, ca=0, vb=0, va=0, tb=0, ta=0,
                     is_sub=int(row.get("is_this_sub_item") or 0), sub_heads=[])
            for sec in sorted(row.get("actuals") or [],
                               key=lambda x: x.get("sequence_id", 0)):
                nm = (sec.get("name") or "").replace("  ", " ").strip().upper()
                b  = flt(sec.get("ytd") or 0)
                a  = flt(sec.get("total_posted_amt_ytd") or 0)
                if nm == "OPERATING EXPENSES":
                    r["ob"] += b; r["oa"] += a
                    for sh in sorted(sec.get("sub_heads") or [],
                                      key=lambda x: x.get("sequence_id", 0)):
                        sn = (sh.get("name") or "").strip()
                        if not sn: continue
                        items = []
                        for it in sorted(sh.get("items") or [],
                                          key=lambda x: x.get("sequence_id", 0)):
                            iname = (it.get("name") or "").strip()
                            if iname:
                                items.append(dict(
                                    name=iname,
                                    b=flt(it.get("ytd") or 0) / 10_000_000,
                                    a=flt(it.get("total_posted_amt") or 0) / 10_000_000,
                                ))
                        r["sub_heads"].append(dict(
                            name=sn,
                            b=flt(sh.get("ytd") or 0) / 10_000_000,
                            a=flt(sh.get("total_posted_amt_ytd") or 0) / 10_000_000,
                            items=items,
                        ))
                elif nm == "CAPITAL EXPENSES":
                    r["cb"] += b; r["ca"] += a
                elif "COVID" in nm:
                    r["vb"] += b; r["va"] += a
                    # Include COVID in opex total (matches extractSection OPEX_NAMES logic)
                    r["ob"] += b; r["oa"] += a
            r["tb"] = r["ob"] + r["cb"]   # ob already includes covid
            r["ta"] = r["oa"] + r["ca"]
            out[lbl] = r
            if lbl not in seen:
                order.append(lbl); seen.add(lbl)
        return out, order

    cur_map, order = parse_main(cur_raw)
    prev_map, _    = parse_main(prev_raw)

    def get_breakup_entries(key):
        """
        Find entries for a given label key from breakup_raw.
        Handles multiple response shapes:
          1. { "table_name": { "label": [entries] } }   ← nested dict of dicts
          2. { "label": [entries] }                       ← flat dict
          3. { "table_name": [entries_with_label_field] } ← list of dicts
        """
        def _filter(entries):
            if not isinstance(entries, list):
                return []
            return [e for e in sorted(entries, key=lambda x: x.get("sequence_id", 0))
                    if (e.get("label") or "") not in ("CONSOLIDATED TOTAL", "")
                    and e.get("settings_doc") != "CONSOLIDATED"]

        if not breakup_raw:
            return []

        # Shape 1: flat dict with key directly
        if key in breakup_raw:
            val = breakup_raw[key]
            if isinstance(val, list):
                return _filter(val)
            if isinstance(val, dict):
                # value is a sub-dict; entries might be inside
                return _filter(list(val.values()))

        # Shape 2: nested — iterate over top-level values
        for grp in breakup_raw.values():
            if isinstance(grp, dict):
                if key in grp:
                    return _filter(grp[key] if isinstance(grp[key], list) else [])
            elif isinstance(grp, list):
                # List of entries — filter by label field
                matched = [e for e in grp
                           if (e.get("label") or "").strip() == key
                           and (e.get("label") or "") not in ("CONSOLIDATED TOTAL",)
                           and e.get("settings_doc") != "CONSOLIDATED"]
                if matched:
                    return sorted(matched, key=lambda x: x.get("sequence_id", 0))

        return []

    def exAct(actuals):
        ob = oa = cb = ca = 0.0
        for sec in sorted(actuals or [], key=lambda x: x.get("sequence_id", 0)):
            nm = (sec.get("name") or "").replace("  ", " ").strip().upper()
            b  = flt(sec.get("ytd") or 0) / 10_000_000
            a  = flt(sec.get("total_posted_amt_ytd") or 0) / 10_000_000
            if not b and not a:
                for sh in (sec.get("sub_heads") or []):
                    b += flt(sh.get("ytd") or 0) / 10_000_000
                    a += flt(sh.get("total_posted_amt_ytd") or 0) / 10_000_000
            if nm == "OPERATING EXPENSES":   ob += b; oa += a
            elif nm == "CAPITAL EXPENSES":   cb += b; ca += a
        return dict(ob=ob, oa=oa, cb=cb, ca=ca, tb=ob+cb, ta=oa+ca)

    # ── Workbook — columns A-G (card 1) + spacer H + I-O (card 2) ────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Monthly MIS"
    ws.sheet_view.showGridLines = False
    # Card 1: A-G (cols 1-7)
    ws.column_dimensions["A"].width = 30
    for col in list("BCDEFG"): ws.column_dimensions[col].width = 11
    # Spacer: H (col 8)
    ws.column_dimensions["H"].width = 2
    # Card 2: I-O (cols 9-15)
    ws.column_dimensions["I"].width = 30
    for col in ["J","K","L","M","N","O"]: ws.column_dimensions[col].width = 11

    # ── Cell helpers — offset-aware ───────────────────────────────────────────
    def _cell(r, c, val, bg, bold=False, fg=C_BLACK, size=9, h="left", italic=False):
        cell = ws.cell(row=r, column=c, value=val)
        cell.fill      = _fill(bg)
        cell.font      = _font(bold=bold, color=fg, size=size, italic=italic)
        cell.alignment = _align(
            h if h != "auto" else
            ("right" if isinstance(val, (int, float)) and val is not None else "left"),
            "center"
        )
        cell.border    = _border(C_WHITE if bg in (C_BLUE, C_ORANGE) else C_BLACK)
        if isinstance(val, float):
            cell.number_format = "#,##0.0"
        return cell

    def hdr(r, c, val, bg, fg=C_WHITE, h="center"):
        return _cell(r, c, val, bg, bold=True, fg=fg, size=9, h=h)

    def dat(r, c, val, bg=C_WHITE, bold=False):
        return _cell(r, c, val, bg, bold=bold, fg=C_WHITE if bg==C_BLUE else C_BLACK,
                     size=9, h="auto")

    # Merge helper
    def merge(r, c1, c2):
        from openpyxl.utils import get_column_letter
        ws.merge_cells(
            f"{get_column_letter(c1)}{r}:{get_column_letter(c2)}{r}"
        )

    # ── Layout builders — offset = 0 (left card) or 8 (right card, skip H) ──
    def section_title(r, text, c_start=1, c_end=7):
        merge(r, c_start, c_end)
        c = ws.cell(row=r, column=c_start, value=text.upper())
        c.fill = _fill(C_WHITE)
        c.font = Font(name="Arial", bold=True, color=C_BLUE, size=11, underline="single")
        c.alignment = _align("left", "center")
        thick = Side(style="medium", color=C_BLUE)
        none  = Side(style=None)
        from openpyxl.styles import Border as B2
        c.border = B2(bottom=thick, left=none, right=none, top=none)
        for col in range(c_start+1, c_end+1):
            ws.cell(row=r, column=col).fill   = _fill(C_WHITE)
            ws.cell(row=r, column=col).border = B2(bottom=thick, left=none, right=none, top=none)
        ws.row_dimensions[r].height = 18
        return r + 1

    def two_row_hdr(r, col_label, grp1, grp2, off=0):
        # off=0 → cols 1-7, off=8 → cols 9-15
        c = off + 1
        merge(r, c, c); merge(r+1, c, c)
        hdr(r, c, col_label, C_BLUE, h="left")
        hdr(r+1, c, "", C_BLUE, h="left")
        merge(r, c+1, c+3); hdr(r, c+1, grp1, C_BLUE, h="center")
        merge(r, c+4, c+6); hdr(r, c+4, grp2, C_BLUE, h="center")
        for i, lbl in enumerate(["Budget","Actuals","% of Budget",
                                   "Budget","Actuals","% of Budget"]):
            hdr(r+1, c+1+i, lbl, C_ORANGE, h="center")
        ws.row_dimensions[r].height   = 14
        ws.row_dimensions[r+1].height = 14
        return r + 2

    def section_hdr_row(r, text, c_start=1, c_end=7):
        merge(r, c_start, c_end)
        c = ws.cell(row=r, column=c_start, value=text)
        c.fill = _fill(C_GREY); c.font = _font(bold=True, color=C_BLUE, size=9)
        c.alignment = _align("left", "center")
        for col in range(c_start+1, c_end+1):
            ws.cell(row=r, column=col).fill   = _fill(C_GREY)
            ws.cell(row=r, column=col).border = _border()
        return r + 1

    def data_row_7(r, lbl, v1, v2, v3, v4, v5, v6, off=0, bg=C_WHITE, bold=False):
        c = off + 1
        dat(r, c,   lbl, bg, bold=bold)
        for i, v in enumerate([v1,v2,v3,v4,v5,v6]):
            dat(r, c+1+i, v, bg, bold=bold)

    def total_row_7(r, cb, ca, pb, pa, off=0, bg=C_LTBLUE):
        c = off + 1
        dat(r, c, "Total", bg, bold=True)
        for i, v in enumerate([cb, ca, pct_val(ca,cb), pb, pa, pct_val(pa,pb)]):
            dat(r, c+1+i, round(v,1) if isinstance(v,float) else v, bg, bold=True)

    def grand_row_7(r, label, cb, ca, pb, pa, off=0):
        c = off + 1
        for col in range(c, c+7):
            ws.cell(row=r, column=col).fill   = _fill(C_BLUE)
            ws.cell(row=r, column=col).font   = _font(bold=True, color=C_WHITE, size=9)
            ws.cell(row=r, column=col).border = _border(C_WHITE)
            ws.cell(row=r, column=col).alignment = _align("right","center")
        ws.cell(row=r, column=c).value     = label
        ws.cell(row=r, column=c).alignment = _align("left","center")
        for i, v in enumerate([cb, ca, pct_val(ca,cb), pb, pa, pct_val(pa,pb)]):
            ws.cell(row=r, column=c+1+i).value = round(v,1) if isinstance(v,float) else v
            if isinstance(ws.cell(row=r, column=c+1+i).value, float):
                ws.cell(row=r, column=c+1+i).number_format = "#,##0.0"

    # ── SECTION WRITERS ───────────────────────────────────────────────────────

    def write_main_section(row, title, bk, ak):
        row = section_title(row, title)
        row = two_row_hdr(row, "Unit", CUR_LBL, PREV_LBL)
        tCB = tCA = tPB = tPA = 0.0
        for lbl in order:
            cm = cur_map.get(lbl, {}); pm = prev_map.get(lbl, {})
            is_sub = cm.get("is_sub",0) or pm.get("is_sub",0)
            bg  = C_SUBROW if is_sub else C_WHITE
            cb  = to_cr(cm.get(bk,0)); ca = to_cr(cm.get(ak,0))
            pb  = to_cr(pm.get(bk,0)); pa = to_cr(pm.get(ak,0))
            data_row_7(row, ("  " if is_sub else "")+lbl,
                       cb, ca, pct_val(ca,cb), pb, pa, pct_val(pa,pb), bg=bg)
            row += 1
            if not is_sub:
                tCB+=flt(cb); tCA+=flt(ca); tPB+=flt(pb); tPA+=flt(pa)
        total_row_7(row, tCB, tCA, tPB, tPA)
        row += 2
        return row

    def write_breakup_section(row, title, keys, col_label="Areas of Work"):
        row = section_title(row, title)
        row = two_row_hdr(row, col_label, "Operating Expense", "Capital Expense")
        g_ob=g_oa=g_cb=g_ca=0.0
        for key in keys:
            entries = get_breakup_entries(key)
            row = section_hdr_row(row, key)
            s_ob=s_oa=s_cb=s_ca=0.0
            for e in entries:
                lbl2 = (e.get("label") or "").strip()
                if not lbl2: continue
                v = exAct(e.get("actuals",[]))
                data_row_7(row, "  "+lbl2,
                           round(v["ob"],1), round(v["oa"],1), pct_val(v["oa"],v["ob"]),
                           round(v["cb"],1), round(v["ca"],1), pct_val(v["ca"],v["cb"]))
                row += 1
                s_ob+=v["ob"]; s_oa+=v["oa"]; s_cb+=v["cb"]; s_ca+=v["ca"]
            total_row_7(row, s_ob, s_oa, s_cb, s_ca)
            row += 1
            g_ob+=s_ob; g_oa+=s_oa; g_cb+=s_cb; g_ca+=s_ca
        grand_row_7(row, f"Total {title}", g_ob, g_oa, g_cb, g_ca)
        row += 2
        return row

    def write_unit_card(ws_row, start_col, display_lbl,
                        sub_heads_cur, sub_heads_prev, opex_bud, opex_act,
                        prev_opex_bud, prev_opex_act):
        """Write one breakdown card starting at (ws_row, start_col)."""
        off = start_col - 1   # offset: 0 for left (col1), 8 for right (col9)
        c_start = start_col; c_end = start_col + 6
        row = section_hdr_row(ws_row, display_lbl, c_start, c_end)
        row = two_row_hdr(row, "Expense Category", CUR_LBL, PREV_LBL, off=off)
        prev_sh_map = {sh["name"]: sh for sh in (sub_heads_prev or [])}
        for sh in (sub_heads_cur or []):
            sn = sh["name"]; psh = prev_sh_map.get(sn, {"b":0,"a":0,"items":[]})
            g_cb=g_ca=g_pb=g_pa=0.0
            if sn.upper().replace("  "," ") == "PROGRAM EXPENSES":
                for it in sh.get("items",[]): 
                    if it["name"]==GRANTS_NAME: g_cb=it["b"]; g_ca=it["a"]
                for it in psh.get("items",[]):
                    if it["name"]==GRANTS_NAME: g_pb=it["b"]; g_pa=it["a"]
            dcb=sh["b"]-g_cb; dca=sh["a"]-g_ca; dpb=psh["b"]-g_pb; dpa=psh["a"]-g_pa
            if not any([dcb,dca,dpb,dpa]): continue
            data_row_7(row, sn, round(dcb,1), round(dca,1), pct_val(dca,dcb),
                       round(dpb,1), round(dpa,1), pct_val(dpa,dpb), off=off)
            row += 1
            if g_cb or g_ca or g_pb or g_pa:
                data_row_7(row, "  Grants", round(g_cb,1), round(g_ca,1), pct_val(g_ca,g_cb),
                           round(g_pb,1), round(g_pa,1), pct_val(g_pa,g_pb), off=off)
                row += 1
        total_row_7(row, opex_bud, opex_act, prev_opex_bud, prev_opex_act, off=off)
        row += 1
        return row

    def write_breakdown_section(start_row):
        """Fix 4: two unit cards side-by-side per row."""
        start_row = section_title(start_row, "Operating Expenses Breakdown", 1, 15)

        # Aggregate for Total Foundation
        agg_cur, agg_prev = {}, {}
        for lbl in order:
            if cur_map.get(lbl,{}).get("is_sub"): continue
            for sh in cur_map.get(lbl,{}).get("sub_heads",[]):
                if sh["name"] not in agg_cur:
                    agg_cur[sh["name"]] = dict(name=sh["name"],b=0.0,a=0.0,items=[])
                agg_cur[sh["name"]]["b"] += sh["b"]; agg_cur[sh["name"]]["a"] += sh["a"]
                agg_cur[sh["name"]]["items"] = sh["items"]
            for sh in prev_map.get(lbl,{}).get("sub_heads",[]):
                if sh["name"] not in agg_prev:
                    agg_prev[sh["name"]] = dict(name=sh["name"],b=0.0,a=0.0,items=[])
                agg_prev[sh["name"]]["b"] += sh["b"]; agg_prev[sh["name"]]["a"] += sh["a"]

        tf_cb  = to_cr(sum(cm.get("ob",0) for cm in cur_map.values()  if not cm.get("is_sub")))
        tf_ca  = to_cr(sum(cm.get("oa",0) for cm in cur_map.values()  if not cm.get("is_sub")))
        tf_pb  = to_cr(sum(pm.get("ob",0) for pm in prev_map.values() if not pm.get("is_sub")))
        tf_pa  = to_cr(sum(pm.get("oa",0) for pm in prev_map.values() if not pm.get("is_sub")))

        # Build unit list: Total Foundation first, then all units
        cards = [("Total Foundation",
                  list(agg_cur.values()), list(agg_prev.values()),
                  tf_cb, tf_ca, tf_pb, tf_pa)]
        for lbl in order:
            cm = cur_map.get(lbl,{}); pm = prev_map.get(lbl,{})
            cards.append((lbl, cm.get("sub_heads",[]), pm.get("sub_heads",[]),
                          to_cr(cm.get("ob",0)), to_cr(cm.get("oa",0)),
                          to_cr(pm.get("ob",0)), to_cr(pm.get("oa",0))))

        # Write in pairs (left col=1, right col=9)
        row = start_row
        i = 0
        while i < len(cards):
            left  = cards[i]
            right = cards[i+1] if i+1 < len(cards) else None
            # Find max row both cards will reach
            row_l = write_unit_card(row, 1, left[0],  left[1],  left[2],
                                    left[3],  left[4],  left[5],  left[6])
            if right:
                row_r = write_unit_card(row, 9, right[0], right[1], right[2],
                                        right[3], right[4], right[5], right[6])
                row = max(row_l, row_r)
            else:
                row = row_l
            row += 1
            i += 2
        return row

    # ── TITLE ROW (Fix 1: smaller text) ──────────────────────────────────────
    row = 1
    # Full width title covers A to O (cols 1-15)
    for col in range(1, 16):
        ws.cell(row=row, column=col).fill = _fill(C_BLUE)
    ws.merge_cells("A1:O1")
    t = ws.cell(row=1, column=1,
                value=f"MIS – Azim Premji Foundation   |   YTD {month}-{financial_year.split('-')[0]}   |   ₹ Cr.")
    t.font      = _font(bold=True, size=10, color=C_WHITE)   # Fix 1: size 10 (was 13)
    t.alignment = _align("center", "center")
    ws.row_dimensions[1].height = 20
    row = 3   # blank row 2 then start

    # ── WRITE ALL SECTIONS ────────────────────────────────────────────────────
    row = write_main_section(row, "Operating Expense",  "ob", "oa")
    row = write_main_section(row, "Capital Expense",    "cb", "ca")
    row = write_main_section(row, "Overall Foundation", "tb", "ta")
    row = write_breakup_section(row, "Education",
          ["Education - District Institutes","Education- Azim Premji Schools"],"States")
    row = write_breakup_section(row, "Health",
          ["Urban Primary care work","Rural Primary care work",
           "Central Initiatives","Hospital","Health Programs Team & Enablers"],"Areas of Work")
    row = write_breakup_section(row, "Livelihoods", ["Livelihoods"], "States")
    row = write_breakup_section(row, "University",
          ["Azim Premji University (Bangalore Campus)",
           "Azim Premji University (Bhopal Campus)",
           "Azim Premji University (Ranchi Campus)"], "Campus / Unit")
    row = write_breakup_section(row, "Enablers", ["Enablers"], "Functions")
    row = write_breakdown_section(row)   # Fix 4: 2 cards per row

    # ── Freeze & finish ───────────────────────────────────────────────────────
    ws.freeze_panes = "A3"
    ws.row_dimensions[1].height = 20

    # Fix 2: Portrait page setup for PDF export
    from openpyxl.worksheet.page import PageMargins, PrintPageSetup
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)
    ws.print_title_rows = "1:2"   # repeat title on every page

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    excel_bytes = buf.read()

    fname_base = f"MIS_APF_{financial_year}_YTD_{month}"

    if str(export_format).lower() == "pdf":
        # Try soffice (LibreOffice via Frappe's configured path) first
        import subprocess, tempfile, os, shutil

        # Frappe bench installs soffice; find it
        soffice_candidates = [
            shutil.which("soffice"),
            shutil.which("libreoffice"),
            "/usr/bin/soffice",
            "/usr/bin/libreoffice",
            "/usr/local/bin/soffice",
            "/opt/libreoffice/program/soffice",
        ]
        soffice_path = next((p for p in soffice_candidates if p and os.path.isfile(p)), None)

        if soffice_path:
            try:
                with tempfile.TemporaryDirectory() as tmpdir:
                    xlsx_path = os.path.join(tmpdir, "report.xlsx")
                    with open(xlsx_path, "wb") as f:
                        f.write(excel_bytes)
                    subprocess.run(
                        [soffice_path, "--headless", "--convert-to", "pdf",
                         "--outdir", tmpdir, xlsx_path],
                        check=True, capture_output=True, timeout=60
                    )
                    pdf_path = os.path.join(tmpdir, "report.pdf")
                    with open(pdf_path, "rb") as f:
                        pdf_bytes = f.read()
                frappe.response["filename"]    = fname_base + ".pdf"
                frappe.response["filecontent"] = pdf_bytes
                frappe.response["type"]        = "binary"
                frappe.response["doctype"]     = None
                return
            except Exception as e:
                frappe.log_error(f"soffice PDF failed: {e}")

        # Fallback: return Excel with a message
        frappe.msgprint(
            "PDF export requires LibreOffice (soffice) to be installed on the server. "
            "Downloading Excel instead.",
            title="PDF Not Available", indicator="orange"
        )
        frappe.response["filename"]    = fname_base + ".xlsx"
        frappe.response["filecontent"] = excel_bytes
        frappe.response["type"]        = "binary"
        frappe.response["doctype"]     = None
    else:
        frappe.response["filename"]    = fname_base + ".xlsx"
        frappe.response["filecontent"] = excel_bytes
        frappe.response["type"]        = "binary"
        frappe.response["doctype"]     = None